"""
local_server.py
───────────────
Mini servidor Flask + SQLite que reemplaza a SharePoint Lists durante
pruebas presenciales. Expone los endpoints REST que la app HTML
consume — el HTML detecta automáticamente que está corriendo en
localhost y apunta a /api/... en lugar de _api/web/lists/... (SharePoint).

Endpoints
─────────
  GET    /                          → sirve app_revision.html
  GET    /api/casos                 → lista de casos (TblCasos)
  GET    /api/historial             → registros históricos (TblHistorial)
  PATCH  /api/casos/<id_caso>       → actualizar campos del caso
  GET    /api/whoami?email=...      → identificar usuario por email
  GET    /api/supervisores          → lista de nombres del MAESTRO
                                       (para el dropdown de login)
  POST   /api/reset                 → drop + recreate desde Excel

Storage
───────
  ALERTAS/REVISION/revision.sqlite — BD operacional
  ALERTAS/REVISION/Casos_Revision_<MM>_<YYYY>.xlsx — snapshot importable

Modo de uso
───────────
  python local_server.py                  # arranca en :5000
  python local_server.py --port 8000      # otro puerto
  python local_server.py --reset          # vacía la BD y re-importa Excel
  python local_server.py --import <xlsx>  # importa un Excel específico
  python local_server.py --export <xlsx>  # exporta a Excel para inspección
"""

from __future__ import annotations

import os
import sys
import json
import sqlite3
import argparse
import datetime as dt
from pathlib import Path

# Forzar UTF-8 en stdout (Windows cp1252 explota con ═/─/▶)
try:
    sys.stdout.reconfigure(encoding="utf-8")
except AttributeError:
    pass

# SCRIPTS/ ya está en path si se importa via run_revision_local; aseguramos.
_SCRIPTS = Path(__file__).resolve().parent
if str(_SCRIPTS) not in sys.path:
    sys.path.insert(0, str(_SCRIPTS))

import paths


# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURACIÓN
# ─────────────────────────────────────────────────────────────────────────────
DIR_REV    = paths.ALERTAS_DIR / "REVISION"
DB_PATH    = DIR_REV / "revision.sqlite"
APP_HTML   = DIR_REV / "app_revision.html"
MAESTRA    = paths.ALERTAS_MAESTRO

EMAIL_COORDINADOR = os.environ.get("EMAIL_COORDINADOR",
                                    "generico_kenvue@eficacia.com.co").lower()
NOMBRE_COORDINADOR = "VICTOR (Coordinador)"

COLS_CASO = [
    "ID_CASO", "SUPERVISOR", "MERCADERISTA_O_MARCA",
    "MODULO", "TIPO", "NIVEL",
    "CAUSA", "DESCRIPCION", "N_AFECTADOS",
    "VALOR_ORIGINAL", "PDVS_AFECTADOS",
    "MES", "ANIO",
    "ESTADO", "DECISION", "VALOR_CORRECTO", "OBSERVACION",
    "METADATA",
    # Auditoría (escritos por la app o el aplicador)
    "PROPUESTO_POR", "PROPUESTO_EN",
    "APROBADO_POR", "APROBADO_EN",
    "APLICADO_EN",
]

COLS_HISTORIAL = [
    "ID_CORRECCION", "ID_PDV", "NOMBRE_PDV", "MES", "ANIO",
    "MODULO", "CAUSA", "DECISION",
    "VALOR_ORIGINAL", "VALOR_CORRECTO",
    "APROBADO_POR", "FECHA", "ID_CASO_ORIGEN",
]

# Campos que la app puede sobreescribir vía PATCH
CAMPOS_EDITABLES = {
    "ESTADO", "DECISION", "VALOR_CORRECTO", "OBSERVACION",
    "PROPUESTO_POR", "PROPUESTO_EN",
    "APROBADO_POR", "APROBADO_EN",
}


# ─────────────────────────────────────────────────────────────────────────────
# CAPA DE BD
# ─────────────────────────────────────────────────────────────────────────────

def _conn() -> sqlite3.Connection:
    DIR_REV.mkdir(parents=True, exist_ok=True)
    c = sqlite3.connect(str(DB_PATH))
    c.row_factory = sqlite3.Row
    c.execute("PRAGMA foreign_keys = ON")
    c.execute("PRAGMA journal_mode = WAL")  # mejor concurrencia
    return c


def crear_schema(conn: sqlite3.Connection) -> None:
    """Crea las tablas si no existen. Idempotente."""
    casos_sql = "CREATE TABLE IF NOT EXISTS casos (\n" + \
        ",\n".join(f'  "{c}" TEXT' for c in COLS_CASO) + ",\n" + \
        '  PRIMARY KEY ("ID_CASO", "MES", "ANIO")\n)'
    conn.execute(casos_sql)

    hist_sql = "CREATE TABLE IF NOT EXISTS historial (\n" + \
        ",\n".join(f'  "{c}" TEXT' for c in COLS_HISTORIAL) + ",\n" + \
        '  PRIMARY KEY ("ID_CORRECCION")\n)'
    conn.execute(hist_sql)

    conn.execute("CREATE INDEX IF NOT EXISTS idx_supervisor ON casos (SUPERVISOR)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_estado     ON casos (ESTADO)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_periodo    ON casos (MES, ANIO)")
    conn.commit()


def importar_excel(conn: sqlite3.Connection, ruta_xlsx: Path) -> tuple[int, int, int]:
    """
    Upsert por (ID_CASO, MES, ANIO):
      • PENDIENTE existente con mismo contenido → no toca
      • Caso nuevo → INSERT
      • Caso existente NO PENDIENTE → conserva (no sobrescribe la decisión)

    Devuelve (n_insertados, n_conservados_decididos, n_total).
    """
    import pandas as pd

    if not ruta_xlsx.exists():
        raise FileNotFoundError(f"No existe: {ruta_xlsx}")

    df_casos = pd.read_excel(ruta_xlsx, sheet_name="Casos")
    df_hist  = pd.read_excel(ruta_xlsx, sheet_name="Historial")

    crear_schema(conn)

    # ── Casos: upsert respetando estado ─────────────────────────────────
    n_ins = 0
    n_skip = 0
    for _, row in df_casos.iterrows():
        id_caso = str(row.get("ID_CASO", "")).strip()
        mes     = str(row.get("MES", "")).strip()
        anio    = str(row.get("ANIO", "")).strip()
        if not id_caso or not mes or not anio:
            continue

        existente = conn.execute(
            "SELECT ESTADO FROM casos WHERE ID_CASO=? AND MES=? AND ANIO=?",
            (id_caso, mes, anio),
        ).fetchone()

        if existente:
            estado_existente = (existente["ESTADO"] or "").upper()
            if estado_existente in ("PROPUESTO", "APROBADO", "RECHAZADO", "APLICADO"):
                n_skip += 1
                continue  # conservar el estado del supervisor/coordinador

        vals = []
        for c in COLS_CASO:
            v = row.get(c, "")
            if v is None:
                v = ""
            elif hasattr(v, "isoformat"):
                v = v.isoformat()
            else:
                v = str(v) if str(v) != "nan" else ""
            vals.append(v)

        placeholders = ",".join("?" * len(COLS_CASO))
        cols_quoted = ",".join(f'"{c}"' for c in COLS_CASO)
        conn.execute(
            f"INSERT OR REPLACE INTO casos ({cols_quoted}) VALUES ({placeholders})",
            vals,
        )
        n_ins += 1

    # ── Historial: upsert por ID_CORRECCION ─────────────────────────────
    for _, row in df_hist.iterrows():
        id_corr = str(row.get("ID_CORRECCION", "")).strip()
        if not id_corr:
            continue
        vals = []
        for c in COLS_HISTORIAL:
            v = row.get(c, "")
            if v is None:
                v = ""
            elif hasattr(v, "isoformat"):
                v = v.isoformat()
            else:
                v = str(v) if str(v) != "nan" else ""
            vals.append(v)
        placeholders = ",".join("?" * len(COLS_HISTORIAL))
        cols_quoted = ",".join(f'"{c}"' for c in COLS_HISTORIAL)
        conn.execute(
            f"INSERT OR REPLACE INTO historial ({cols_quoted}) VALUES ({placeholders})",
            vals,
        )

    conn.commit()
    n_total = conn.execute("SELECT COUNT(*) FROM casos").fetchone()[0]
    return n_ins, n_skip, n_total


def exportar_excel(conn: sqlite3.Connection, ruta_xlsx: Path) -> Path:
    """Vuelca SQLite → Excel con tablas nombradas (snapshot legible)."""
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.utils import get_column_letter

    df_c = pd.read_sql_query(f"SELECT {','.join(COLS_CASO)} FROM casos ORDER BY ID_CASO", conn)
    df_h = pd.read_sql_query(f"SELECT {','.join(COLS_HISTORIAL)} FROM historial ORDER BY ID_CORRECCION", conn)

    wb = Workbook()
    ws_c = wb.active
    ws_c.title = "Casos"
    ws_c.append(COLS_CASO)
    for _, row in df_c.iterrows():
        ws_c.append([row[c] for c in COLS_CASO])
    if len(df_c) > 0:
        rango = f"A1:{get_column_letter(len(COLS_CASO))}{len(df_c)+1}"
        t = Table(displayName="TblCasos", ref=rango)
        t.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9",
                                           showRowStripes=True)
        ws_c.add_table(t)

    ws_h = wb.create_sheet("Historial")
    ws_h.append(COLS_HISTORIAL)
    for _, row in df_h.iterrows():
        ws_h.append([row[c] for c in COLS_HISTORIAL])
    n_hist_rows = max(len(df_h), 1)
    if len(df_h) == 0:
        ws_h.append([""] * len(COLS_HISTORIAL))
    rango_h = f"A1:{get_column_letter(len(COLS_HISTORIAL))}{n_hist_rows+1}"
    t_h = Table(displayName="TblHistorial", ref=rango_h)
    t_h.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9",
                                          showRowStripes=True)
    ws_h.add_table(t_h)

    wb.save(ruta_xlsx)
    return ruta_xlsx


def reset_db(conn: sqlite3.Connection) -> None:
    """Drop completo de tablas — usado por --reset."""
    conn.execute("DROP TABLE IF EXISTS casos")
    conn.execute("DROP TABLE IF EXISTS historial")
    conn.commit()
    crear_schema(conn)


# ─────────────────────────────────────────────────────────────────────────────
# CAPA DE APLICACIÓN — Flask
# ─────────────────────────────────────────────────────────────────────────────

from flask import Flask, jsonify, request, send_from_directory, abort

app = Flask(__name__)


def _row_to_dict(row: sqlite3.Row) -> dict:
    d = dict(row)
    # METADATA viene serializado como JSON string en la BD
    md = d.get("METADATA", "")
    if md:
        try:
            d["METADATA"] = json.loads(md)
        except (ValueError, TypeError):
            pass
    # Coerciones útiles para el frontend
    try:
        d["N_AFECTADOS"] = int(d.get("N_AFECTADOS") or 0)
    except (ValueError, TypeError):
        d["N_AFECTADOS"] = 0
    try:
        d["MES"] = int(d.get("MES") or 0)
    except (ValueError, TypeError):
        d["MES"] = 0
    try:
        d["ANIO"] = int(d.get("ANIO") or 0)
    except (ValueError, TypeError):
        d["ANIO"] = 0
    return d


@app.route("/")
def root():
    """Sirve la app HTML directamente desde ALERTAS/REVISION/."""
    if not APP_HTML.exists():
        return f"app_revision.html no encontrado en {APP_HTML}", 404
    return send_from_directory(str(DIR_REV), "app_revision.html")


@app.route("/api/casos", methods=["GET"])
def get_casos():
    """
    GET /api/casos?supervisor=NOMBRE&estado=PENDIENTE
    Filtros opcionales por query string. Sin filtros → todos los casos.
    """
    supervisor = request.args.get("supervisor", "").strip().upper()
    estado     = request.args.get("estado", "").strip().upper()
    modulo     = request.args.get("modulo", "").strip().upper()

    sql = "SELECT * FROM casos WHERE 1=1"
    params: list = []
    if supervisor:
        sql += " AND UPPER(SUPERVISOR) = ?"
        params.append(supervisor)
    if estado:
        sql += " AND UPPER(ESTADO) = ?"
        params.append(estado)
    if modulo:
        sql += " AND UPPER(MODULO) = ?"
        params.append(modulo)
    sql += " ORDER BY NIVEL DESC, MODULO, ID_CASO"

    with _conn() as c:
        rows = c.execute(sql, params).fetchall()
    return jsonify([_row_to_dict(r) for r in rows])


@app.route("/api/historial", methods=["GET"])
def get_historial():
    """
    GET /api/historial?id_pdv=123&marca=NEUTROGENA&modulo=SOS
    Filtros opcionales. Sin filtros → todo el historial.
    """
    id_pdv = request.args.get("id_pdv", "").strip()
    modulo = request.args.get("modulo", "").strip().upper()

    sql = "SELECT * FROM historial WHERE 1=1"
    params: list = []
    if id_pdv:
        sql += " AND ID_PDV = ?"
        params.append(id_pdv)
    if modulo:
        sql += " AND UPPER(MODULO) = ?"
        params.append(modulo)
    sql += " ORDER BY FECHA DESC LIMIT 500"

    with _conn() as c:
        rows = c.execute(sql, params).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route("/api/casos/<id_caso>", methods=["PATCH"])
def patch_caso(id_caso: str):
    """
    PATCH /api/casos/<ID_CASO>
    Body JSON: { ESTADO, DECISION, VALOR_CORRECTO, OBSERVACION, ... }
    """
    body = request.get_json(silent=True) or {}
    if not isinstance(body, dict) or not body:
        return jsonify({"error": "Body JSON requerido"}), 400

    # Audit fields automáticos según el estado nuevo
    ahora = dt.datetime.now().isoformat(timespec="seconds")
    usuario = (body.pop("__usuario__", "") or "").strip().upper()

    if "ESTADO" in body and body["ESTADO"] == "PROPUESTO":
        body.setdefault("PROPUESTO_POR", usuario)
        body.setdefault("PROPUESTO_EN",  ahora)
    if "ESTADO" in body and body["ESTADO"] in ("APROBADO", "RECHAZADO"):
        body.setdefault("APROBADO_POR", usuario)
        body.setdefault("APROBADO_EN",  ahora)

    # Sólo campos permitidos
    asign = {k: v for k, v in body.items() if k in CAMPOS_EDITABLES}
    if not asign:
        return jsonify({"error": "Ningún campo editable en el body",
                        "permitidos": sorted(CAMPOS_EDITABLES)}), 400

    set_clause = ", ".join(f'"{k}"=?' for k in asign)
    params = list(asign.values())

    with _conn() as c:
        existe = c.execute("SELECT 1 FROM casos WHERE ID_CASO=?", (id_caso,)).fetchone()
        if not existe:
            return jsonify({"error": f"Caso no encontrado: {id_caso}"}), 404
        c.execute(f"UPDATE casos SET {set_clause} WHERE ID_CASO=?", params + [id_caso])
        c.commit()
        row = c.execute("SELECT * FROM casos WHERE ID_CASO=?", (id_caso,)).fetchone()

    return jsonify(_row_to_dict(row))


@app.route("/api/whoami", methods=["GET"])
def whoami():
    """
    GET /api/whoami?email=...
    Devuelve nombre + rol para login. En modo local sin M365, la app
    pasa el correo seleccionado en el dropdown.
    """
    email = (request.args.get("email", "") or "").strip().lower()
    if not email:
        return jsonify({"email": "", "nombre": "", "esCoordinador": False})

    if email == EMAIL_COORDINADOR:
        return jsonify({
            "email": email,
            "nombre": NOMBRE_COORDINADOR,
            "esCoordinador": True,
        })

    # Buscar en la maestra
    import pandas as pd
    if not MAESTRA.exists():
        return jsonify({"email": email, "nombre": "", "esCoordinador": False,
                        "error": "Maestra no encontrada"})
    df = pd.read_excel(MAESTRA)
    df.columns = df.columns.str.strip().str.upper()
    if "CORREO" not in df.columns or "NOMBRE_SUPERVISOR" not in df.columns:
        return jsonify({"email": email, "nombre": "", "esCoordinador": False,
                        "error": "Maestra sin columnas esperadas"})
    df["CORREO_N"] = df["CORREO"].astype(str).str.strip().str.lower()
    match = df[df["CORREO_N"] == email]
    if match.empty:
        return jsonify({"email": email, "nombre": "",
                        "esCoordinador": False,
                        "error": "Email no registrado"})

    nombre = str(match.iloc[0]["NOMBRE_SUPERVISOR"]).strip().upper()
    return jsonify({"email": email, "nombre": nombre, "esCoordinador": False})


@app.route("/api/supervisores", methods=["GET"])
def get_supervisores():
    """
    GET /api/supervisores
    Devuelve [{nombre, correo, esCoordinador}, ...] para el dropdown
    de login en modo local.
    """
    import pandas as pd
    out = [{
        "nombre": NOMBRE_COORDINADOR,
        "correo": EMAIL_COORDINADOR,
        "esCoordinador": True,
    }]
    if MAESTRA.exists():
        try:
            df = pd.read_excel(MAESTRA)
            df.columns = df.columns.str.strip().str.upper()
            for _, r in df.iterrows():
                nombre = str(r.get("NOMBRE_SUPERVISOR", "")).strip().upper()
                correo = str(r.get("CORREO", "")).strip().lower()
                if nombre and correo and correo != "NAN":
                    out.append({"nombre": nombre, "correo": correo,
                                "esCoordinador": False})
        except Exception as e:
            out.append({"error": str(e)})
    return jsonify(out)


@app.route("/api/reset", methods=["POST"])
def api_reset():
    """POST /api/reset?importar=<ruta_xlsx>  — drop + recreate desde Excel."""
    ruta = request.args.get("importar", "")
    with _conn() as c:
        reset_db(c)
        if ruta:
            ruta_p = Path(ruta)
            if not ruta_p.is_absolute():
                ruta_p = DIR_REV / ruta_p
            ins, skip, total = importar_excel(c, ruta_p)
            return jsonify({"reset": True, "import": str(ruta_p),
                            "insertados": ins, "conservados": skip,
                            "total": total})
    return jsonify({"reset": True})


@app.route("/api/health", methods=["GET"])
def health():
    """Smoke test para el orquestador."""
    with _conn() as c:
        n = c.execute("SELECT COUNT(*) FROM casos").fetchone()[0]
    return jsonify({"ok": True, "db": str(DB_PATH), "casos": n})


# ─────────────────────────────────────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────────────────────────────────────

def _excel_mas_reciente() -> Path | None:
    """Encuentra el Casos_Revision_*.xlsx más reciente."""
    candidatos = sorted(DIR_REV.glob("Casos_Revision_*.xlsx"))
    return candidatos[-1] if candidatos else None


def main():
    parser = argparse.ArgumentParser(description="Servidor local de revisión")
    parser.add_argument("--port",    type=int, default=5000)
    parser.add_argument("--host",    default="127.0.0.1",
                        help="127.0.0.1 (kiosko) o 0.0.0.0 (multi-PC en red local)")
    parser.add_argument("--reset",   action="store_true",
                        help="Borra la BD antes de arrancar (re-importa desde Excel)")
    parser.add_argument("--import",  dest="importar", default=None,
                        help="Importa un Excel específico al arrancar")
    parser.add_argument("--export",  default=None,
                        help="Exporta la BD a un Excel y termina (sin levantar server)")
    parser.add_argument("--no-server", action="store_true",
                        help="Solo importa/exporta y sale (no levanta el server)")
    args = parser.parse_args()

    print("═" * 70)
    print(f"  SERVIDOR LOCAL DE REVISIÓN — Eficacia")
    print(f"  DB    : {DB_PATH}")
    print(f"  HTML  : {APP_HTML}")
    print("═" * 70)

    with _conn() as c:
        crear_schema(c)

        if args.reset:
            print("  ⚠️  --reset: borrando BD")
            reset_db(c)

        # Importar
        ruta_imp = None
        if args.importar:
            ruta_imp = Path(args.importar)
            if not ruta_imp.is_absolute():
                ruta_imp = DIR_REV / ruta_imp
        else:
            ruta_imp = _excel_mas_reciente()

        if ruta_imp and ruta_imp.exists():
            print(f"  Importando: {ruta_imp.name}")
            ins, skip, total = importar_excel(c, ruta_imp)
            print(f"     Insertados/actualizados: {ins}")
            print(f"     Conservados (ya decididos): {skip}")
            print(f"     Total en BD: {total}")
        else:
            print(f"  ⚠️  Sin Excel para importar (esperaba {DIR_REV}/Casos_Revision_*.xlsx)")

        # Exportar
        if args.export:
            ruta_exp = Path(args.export)
            if not ruta_exp.is_absolute():
                ruta_exp = DIR_REV / ruta_exp
            exportar_excel(c, ruta_exp)
            print(f"  Exportado: {ruta_exp}")
            return

    if args.no_server:
        print("  --no-server: termino aquí.")
        return

    print(f"\n  ▶ Servidor escuchando en http://{args.host}:{args.port}")
    print(f"     Abre el navegador en: http://localhost:{args.port}/")
    print(f"     Ctrl+C para detener.\n")
    app.run(host=args.host, port=args.port, debug=False, use_reloader=False)


if __name__ == "__main__":
    main()
