"""
alertas_email.py
────────────────
Fase C del sistema de alertas de Eficacia.

Genera los adjuntos Excel por supervisor y dispara la macro VBA
de EnviarCorreos.xlsm para que Outlook envíe los correos.

Flujo de trabajo
────────────────
1. Lee MAESTRO_SUPERVISORES.xlsx para obtener correos.
2. Lee los adjuntos ya generados por calcular_cumplimientos.py
   (uno por supervisor en ALERTAS/ADJUNTOS/).
3. Escribe la hoja "COLA" de EnviarCorreos.xlsm con:
      CORREO | ASUNTO | CUERPO_HTML | RUTA_ADJUNTO | ENVIADO
4. Abre EnviarCorreos.xlsm con win32com (Excel COM) y ejecuta
   la macro Sub EnviarTodos().
5. Excel+VBA itera la cola y envía cada correo via Outlook Desktop.

Por qué win32com → VBA en lugar de win32com → Outlook directo
──────────────────────────────────────────────────────────────
  • VBA maneja mejor los diálogos de seguridad de Outlook.
  • La macro puede marcar cada fila como ENVIADO antes de avanzar,
    dando trazabilidad sin escribir desde Python mientras COM está activo.
  • Si un envío falla, la macro lo registra en la columna ESTADO
    sin detener los demás.

Prerequisitos
─────────────
  · win32com instalado: pip install pywin32
  · Outlook Desktop abierto y con sesión activa
  · MAESTRO_SUPERVISORES.xlsx con CORREO completo
  · EnviarCorreos.xlsm en la carpeta ALERTAS/
  · Adjuntos generados por calcular_cumplimientos.py
"""

import os
import sys
import time
import pandas as pd
from pathlib import Path
from datetime import datetime

# Cargar config.env antes de leer UMBRAL_OK/WARNING desde os.environ
from config_loader import cargar_config
cargar_config()

from alertas_logger import get_logger
_log = get_logger("alertas_email")

import paths   # alertas_logger ya añadió SCRIPTS/ a sys.path
BASE         = paths.BASE
DIR_ALERTAS  = paths.ALERTAS_DIR
DIR_ADJUNTOS = paths.ALERTAS_ADJUNTOS
RUTA_MAESTRA = paths.ALERTAS_MAESTRO
RUTA_XLSM    = paths.ALERTAS_XLSM

UMBRAL_OK      = float(os.environ.get("UMBRAL_OK",      "0.90"))

# Contacto del analista para reportar inconsistencias (Sprint 13.3)
ANALISTA_NOMBRE = os.environ.get("ANALISTA_NOMBRE", "Giovanny Restrepo")
ANALISTA_EMAIL  = os.environ.get("ANALISTA_EMAIL",  "giovanny_restrepo@eficacia.com.co")
UMBRAL_WARNING = float(os.environ.get("UMBRAL_WARNING",  "0.70"))


# ─────────────────────────────────────────────────────────────────────────────
# CONSTRUCCIÓN DEL CUERPO HTML DEL CORREO
# ─────────────────────────────────────────────────────────────────────────────

def _umbral_ok_de(col: str | None) -> float:
    """Resuelve el UMBRAL_OK específico de la columna (Sprint 13.4)."""
    if not col:
        return UMBRAL_OK
    try:
        from calcular_cumplimientos import umbral_de
        return umbral_de(col)
    except Exception:
        return UMBRAL_OK


def _color_celda(valor, col: str | None = None) -> str:
    """Color de fondo HTML según umbral específico del KPI (Sprint 13.4)."""
    if pd.isna(valor):
        return "#F2F2F2"
    umb_ok = _umbral_ok_de(col)
    if valor >= umb_ok:
        return "#C6EFCE"
    if valor >= UMBRAL_WARNING:
        return "#FFEB9C"
    return "#FFC7CE"


def _semaforo(valor, col: str | None = None) -> str:
    if pd.isna(valor):
        return "—"
    umb_ok = _umbral_ok_de(col)
    if valor >= umb_ok:
        return "✅"
    if valor >= UMBRAL_WARNING:
        return "⚠️"
    return "❌"


def _pct(valor) -> str:
    if pd.isna(valor):
        return "N/D"
    return f"{valor * 100:.1f}%"


def _label_global_rol(rol: str) -> str:
    """Sprint 17.15 — label del cumplimiento global según rol."""
    if rol == "GDD":
        return "Cumplimiento global del periodo"
    if rol == "LIDER":
        return "Cumplimiento global de tu zona"
    return "Cumplimiento global del equipo"


def _intro_rol(rol: str, n_gestores: int) -> str:
    """Sprint 17.15 — saludo según rol del destinatario."""
    if rol == "GDD":
        return "A continuación encontrará su <strong>cumplimiento personal</strong>"
    if rol == "LIDER":
        if n_gestores < 1:
            return "A continuación encontrará su <strong>cumplimiento personal</strong>"
        return (
            f"A continuación encontrará el resumen de cumplimiento "
            f"de los <strong>{n_gestores} supervisores</strong> a su cargo "
            f"y su cumplimiento personal"
        )
    # SUPERVISOR (default)
    if n_gestores == 0:
        return "A continuación encontrará su <strong>cumplimiento personal</strong>"
    # Restamos 1 por la fila del propio supervisor cuando aplica.
    n_mercs = max(n_gestores - 1, 1)
    return (
        f"A continuación encontrará el resumen de cumplimiento de su equipo de "
        f"<strong>{n_mercs} mercaderistas</strong>"
    )


def construir_cuerpo_html(
    supervisor: str,
    df_equipo: pd.DataFrame,
    mes: int,
    anio: int,
    rango_periodo: dict | None = None,
    rol_destinatario: str = "SUPERVISOR",
) -> str:
    """
    Genera el HTML del cuerpo del correo para un supervisor.
    Incluye tabla de resumen del equipo y nota sobre el adjunto.

    rango_periodo (Sprint 13.1): dict de detectar_rango_periodo() con
    fecha_inicio/fecha_fin/rango_legible/avance_esperado_pct. Si None,
    el correo solo muestra mes/año sin rango específico.
    """
    meses_es = {
        1:"Enero", 2:"Febrero", 3:"Marzo", 4:"Abril",
        5:"Mayo", 6:"Junio", 7:"Julio", 8:"Agosto",
        9:"Septiembre", 10:"Octubre", 11:"Noviembre", 12:"Diciembre",
    }
    mes_nombre = meses_es.get(mes, str(mes))
    rango_periodo = rango_periodo or {}
    rango_legible = rango_periodo.get("rango_legible", "")
    avance_esperado = rango_periodo.get("avance_esperado_pct", 0.0)
    dias_trans = rango_periodo.get("dias_transcurridos", 0)
    dias_total = rango_periodo.get("dias_mes_total", 0)

    # Tabla HTML con los mercaderistas del equipo.
    # Sprint 17.12 — la columna "Exh. Pagadas" muestra CAPTURA (no ejecución).
    # Sprint 17.13 — la columna "Exh. Gratis" pasa de conteo (#) a % cumplimiento
    # promedio = (CUMP_ALTO + CUMP_MEDIO) / 2, con targets por canal del gestor
    # (PROXIMITY/TAT → 3 alto + 5 medio; resto → 2 alto + 8 medio).
    cols_modulos = {
        # Operativos
        "CIF_%"                  : "CIF<br>(Visitas)",
        "NP_%"                   : "No<br>Presencia",
        "PRECIOS_%"              : "Precios",
        "SOS_%"                  : "SOS",
        "EXHIB_PAG_CAPTURA_%"    : "Exh.<br>Pagadas",
        # Comerciales D&P
        "VENTA_%"                : "Venta<br>vs Cuota",
        "IMPACTOS_%"             : "Impactos",
        "MSL_%"                  : "MSL",
        "PROD_NUEVOS_%"          : "Productos<br>Nuevos",
        "EXHIB_GRATIS_PROM_%"    : "Exh.<br>Gratis",
        "CUMPL_GLOBAL_%"         : "Global",
    }
    if rol_destinatario == "LIDER":
        for _c in ("VENTA_%", "IMPACTOS_%", "MSL_%", "PROD_NUEVOS_%"):
            cols_modulos.pop(_c, None)

    # Sprint 17.13 — derivar EXHIB_GRATIS_PROM_% = (CUMP_ALTO + CUMP_MEDIO)/2.
    # Sprint 17.14 — cada componente se capa a 1.0 (100%) antes del promedio
    # para evitar que outliers (gestores con conteos anómalos) inflen el global.
    if {"EXHIB_GRA_ALTO_%", "EXHIB_GRA_MEDIO_%"}.issubset(df_equipo.columns):
        df_equipo = df_equipo.copy()
        alto_cap  = df_equipo["EXHIB_GRA_ALTO_%"].clip(upper=1.0)
        medio_cap = df_equipo["EXHIB_GRA_MEDIO_%"].clip(upper=1.0)
        df_equipo["EXHIB_GRATIS_PROM_%"] = pd.concat(
            [alto_cap, medio_cap], axis=1
        ).mean(axis=1, skipna=True)

    # CUMPL_GLOBAL: SIEMPRE recalcular para que use EXHIB_GRATIS_PROM_% capeado
    # (no los componentes ALTO/MEDIO sin techo, ni el EXHIB_PAG_% de ejecución
    # que ya no se reporta en el HTML).
    cols_kpi = ["CIF_%", "NP_%", "PRECIOS_%", "SOS_%", "EXHIB_PAG_CAPTURA_%",
                "EXHIB_GRATIS_PROM_%"]
    if rol_destinatario != "LIDER":
        cols_kpi += ["VENTA_%", "IMPACTOS_%", "MSL_%", "PROD_NUEVOS_%"]
    cols_disp = [c for c in cols_kpi if c in df_equipo.columns]
    if cols_disp:
        df_equipo = df_equipo.copy()
        df_equipo["CUMPL_GLOBAL_%"] = df_equipo[cols_disp].mean(axis=1, skipna=True)

    filas_html = ""
    for i, (_, row) in enumerate(df_equipo.iterrows()):
        bg = "#FFFFFF" if i % 2 == 0 else "#F7F7F7"
        nombre = row.get("NOMBRE", "—")
        filas_html += f'<tr style="background:{bg}">\n'
        filas_html += f'  <td style="padding:6px 10px;border:1px solid #ddd">{nombre}</td>\n'
        for col in cols_modulos:
            if col in row.index:
                val = row[col]
                # Semáforo y color según umbral del KPI (todas las cols ahora son %).
                color = _color_celda(val, col)
                texto = f"{_pct(val)} {_semaforo(val, col)}"
                filas_html += (
                    f'  <td style="padding:6px 10px;border:1px solid #ddd;'
                    f'background:{color};text-align:center">{texto}</td>\n'
                )
        filas_html += "</tr>\n"

    encabezados_html = '<th style="padding:8px 10px;background:#1F3864;color:#FFF;border:1px solid #ddd">Mercaderista</th>\n'
    for etiqueta in cols_modulos.values():
        encabezados_html += (
            f'<th style="padding:8px 10px;background:#1F3864;color:#FFF;'
            f'border:1px solid #ddd;text-align:center">{etiqueta}</th>\n'
        )

    n_gestores = len(df_equipo)
    global_equipo_val = df_equipo["CUMPL_GLOBAL_%"].mean() if "CUMPL_GLOBAL_%" in df_equipo.columns else float("nan")
    global_equipo_str = _pct(global_equipo_val)
    global_sem        = _semaforo(global_equipo_val, "CUMPL_GLOBAL_%")

    html = f"""
<!DOCTYPE html>
<html>
<head><meta charset="UTF-8"></head>
<body style="font-family:Arial,sans-serif;font-size:13px;color:#333;margin:0;padding:0">

  <div style="background:#1F3864;padding:18px 24px">
    <h2 style="margin:0;color:#FFFFFF;font-size:16px">
      📊 Reporte de Cumplimiento — {mes_nombre} {anio}
    </h2>
    {f'<p style="margin:4px 0 0;color:#C5D4E8;font-size:12px">Corte del {rango_legible} — día {dias_trans} de {dias_total} ({int(avance_esperado*100)}% del mes transcurrido)</p>' if rango_legible else '<p style="margin:4px 0 0;color:#C5D4E8;font-size:12px">Eficacia</p>'}
  </div>

  <div style="padding:20px 24px">

    <p>Estimada/o <strong>{supervisor.title()}</strong>,</p>
    <p>{_intro_rol(rol_destinatario, n_gestores)} {f'para el periodo del <strong>{rango_legible} de {anio}</strong>' if rango_legible else f'para el periodo <strong>{mes_nombre} {anio}</strong>'}.</p>

    <p><strong>{_label_global_rol(rol_destinatario)}: {global_equipo_str} {global_sem}</strong></p>

    <table style="border-collapse:collapse;width:100%;margin-top:12px">
      <thead>
        <tr>{encabezados_html}</tr>
      </thead>
      <tbody>
        {filas_html}
      </tbody>
    </table>

    <p style="margin-top:20px;font-size:12px;color:#666">
      El archivo adjunto contiene el detalle completo por punto de venta.<br>
      Ante cualquier inconsistencia, comunícate con <strong>{ANALISTA_NOMBRE}</strong> &mdash;
      <a href="mailto:{ANALISTA_EMAIL}" style="color:#1F3864;text-decoration:underline">{ANALISTA_EMAIL}</a>.
    </p>

  </div>

  <div style="background:#F2F2F2;padding:10px 24px;font-size:11px;color:#888">
    Eficacia S.A. — Reporte generado automáticamente el {datetime.now().strftime("%d/%m/%Y %H:%M")}
  </div>

</body>
</html>
"""
    return html


# ─────────────────────────────────────────────────────────────────────────────
# ESCRITURA DE LA HOJA COLA EN EnviarCorreos.xlsm
# ─────────────────────────────────────────────────────────────────────────────

def escribir_cola_envios(
    df_detalle: pd.DataFrame,
    df_maestro: pd.DataFrame,
    rutas_adjuntos: dict,
    mes: int,
    anio: int,
    rango_periodo: dict | None = None,
) -> int:
    """
    Escribe la hoja COLA del xlsm con una fila por supervisor.
    Devuelve el número de filas escritas.

    Columnas de la hoja COLA:
      CORREO | ASUNTO | CUERPO_HTML | RUTA_ADJUNTO | ESTADO
    """
    from openpyxl import load_workbook

    if not RUTA_XLSM.exists():
        raise FileNotFoundError(
            f"No se encontró EnviarCorreos.xlsm en: {RUTA_XLSM}\n"
            "Asegúrate de que el archivo esté en la carpeta ALERTAS/."
        )

    # Normalizar maestra
    df_m = df_maestro.copy()
    df_m.columns = df_m.columns.str.strip().str.upper()
    df_m["NOMBRE_SUPERVISOR"] = df_m["NOMBRE_SUPERVISOR"].astype(str).str.strip().str.upper()
    df_m["CORREO"]            = df_m["CORREO"].astype(str).str.strip()

    # Construir filas de la cola
    meses_es = {
        1:"Enero", 2:"Febrero", 3:"Marzo", 4:"Abril",
        5:"Mayo", 6:"Junio", 7:"Julio", 8:"Agosto",
        9:"Septiembre", 10:"Octubre", 11:"Noviembre", 12:"Diciembre",
    }
    mes_nombre = meses_es.get(mes, str(mes))
    # Sprint 13.1: el asunto incluye el rango de fechas si está disponible
    if rango_periodo and rango_periodo.get("rango_legible"):
        asunto_base = (
            f"Cumplimiento {mes_nombre} {anio} "
            f"(corte {rango_periodo['rango_legible']}) — Tu equipo Eficacia"
        )
    else:
        asunto_base = f"Cumplimiento {mes_nombre} {anio} — Tu equipo Eficacia"

    filas = []
    sin_correo = []

    # Sprint 17.15 — destinatarios: SUPERVISOR + GDD + LIDER.
    # Cada uno recibe correo propio.
    if {"ES_SUPERVISOR", "ES_GDD", "ES_LIDER"}.issubset(df_detalle.columns):
        sups_df = (
            df_detalle[
                (df_detalle["ES_SUPERVISOR"] == True) |
                (df_detalle["ES_GDD"]        == True) |
                (df_detalle["ES_LIDER"]      == True)
            ][["ACRONIMO", "NOMBRE"]]
                     .drop_duplicates(subset=["ACRONIMO"])
                     .sort_values("NOMBRE")
        )
    elif "ES_SUPERVISOR" in df_detalle.columns:
        sups_df = (
            df_detalle[df_detalle["ES_SUPERVISOR"] == True][["ACRONIMO", "NOMBRE"]]
                     .drop_duplicates(subset=["ACRONIMO"])
                     .sort_values("NOMBRE")
        )
    else:
        sups_df = pd.DataFrame({
            "ACRONIMO": [""] * df_detalle["SUPERVISOR_LIDER"].nunique(),
            "NOMBRE":   sorted(df_detalle["SUPERVISOR_LIDER"].dropna().unique()),
        })

    # Mapping LIDER_POR_CIUDAD y helper para construir el equipo del líder
    # (filtrado a canal DIRECTO + filas sintéticas con promedios del equipo)
    # se importan de calcular_cumplimientos para mantener una sola fuente
    # de verdad entre el adjunto Excel y el cuerpo HTML del correo.
    try:
        from calcular_cumplimientos import (
            LIDER_POR_CIUDAD as _LIDER_POR_CIUDAD,
            construir_equipo_lider as _construir_equipo_lider,
        )
    except ImportError:
        _LIDER_POR_CIUDAD = {}
        _construir_equipo_lider = None

    for _, fila_sup in sups_df.iterrows():
        acr_sup    = fila_sup["ACRONIMO"]
        nombre_sup = str(fila_sup["NOMBRE"]).strip().upper()
        if not nombre_sup:
            continue

        correo_row = df_m[df_m["NOMBRE_SUPERVISOR"] == nombre_sup]
        if correo_row.empty or correo_row["CORREO"].iloc[0] in ("", "nan", "NAN"):
            sin_correo.append(nombre_sup)
            continue
        correo = correo_row["CORREO"].iloc[0]

        # Sprint 17.15 — equipo según rol del destinatario.
        propio = df_detalle[df_detalle["ACRONIMO"] == acr_sup]
        es_gdd   = bool(propio["ES_GDD"].iloc[0])   if not propio.empty and "ES_GDD"   in propio.columns else False
        es_lider = bool(propio["ES_LIDER"].iloc[0]) if not propio.empty and "ES_LIDER" in propio.columns else False
        if es_gdd:
            df_equipo = propio.copy()
        elif es_lider and _construir_equipo_lider is not None:
            df_equipo = _construir_equipo_lider(df_detalle, acr_sup)
        elif "ACRONIMO_SUP" in df_detalle.columns and acr_sup:
            # SUPERVISOR — gestores a cargo + el propio supervisor.
            gestores = df_detalle[df_detalle["ACRONIMO_SUP"] == acr_sup]
            df_equipo = pd.concat([gestores, propio], ignore_index=True)
        else:
            df_equipo = df_detalle[df_detalle.get("SUPERVISOR_LIDER", "") == nombre_sup].copy()

        cuerpo  = construir_cuerpo_html(
            nombre_sup, df_equipo, mes, anio,
            rango_periodo=rango_periodo,
            rol_destinatario=("GDD" if es_gdd else ("LIDER" if es_lider else "SUPERVISOR")),
        )
        adjunto = rutas_adjuntos.get(nombre_sup, "")

        filas.append({
            "CORREO"      : correo,
            "ASUNTO"      : asunto_base,
            "CUERPO_HTML" : cuerpo,
            "RUTA_ADJUNTO": adjunto,
            "ESTADO"      : "PENDIENTE",
        })

    if sin_correo:
        print(f"  ⚠️  {len(sin_correo)} supervisor(es) sin correo en la maestra:")
        for s in sin_correo:
            print(f"     · {s}")

    # Abrir xlsm y escribir hoja COLA (o crearla si no existe)
    wb = load_workbook(RUTA_XLSM, keep_vba=True)

    if "COLA" in wb.sheetnames:
        ws_cola = wb["COLA"]
        # Limpiar filas anteriores (conservar encabezado)
        for row in ws_cola.iter_rows(min_row=2):
            for cell in row:
                cell.value = None
    else:
        ws_cola = wb.create_sheet("COLA")

    # Encabezados
    headers = ["CORREO", "ASUNTO", "CUERPO_HTML", "RUTA_ADJUNTO", "ESTADO"]
    for col_idx, hdr in enumerate(headers, start=1):
        ws_cola.cell(row=1, column=col_idx, value=hdr)

    # Datos
    for row_idx, fila in enumerate(filas, start=2):
        for col_idx, hdr in enumerate(headers, start=1):
            ws_cola.cell(row=row_idx, column=col_idx, value=fila[hdr])

    wb.save(RUTA_XLSM)
    print(f"  ✓ Hoja COLA escrita: {len(filas)} filas en {RUTA_XLSM.name}")
    return len(filas)


# ─────────────────────────────────────────────────────────────────────────────
# DISPARO DE LA MACRO VBA via win32com
# ─────────────────────────────────────────────────────────────────────────────

def _emitir_log_macro_y_estados() -> dict:
    """
    Tras ejecutar la macro, lee:
      1. ALERTAS/logs/macro_envio.log  (escrito por el VBA — INFO/WARN/ERROR).
         De ahi extraemos el resumen enviados/errores (fuente autoritativa).
      2. La columna ESTADO de la hoja COLA (best-effort) para detalles de error.
         Reintenta si el archivo aun esta locked por Excel/OneDrive.
    Emite al logger Python y devuelve un dict con resumen.
    """
    enviados = errores = total = 0
    log_path = paths.ALERTAS_DIR / "logs" / "macro_envio.log"
    if log_path.exists():
        try:
            with open(log_path, "r", encoding="utf-8", errors="replace") as f:
                lineas = [ln.rstrip() for ln in f if ln.strip()]
            for ln in lineas:
                nivel = ln.split("|", 2)[1].strip().upper() if "|" in ln else "INFO"
                msg = f"[macro] {ln}"
                if nivel == "ERROR":   _log.error(msg)
                elif nivel == "WARN":  _log.warning(msg)
                else:                  _log.info(msg)
                # Parse resumen final: "Resumen -- enviados=X errores=Y total=Z"
                if "Resumen" in ln and "enviados=" in ln:
                    import re
                    m_e = re.search(r"enviados=(\d+)", ln)
                    m_x = re.search(r"errores=(\d+)",  ln)
                    m_t = re.search(r"total=(\d+)",    ln)
                    if m_e: enviados = int(m_e.group(1))
                    if m_x: errores  = int(m_x.group(1))
                    if m_t: total    = int(m_t.group(1))
            log_path.unlink()
        except Exception as e:
            _log.warning(f"No se pudo leer {log_path.name}: {e}")
    else:
        _log.warning("No se encontró logs/macro_envio.log — la macro no produjo log.")

    # Detalles de error por fila desde la hoja COLA (best-effort, con reintentos).
    detalles_error: list[str] = []
    try:
        from openpyxl import load_workbook
        wb = None
        ultimo_error = None
        for intento in range(5):
            try:
                wb = load_workbook(RUTA_XLSM, read_only=True, keep_vba=True)
                break
            except PermissionError as pe:
                ultimo_error = pe
                time.sleep(0.5 * (intento + 1))
        if wb is None:
            _log.warning(f"No se pudo abrir COLA para detalles de error: {ultimo_error}")
        else:
            if "COLA" in wb.sheetnames:
                ws = wb["COLA"]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row or row[0] is None:
                        continue
                    estado = str(row[4] or "").strip().upper()
                    if estado.startswith("ERROR"):
                        detalles_error.append(f"{row[0]} — {row[4]}")
            wb.close()
    except Exception as e:
        _log.warning(f"No se pudo leer hoja COLA para detalles: {e}")

    _log.info(f"Email — enviados={enviados} errores={errores} total={total}")
    for d in detalles_error:
        _log.error(f"[cola] {d}")
    return {"enviados": enviados, "errores": errores, "total": total}


def disparar_macro(modo_prueba: bool = False) -> bool:
    """
    Abre EnviarCorreos.xlsm con Excel COM y ejecuta Sub EnviarTodos().
    La macro NO muestra MsgBox: escribe resumen y errores a
    ALERTAS/logs/macro_envio.log, que Python lee y emite al logger
    tras la ejecución.
    Devuelve True si la macro terminó sin excepción.
    """
    if modo_prueba:
        print("  ⚠️  MODO PRUEBA — la macro no se ejecutará")
        print(f"  El archivo EnviarCorreos.xlsm está listo en: {RUTA_XLSM}")
        return True

    try:
        import win32com.client as win32
    except ImportError:
        print("  ❌ win32com no está instalado.")
        print("     Instala con: pip install pywin32")
        print(f"     Alternativamente, abre {RUTA_XLSM} y ejecuta la macro manualmente.")
        return False

    print("  Abriendo Excel y ejecutando macro EnviarTodos()...")
    xl   = None
    wb   = None
    ok   = False

    # Defensive: matar cualquier EXCEL.EXE residual antes de Dispatch.
    # Dispatch puede reutilizar un proceso existente, y si ese proceso tiene
    # un dialogo modal o un workbook bloqueado, la llamada COM falla con
    # RPC_E_CALL_REJECTED (-2147418111).
    try:
        import subprocess
        subprocess.run(["taskkill", "/F", "/IM", "EXCEL.EXE"],
                       capture_output=True, timeout=10)
        time.sleep(1)
    except Exception:
        pass

    try:
        xl = win32.Dispatch("Excel.Application")
        # Excel invisible: la macro ya no muestra MsgBox.
        xl.Visible = False
        xl.DisplayAlerts = False
        try:
            xl.AutomationSecurity = 1   # msoAutomationSecurityLow
        except Exception:
            pass

        wb = xl.Workbooks.Open(str(RUTA_XLSM.resolve()))
        time.sleep(2)               # esperar a que el VBA cargue

        # El .xlsm puede estar en OneDrive/SharePoint sincronizado, en cuyo caso
        # ThisWorkbook.Path en VBA devuelve la URL https y Open # falla (error 52).
        # Le pasamos el path local absoluto del log en Hoja1!Y1 para que la
        # macro escriba ahi.
        log_path = paths.ALERTAS_DIR / "logs" / "macro_envio.log"
        log_path.parent.mkdir(parents=True, exist_ok=True)
        try:
            wb.Sheets("Hoja1").Range("Y1").Value = str(log_path.resolve())
        except Exception:
            pass

        # Run con retry: Excel puede rechazar la llamada COM si esta
        # procesando algo (RPC_E_CALL_REJECTED = -2147418111).
        ultimo_error = None
        for intento in range(5):
            try:
                xl.Run("EnviarCorreos.xlsm!EnviarTodos")
                ultimo_error = None
                break
            except Exception as e:
                ultimo_error = e
                code = getattr(e, "args", [None])[0]
                if code == -2147418111:
                    print(f"     Reintento {intento+1}/5 tras RPC_E_CALL_REJECTED...")
                    time.sleep(2 * (intento + 1))
                    continue
                raise
        if ultimo_error is not None:
            raise ultimo_error

        # La macro VBA ya hace `ThisWorkbook.Save` al final, asi que el
        # workbook esta persistido. Este Save Python es redundante y a
        # veces falla con RPC_E_CALL_REJECTED porque OneDrive esta haciendo
        # un sync en background — no es fatal.
        try:
            wb.Save()
        except Exception as e_save:
            print(f"     (wb.Save Python falló, no es fatal — VBA ya guardó: {e_save})")
        print("  ✅ Macro ejecutada correctamente")
        ok = True
        return True

    except Exception as e:
        print(f"  ❌ Error ejecutando la macro: {e}")
        print(f"     Puedes abrir {RUTA_XLSM} y ejecutar Sub EnviarTodos() manualmente.")
        return False

    finally:
        # Cleanup: SaveChanges=False (ya guardamos arriba con wb.Save()).
        # Sin MsgBox modales, Excel cierra limpio.
        if wb is not None:
            try:
                wb.Close(SaveChanges=False)
            except Exception:
                pass
        if xl is not None:
            try:
                xl.Quit()
            except Exception:
                pass
        # Emitir el log de la macro y el resumen de la hoja COLA tras cerrar
        # Excel (para evitar contención de archivo).
        if ok:
            try:
                _emitir_log_macro_y_estados()
            except Exception as e:
                _log.warning(f"No se pudo emitir resumen post-macro: {e}")
        # Liberar referencias COM explícitamente para forzar el GC
        wb = None
        xl = None


# ─────────────────────────────────────────────────────────────────────────────
# FUNCIÓN PRINCIPAL EXPORTADA
# ─────────────────────────────────────────────────────────────────────────────

def enviar_correos(
    df_detalle: pd.DataFrame,
    df_maestro: pd.DataFrame,
    rutas_adjuntos: dict,
    mes: int,
    anio: int,
    modo_prueba: bool = False,
    rango_periodo: dict | None = None,
) -> dict:
    """
    Orquesta la generación de la cola y el disparo de la macro.

    Parámetros
    ──────────
    df_detalle      : de calcular_cumplimientos.main()['df_detalle']
    df_maestro      : leído de MAESTRO_SUPERVISORES.xlsx
    rutas_adjuntos  : de calcular_cumplimientos.main()['rutas_adjuntos']
    mes, anio       : periodo
    modo_prueba     : si True, no abre Excel ni envía nada

    Retorna dict con claves 'filas_cola', 'macro_ok'.
    """
    print("\n" + "═" * 55)
    print("  ALERTAS EMAIL — Eficacia")
    print("═" * 55)

    if modo_prueba:
        print("  ⚠️  MODO PRUEBA — no se abrirá Excel ni se enviarán correos\n")

    print("\nEscribiendo cola de envíos:")
    try:
        n_filas = escribir_cola_envios(df_detalle, df_maestro, rutas_adjuntos, mes, anio, rango_periodo=rango_periodo)
    except FileNotFoundError as e:
        print(f"  ❌ {e}")
        return {"filas_cola": 0, "macro_ok": False}

    if n_filas == 0:
        print("  ⚠️  Cola vacía — no hay supervisores con correo configurado.")
        return {"filas_cola": 0, "macro_ok": False}

    print("\nDisparo de macro Outlook:")
    macro_ok = disparar_macro(modo_prueba=modo_prueba)

    _log.info(
        f"Email — filas_cola={n_filas} macro_ok={macro_ok} "
        f"prueba={modo_prueba}"
    )

    print("\n" + "═" * 55)
    print(f"  Correos en cola: {n_filas}")
    print(f"  Macro ejecutada: {'✅' if macro_ok else '❌'}")
    print("═" * 55 + "\n")

    return {"filas_cola": n_filas, "macro_ok": macro_ok}


# ─────────────────────────────────────────────────────────────────────────────
# EJECUCIÓN STANDALONE
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Envío de correos Outlook — Eficacia")
    parser.add_argument("--prueba", action="store_true",
                        help="Modo prueba: genera cola sin abrir Excel")
    args = parser.parse_args()

    # Sprint 17.11 — invocar calcular_cumplimientos.main() para obtener TODO
    # en memoria (incluye rango_periodo con fechas legibles). El standalone
    # antes leía DETALLE_*.xlsx y NO podía propagar rango_periodo, así que
    # el cuerpo del correo decía "para el periodo Mayo 2026" en vez de
    # "para el periodo del 1 al 31 de mayo de 2026".
    import calcular_cumplimientos as cc
    out = cc.main()
    df_detalle     = out["df_detalle"]
    mes            = out["mes"]
    anio           = out["anio"]
    rutas_adj      = out["rutas_adjuntos"]
    rango_periodo  = out.get("rango_periodo")

    df_maestro = pd.read_excel(RUTA_MAESTRA) if RUTA_MAESTRA.exists() else pd.DataFrame()
    if df_maestro.empty:
        print("❌ MAESTRO_SUPERVISORES.xlsx no encontrado.")
        sys.exit(1)

    enviar_correos(
        df_detalle, df_maestro, rutas_adj, mes, anio,
        modo_prueba=args.prueba,
        rango_periodo=rango_periodo,
    )
