import os
import sys
import io
import re
import tempfile
import unicodedata
import requests
import urllib.parse
import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from azure.identity import ClientSecretCredential
from dotenv import load_dotenv

# Cargar variables del .env
load_dotenv()

# Cargar config.env antes de leer UMBRAL_OK/WARNING desde os.environ
from config_loader import cargar_config
cargar_config()

from alertas_logger import get_logger
_log = get_logger("calcular_cumplimientos")

# alertas_logger ya añade SCRIPTS/ a sys.path
import paths

# KPIs comerciales D&P (venta, impactos, MSL, productos nuevos)
import cumplimiento_dyp

# Tabla maestra de personas (ACRONIMO como llave canónica)
import base_cupos as bcm


# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURACIÓN DE CONEXIÓN NUBE (Microsoft Graph API usando Drive ID)
# ─────────────────────────────────────────────────────────────────────────────
AZURE_TENANT_ID = os.getenv("AZURE_TENANT_ID")
AZURE_CLIENT_ID = os.getenv("AZURE_CLIENT_ID")
AZURE_CLIENT_SECRET = os.getenv("AZURE_CLIENT_SECRET")

def _obtener_token_azure():
    """Genera el token de acceso para la API de Microsoft Graph."""
    credential = ClientSecretCredential(
        tenant_id=AZURE_TENANT_ID,
        client_id=AZURE_CLIENT_ID,
        client_secret=AZURE_CLIENT_SECRET
    )
    return credential.get_token("https://graph.microsoft.com/.default").token

_DRIVE_ID_CACHE: dict = {}


def _obtener_default_drive_id(token: str) -> str:
    """
    Obtiene el ID del drive conectándose al sitio corporativo específico de
    SharePoint. Usa `paths.SHAREPOINT_SITE_NAME` (ej. "JJ451") para buscar el
    sitio correcto — antes esta función buscaba un texto fijo ("eficacia")
    que no necesariamente coincide con el nombre real del sitio y podía
    resolver el drive equivocado (causa típica de 404 en TODAS las rutas
    candidatas, incluso siendo la ruta relativa correcta).
    """
    if "drive_id" in _DRIVE_ID_CACHE:
        return _DRIVE_ID_CACHE["drive_id"]

    headers = {"Authorization": f"Bearer {token}"}
    nombre_sitio = getattr(paths, "SHAREPOINT_SITE_NAME", "eficacia")

    # Búsqueda del sitio corporativo por el nombre configurado en paths.py
    res_site = requests.get(
        f"https://graph.microsoft.com/v1.0/sites?search={urllib.parse.quote(nombre_sitio)}",
        headers=headers,
    )
    if res_site.status_code == 200:
        sites = res_site.json().get("value", [])
        if sites:
            # Si hay varios resultados, preferir un match exacto de nombre.
            site = next(
                (s for s in sites
                 if nombre_sitio.lower() in (s.get("name", "").lower(), s.get("displayName", "").lower())),
                sites[0],
            )
            site_id = site.get("id")
            print(f"    ℹ️  Sitio SharePoint resuelto: {site.get('displayName') or site.get('name')} "
                  f"({site.get('webUrl')})")
            res_drive = requests.get(f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive", headers=headers)
            if res_drive.status_code == 200:
                drive = res_drive.json()
                print(f"    ℹ️  Drive resuelto: {drive.get('name')} ({drive.get('webUrl')})")
                _DRIVE_ID_CACHE["drive_id"] = drive.get("id")
                return drive.get("id")

    # Plan B: Si no lo encuentra por búsqueda, intenta con el raíz del tenant
    res_root = requests.get("https://graph.microsoft.com/v1.0/sites/root", headers=headers)
    if res_root.status_code == 200:
        site_id = res_root.json().get("id")
        res_drive = requests.get(f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive", headers=headers)
        if res_drive.status_code == 200:
            print("    ⚠️  No se encontró el sitio por nombre; usando el sitio raíz del tenant como fallback.")
            _DRIVE_ID_CACHE["drive_id"] = res_drive.json().get("id")
            return res_drive.json().get("id")

    raise Exception(f"No se pudo obtener el Drive ID del sitio SharePoint '{nombre_sitio}': {res_site.text}")

def _limpiar_ruta_graph(ruta_sharepoint: str) -> str:
    """
    Normaliza separadores de la ruta relativa dentro del drive.
    IMPORTANTE: NO remueve 'Equipo Información/' — esa es una carpeta REAL
    dentro de la biblioteca 'Documentos compartidos' del sitio (así lo
    define paths.SHAREPOINT_BASE_DIR), no un alias del nombre del sitio.
    Solo se remueven prefijos que sí son alias genéricos del nombre de la
    biblioteca documental por defecto.
    """
    ruta_sharepoint = ruta_sharepoint.replace("\\", "/")
    for prefijo in [
        "Documentos compartidos/",
        "Shared Documents/",
    ]:
        if ruta_sharepoint.startswith(prefijo):
            ruta_sharepoint = ruta_sharepoint.replace(prefijo, "", 1)
    return ruta_sharepoint.lstrip("/")

def _leer_excel_cloud(ruta_sharepoint: str, descripcion: str) -> pd.DataFrame:
    """Lee un archivo Excel directamente desde SharePoint usando el Drive ID en Graph API con reintentos de ruta."""
    print(f"  ⏳ Leyendo {descripcion} desde SharePoint ({ruta_sharepoint})...")
    token = _obtener_token_azure()
    headers = {"Authorization": f"Bearer {token}"}
    
    drive_id = _obtener_default_drive_id(token)
    ruta_limpia = _limpiar_ruta_graph(ruta_sharepoint)
    
    rutas_candidatas = [
        ruta_limpia,
        f"Documentos compartidos/{ruta_limpia}",
    ]
    
    response = None
    url_usada = ""
    for r in rutas_candidatas:
        ruta_codificada = urllib.parse.quote(r)
        url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/root:/{ruta_codificada}:/content"
        print(f"    🔍 Probando URL Graph API: {url}")
        response = requests.get(url, headers=headers)
        if response.status_code == 200:
            url_usada = url
            break

    if response is not None and response.status_code == 200:
        print(f"    ✓ Archivo descargado exitosamente.")
        return pd.read_excel(io.BytesIO(response.content))
    else:
        status = response.status_code if response is not None else "N/A"
        text = response.text if response is not None else "Sin respuesta"
        raise Exception(f"Error al descargar {descripcion} desde Graph API (404/Error). Última URL probada: {url_usada} | Status: {status} - {text}")

def _guardar_excel_cloud(df: pd.DataFrame, ruta_sharepoint: str, descripcion: str):
    """Sube un DataFrame convertido a Excel directamente a SharePoint vía Graph API usando el Drive ID."""
    print(f"  ⏳ Guardando {descripcion} en SharePoint ({ruta_sharepoint})...")
    token = _obtener_token_azure()
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    }
    
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)
    excel_bytes = output.getvalue()
    
    drive_id = _obtener_default_drive_id(token)
    ruta_limpia = _limpiar_ruta_graph(ruta_sharepoint)
    ruta_codificada = urllib.parse.quote(ruta_limpia)
    
    url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/root:/{ruta_codificada}:/content"
    response = requests.put(url, headers=headers, data=excel_bytes)
    if response.status_code in [200, 201]:
        print(f"  ✓ {descripcion} guardado exitosamente en SharePoint")
    else:
        raise Exception(f"Error al subir {descripcion} a Graph API: {response.status_code} - {response.text}")


def _subir_bytes_cloud(data: bytes, ruta_sharepoint: str, descripcion: str,
                       content_type: str = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"):
    """
    Sube bytes ya generados (p.ej. un Workbook de openpyxl volcado a BytesIO)
    directamente a SharePoint vía Graph API. Equivalente a
    `_guardar_excel_cloud` pero sin pasar por un DataFrame, para archivos con
    formato/multiples hojas.
    """
    token = _obtener_token_azure()
    headers = {"Authorization": f"Bearer {token}", "Content-Type": content_type}
    drive_id = _obtener_default_drive_id(token)
    ruta_codificada = urllib.parse.quote(_limpiar_ruta_graph(ruta_sharepoint))
    url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/root:/{ruta_codificada}:/content"
    response = requests.put(url, headers=headers, data=data)
    if response.status_code in (200, 201):
        return True
    raise Exception(f"Error al subir {descripcion} a Graph API: {response.status_code} - {response.text}")


def _descargar_a_temp(ruta_sharepoint: str, descripcion: str) -> Path | None:
    """
    Baja un archivo de SharePoint a un temporal local y devuelve su Path.

    Necesario para los insumos D&P (Consolidado de Impactos, Consolidado de
    Ventas, Rutero y Listas), que se movieron a la nube pero siguen siendo
    consumidos por módulos (`cumplimiento_dyp`, `etl_impactos_segmentos`) que
    esperan un archivo en disco. Así se migra la ubicación sin reescribir esos
    módulos. Devuelve None si no se pudo descargar (el llamador decide qué
    hacer, en vez de fallar silenciosamente).
    """
    try:
        token = _obtener_token_azure()
        headers = {"Authorization": f"Bearer {token}"}
        drive_id = _obtener_default_drive_id(token)
        ruta_codificada = urllib.parse.quote(_limpiar_ruta_graph(ruta_sharepoint))
        url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/root:/{ruta_codificada}:/content"
        response = requests.get(url, headers=headers)
        if response.status_code != 200:
            print(f"    ⚠️  [{descripcion}] no disponible en la nube "
                  f"({response.status_code}): {ruta_sharepoint}")
            _log.warning("No pude descargar %s (%s): %s",
                         descripcion, response.status_code, ruta_sharepoint)
            return None
        destino = Path(tempfile.gettempdir()) / Path(ruta_sharepoint).name
        destino.write_bytes(response.content)
        print(f"    ✓ [{descripcion}] descargado ({len(response.content):,} bytes)")
        return destino
    except Exception as e:
        print(f"    ⚠️  [{descripcion}] error al descargar: {e}")
        _log.warning("Error descargando %s: %s", descripcion, e)
        return None


def _listar_hijos_cloud(ruta_carpeta: str) -> list:
    """
    Lista los archivos/subcarpetas de `ruta_carpeta` en SharePoint vía Graph
    API (equivalente en la nube de recorrer un directorio local). Cada
    elemento devuelto es el dict nativo de Graph (trae 'name',
    'lastModifiedDateTime' y 'file' o 'folder' según el tipo de ítem).
    """
    token = _obtener_token_azure()
    headers = {"Authorization": f"Bearer {token}"}
    drive_id = _obtener_default_drive_id(token)
    ruta_limpia = _limpiar_ruta_graph(ruta_carpeta)
    ruta_codificada = urllib.parse.quote(ruta_limpia)

    url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/root:/{ruta_codificada}:/children"
    items = []
    while url:
        response = requests.get(url, headers=headers)
        if response.status_code != 200:
            raise Exception(
                f"Error al listar carpeta '{ruta_carpeta}' en Graph API: "
                f"{response.status_code} - {response.text}"
            )
        data = response.json()
        items.extend(data.get("value", []))
        url = data.get("@odata.nextLink")
    return items


def _carpeta_periodo(raiz: str, anio: int | None) -> str:
    """
    Devuelve la carpeta de BASES correspondiente al AÑO DEL PERIODO procesado.

    En `paths.py` las carpetas de BASES se construyen con
    `AÑO_ACTUAL = datetime.datetime.now().year`, es decir el año del RELOJ, no
    el del periodo que se está calculando. Eso funciona mientras se procese el
    mes en curso, pero falla en dos casos reales y silenciosos:

      • Cerrar diciembre durante enero → busca en .../2027 cuando el dato
        está en .../2026, y todas las hojas que dependen de BASES salen vacías.
      • Recalcular un periodo anterior (backfill) → misma historia.

    Esta función normaliza: si la raíz ya termina en un año, lo REEMPLAZA por
    el del periodo (no lo anida, que produciría .../2026/2026); si no lo trae,
    lo agrega. Con `anio=None` devuelve la raíz sin tocar.
    """
    raiz = str(raiz).rstrip("/")
    if not anio:
        return raiz
    anio_str = str(int(anio))
    if raiz.endswith(f"/{anio_str}"):
        return raiz
    if re.search(r"/(19|20)\d{2}$", raiz):
        nueva = re.sub(r"/(19|20)\d{2}$", f"/{anio_str}", raiz)
        if nueva != raiz:
            print(f"    ℹ️  Carpeta ajustada al año del periodo: {nueva}")
        return nueva
    return f"{raiz}/{anio_str}"


def _resolver_exhib_data_dir_cloud(anio: int | None = None) -> str:
    """
    Equivalente en la nube de `paths._resolver_exhib_data_dir()`: la carpeta
    de BASES de Exhibiciones suele contener subcarpetas tipo "01. ...",
    "02. ..." (una por corte/actualización); se toma la más reciente. Si no
    hay subcarpetas que matcheen el patrón, se usa la carpeta raíz tal cual.

    FIX: los archivos de planning viven en
    `BASES DE RESPUESTAS/EXHIBICIONES/<año>`. Antes se usaba
    `paths.RUTA_CARPETA_BASES_EXHIB` tal cual, que trae el año del RELOJ
    (`AÑO_ACTUAL`), no el del periodo. Ahora se normaliza con
    `_carpeta_periodo()`.
    """
    raiz = _carpeta_periodo(paths.RUTA_CARPETA_BASES_EXHIB, anio)

    try:
        hijos = _listar_hijos_cloud(raiz)
    except Exception:
        return raiz
    candidatos = [
        h for h in hijos
        if h.get("folder") and re.match(r"^\d{2}\.\s", h.get("name", ""))
    ]
    if not candidatos:
        return raiz
    candidatos.sort(key=lambda h: h.get("lastModifiedDateTime", ""), reverse=True)
    return f"{raiz}/{candidatos[0]['name']}"
    candidatos = [
        h for h in hijos
        if h.get("folder") and re.match(r"^\d{2}\.\s", h.get("name", ""))
    ]
    if not candidatos:
        return raiz
    candidatos.sort(key=lambda h: h.get("lastModifiedDateTime", ""), reverse=True)
    return f"{raiz}/{candidatos[0]['name']}"


def _buscar_archivo_cloud(ruta_carpeta: str, patron_regex: str, contexto: str) -> str:
    """
    Localiza UN único archivo cuyo nombre matchee `patron_regex` dentro de
    `ruta_carpeta` en SharePoint. Equivalente en la nube de
    `periodo_resolver._find_unico`. Ignora temporales de Excel (~$*).
    Devuelve la ruta SharePoint completa del archivo encontrado.
    """
    hijos = _listar_hijos_cloud(ruta_carpeta)
    matches = [
        h["name"] for h in hijos
        if h.get("file") and re.match(patron_regex, h.get("name", ""))
        and not h["name"].startswith("~$")
    ]
    if not matches:
        raise FileNotFoundError(
            f"[{contexto}] No se encontró archivo con patrón '{patron_regex}' en {ruta_carpeta}"
        )
    if len(matches) > 1:
        raise RuntimeError(
            f"[{contexto}] Múltiples coincidencias en {ruta_carpeta}: {matches}. Esperaba exactamente uno."
        )
    return f"{ruta_carpeta}/{matches[0]}"


# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURACIÓN DE RUTAS USANDO EXCLUSIVAMENTE `paths.py`
# ─────────────────────────────────────────────────────────────────────────────
DIR_NP_OUT      = paths.RUTA_CARPETA_SALIDAS_NP
DIR_PRECIOS_OUT = paths.RUTA_CARPETA_SALIDAS_PRECIOS
DIR_ALERTAS     = getattr(paths, 'ALERTAS_DIR', f"{paths._SALIDAS_ROOT}/ALERTAS")
RUTA_MAESTRA    = getattr(paths, 'ALERTAS_MAESTRO', f"{DIR_ALERTAS}/maestra_supervisores.xlsx")


# ── Archivos de DETALLE (los que alimentan las hojas de los adjuntos) ────────
#
# FIX: antes estos cuatro se leían con un nombre SIN periodo
# ("Plan de trabajo.xlsx", "Cumplimiento_Captura_SOS.xlsx",
# "REPORTE_NO_PRESENCIA.xlsx", "REPORTE_CAPTURA_PRECIOS.xlsx"). Ningún ETL
# escribe esos nombres: eran copias que alguien mantenía a mano y que dejaron
# de actualizarse el 31/07/2026. Como los KPIs sí salían del periodo correcto,
# el síntoma era silencioso — al filtrar el detalle a (08, 2026) quedaba vacío
# y los supervisores recibían "N/D" en todo el bloque operativo.
#
# Los ETLs publican SOS/NP/Precios con el periodo en el nombre
# (`..._AGOSTO_2026.xlsx`). CIF es la excepción: `CIF.xlsx` acumula todos los
# meses en un solo archivo y se filtra por MES/AÑO más abajo.
def _periodo_nom(mes: int, anio: int) -> str:
    """'AGOSTO_2026' — la convención de nombre que usan los ETLs."""
    import periodo_resolver as _pr
    spec = _pr.resolver(int(mes), int(anio))
    return f"{spec.mes_str_upper}_{spec.anio}"


def ruta_cif_detalle() -> str:
    return f"{paths.RUTA_CARPETA_SALIDAS_CIF}/{paths.CIF_OUT_CIF.name}"


def ruta_sos_detalle(mes: int, anio: int) -> str:
    return (f"{paths.RUTA_CARPETA_SALIDAS_SOS}/"
            f"Cumplimiento_Captura_SOS_{_periodo_nom(mes, anio)}.xlsx")


def ruta_np_detalle(mes: int, anio: int) -> str:
    return f"{DIR_NP_OUT}/REPORTE_NO_PRESENCIA_{_periodo_nom(mes, anio)}.xlsx"


def ruta_pr_detalle(mes: int, anio: int) -> str:
    return f"{DIR_PRECIOS_OUT}/REPORTE_CAPTURA_PRECIOS_{_periodo_nom(mes, anio)}.xlsx"

# paths.py sólo define el directorio local de ALERTAS (ALERTAS_DIR, usado para
# cachear adjuntos antes de enviarlos por correo). No existe un
# `RUTA_CARPETA_SALIDAS_ALERTAS` en la nube, así que lo construimos aquí
# siguiendo la misma convención que el resto de módulos (_SALIDAS_ROOT/<módulo>).
RUTA_CARPETA_SALIDAS_ALERTAS = f"{paths._SALIDAS_ROOT}/ALERTAS"
RUTA_MAESTRA_CLOUD           = f"{paths._BASES_ROOT}/ALERTAS/MAESTRO_SUPERVISORES.xlsx"
# Carpeta en la nube donde deben quedar los adjuntos por supervisor
# (…/BI/INVOLVES/SALIDAS/ALERTAS/ADJUNTOS)
RUTA_CARPETA_ADJUNTOS_CLOUD  = f"{RUTA_CARPETA_SALIDAS_ALERTAS}/ADJUNTOS"


# ─────────────────────────────────────────────────────────────────────────────
# F) PUNTO DE ENTRADA
# ─────────────────────────────────────────────────────────────────────────────

def _rutas_insumos(mes: int, anio: int) -> list[tuple[str, str, bool]]:
    """
    Punto ÚNICO de verdad de dónde vive cada insumo en SharePoint.

    Devuelve [(descripción, ruta_cloud, es_obligatorio)]. Centralizarlo evita
    lo que venía pasando: rutas construidas en 5 sitios distintos del archivo,
    con el año a veces fijo y a veces derivado del periodo, y sin forma de
    saber cuál está mal sin leer todo el log.

    Estructura vigente (todo en la nube):
      BASES DE RESPUESTAS/CIF/<año>/            → Involves_Procesado_<MM>_<AAAA>.csv
      BASES DE RESPUESTAS/EXHIBICIONES/<año>/   → PLANNING DE <MES> <AÑO>.xlsx
                                                  Base Exhibiciones Planning <Mes> <Año>.xlsx
      BASES DE RESPUESTAS/DYP/<año>/Rutero/     → RUTERO_<MES>_<AÑO>_D_P.xlsx
      BASES DE RESPUESTAS/DYP/<año>/Listas/     → listas_referencia.xlsx
      SALIDAS/DYP/                              → Consolidado_Impactos.csv
                                                  Consolidado_Ventas.csv
    """
    import periodo_resolver as _pr
    spec = _pr.resolver(mes, anio)

    # Se prefieren los helpers de paths.py (punto único de verdad). Si aún no
    # están definidos, se cae al armado local normalizando el año del periodo.
    def _p(nombre_helper, *args, fallback=""):
        fn = getattr(paths, nombre_helper, None)
        return fn(*args) if callable(fn) else fallback

    bases_cif   = _carpeta_periodo(getattr(paths, "RUTA_CARPETA_BASES_CIF",   f"{paths._BASES_ROOT}/CIF"), anio)
    bases_dyp   = _carpeta_periodo(getattr(paths, "RUTA_CARPETA_BASES_DYP",   f"{paths._BASES_ROOT}/DYP"), anio)
    bases_exh   = _carpeta_periodo(getattr(paths, "RUTA_CARPETA_BASES_EXHIB", f"{paths._BASES_ROOT}/EXHIBICIONES"), anio)
    salidas_dyp = getattr(paths, "RUTA_CARPETA_SALIDAS_DYP", f"{paths._SALIDAS_ROOT}/DYP")

    return [
        ("CIF — detalle",           ruta_cif_detalle(), True, []),
        ("SOS — detalle",           ruta_sos_detalle(mes, anio), True, []),
        ("NP — detalle",            ruta_np_detalle(mes, anio), True, []),
        ("Precios — detalle",       ruta_pr_detalle(mes, anio), True, []),
        ("KPIs CIF",                f"{paths.RUTA_CARPETA_SALIDAS_CIF}/{paths.CIF_OUT_KPIS.name}", True,
                                     ["KPIS_CIF.xlsx", "CIF_KPIS.xlsx"]),
        ("KPIs SOS",                f"{paths.RUTA_CARPETA_SALIDAS_SOS}/{paths.SOS_OUT_KPIS.name}", True,
                                     ["SOS_KPIS.xlsx", "KPIS_SOS.xlsx"]),
        ("KPIs NP",                 f"{paths.RUTA_CARPETA_SALIDAS_NP}/{paths.NP_OUT_KPIS.name}", True,
                                     ["NP_KPIS.xlsx", "NO_PRESENCIA_KPIS.xlsx"]),
        ("KPIs Precios",            f"{paths.RUTA_CARPETA_SALIDAS_PRECIOS}/{paths.PR_OUT_KPIS.name}", True,
                                     ["PRECIOS_KPIS.xlsx", "KPIS_PRECIOS.xlsx"]),
        ("KPIs Exhib. Pagadas",     f"{paths.RUTA_CARPETA_SALIDAS_EXHIB}/{paths.EXHIB_PAG_OUT_KPIS.name}", True,
                                     ["EXHIBICIONES_PAGADAS_KPIS.xlsx", "KPI_Exhibiciones_Pagadas.xlsx"]),
        # FIX: este era el que te disparaba el ERROR falso. El nombre canónico
        # de paths.py (EXHIBICIONES_GRATIS_KPIS.xlsx) no existe en SharePoint,
        # pero _read_kpi() SÍ lo carga bien por el alterno
        # (KPI_Exhibiciones_Gratis.xlsx). El verificador antes solo probaba el
        # canónico, así que marcaba "OBLIGATORIO ausente" y lanzaba
        # _log.error() aunque el dato se hubiera cargado correctamente 40
        # líneas más abajo. Ahora prueba también los alternos, en el MISMO
        # orden que usa _read_kpi, así el diagnóstico coincide con lo que en
        # verdad va a pasar.
        ("KPIs Exhib. Gratis",      f"{paths.RUTA_CARPETA_SALIDAS_EXHIB}/{paths.EXHIB_GRA_OUT_KPIS.name}", True,
                                     ["KPI_Exhibiciones_Gratis.xlsx", "EXHIBICIONES_GRATIS_KPIS.xlsx"]),
        ("Informe de visitas",      _p("cloud_cif_visitas", mes, anio,
                                       fallback=f"{bases_cif}/Involves_Procesado_{mes:02d}_{anio}.csv"), False, []),
        ("Exh — Planning maestro",  _p("cloud_exhib_planning", mes, anio,
                                       fallback=f"{bases_exh}/PLANNING DE {spec.mes_str_upper} {anio}.xlsx"), False, []),
        ("Exh — Base Planning",     _p("cloud_exhib_base_planning", mes, anio,
                                       fallback=f"{bases_exh}/Base Exhibiciones Planning {spec.mes_str} {anio}.xlsx"), False, []),
        ("Exh — Fuera de regla",    f"{paths.RUTA_CARPETA_SALIDAS_EXHIB}/Exh_Gratis_Fuera_de_Regla.xlsx", False, []),
        # Archivos ÚNICOS acumulados: el periodo se filtra por MES/AÑO, no va
        # en el nombre. El fallback ya no lleva el mes por lo mismo.
        ("D&P — Consolidado Impactos", _p("cloud_dyp_impactos", mes, anio,
                                       fallback=f"{salidas_dyp}/Consolidado_Impactos.csv"), False, []),
        ("D&P — Consolidado Ventas",   _p("cloud_dyp_ventas", mes, anio,
                                       fallback=f"{salidas_dyp}/Consolidado_Ventas.csv"), False, []),
        # FIX: el nombre con mes/año casi nunca existe — en producción el
        # archivo real es genérico (Rutero_Droguerias.xlsx, sin mes ni año).
        # Se agrega como alterno para que el verificador no marque ausente
        # algo que _precargar_segmentos_nuevos sí va a encontrar 2 intentos
        # después, y para no generar 2 peticiones 404 de más en cada corrida.
        # Nombre confirmado: "RUTERO <MES> <AÑO> D&P.xlsx", con espacios y &,
        # cambia cada mes. Se deja la variante con guiones bajos como alterno.
        ("D&P — Rutero",            _p("cloud_dyp_rutero", mes, anio,
                                       fallback=f"{bases_dyp}/Rutero/RUTERO {spec.mes_str_upper} {anio} D&P.xlsx"), False,
                                     [f"{bases_dyp}/Rutero/RUTERO_{spec.mes_str_upper}_{anio}_D_P.xlsx"]),
        ("D&P — Listas",            _p("cloud_dyp_listas", anio,
                                       fallback=f"{bases_dyp}/Listas/{Path(paths.DYP_LISTAS_FILE).name}"), False, []),
        ("Base cupos",              _p("cloud_dyp_base_cupos", anio,
                                       fallback=f"{bases_dyp}/Base_cupos.xlsx"), False, []),
    ]


def verificar_insumos(mes: int, anio: int) -> dict:
    """
    Comprueba la existencia de cada insumo ANTES de procesar, y reporta un
    cuadro claro. Así un archivo renombrado o movido se detecta de entrada,
    en vez de manifestarse como una hoja "No aplica" o un KPI en blanco 20
    minutos después.
    """
    print("\n" + "─" * 60)
    print(f"  VERIFICACIÓN DE INSUMOS — periodo {mes:02d}/{anio}")
    print("─" * 60)

    ok, faltan_obl, faltan_opc = [], [], []
    for desc, ruta, obligatorio, alternos in _rutas_insumos(mes, anio):
        def _existe(r: str) -> bool:
            try:
                token = _obtener_token_azure()
                headers = {"Authorization": f"Bearer {token}"}
                drive_id = _obtener_default_drive_id(token)
                r_cod = urllib.parse.quote(_limpiar_ruta_graph(r))
                url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/root:/{r_cod}"
                return requests.get(url, headers=headers).status_code == 200
            except Exception:
                return False

        # FIX: antes solo se probaba `ruta` (el nombre canónico). Si el
        # archivo real tenía otro nombre pero SÍ era encontrado más adelante
        # vía alterno (en _read_kpi o _primer_disponible), el verificador
        # igual lo marcaba "OBLIGATORIO ausente" y disparaba _log.error(),
        # un falso positivo que no reflejaba lo que en verdad pasaba.
        #
        # FIX 2: los alternos pueden venir como RUTA COMPLETA (caso Rutero) o
        # como SOLO EL NOMBRE del archivo (caso KPIs, que es como los declara
        # _read_kpi). Antes se probaban todos como ruta completa, así que un
        # alterno tipo "KPI_Exhibiciones_Gratis.xlsx" se consultaba contra la
        # RAÍZ del drive y siempre daba 404 — el verificador seguía reportando
        # ausente un archivo que la carga real sí encontraba. Ahora, si el
        # alterno no trae "/", se resuelve contra la carpeta de la canónica.
        carpeta_base = ruta.rsplit("/", 1)[0] if "/" in ruta else ""

        def _resolver_alterno(a: str) -> str:
            return a if "/" in a else (f"{carpeta_base}/{a}" if carpeta_base else a)

        candidatas = [ruta] + [_resolver_alterno(a) for a in alternos]
        vistas, candidatas_unicas = set(), []
        for c in candidatas:
            if c not in vistas:
                vistas.add(c)
                candidatas_unicas.append(c)

        ruta_ok = None
        for candidata in candidatas_unicas:
            if _existe(candidata):
                ruta_ok = candidata
                break

        if ruta_ok is not None:
            ok.append(desc)
            if ruta_ok == ruta:
                print(f"  ✓ {desc}")
            else:
                print(f"  ✓ {desc}  (vía alterno: {ruta_ok})")
        elif obligatorio:
            faltan_obl.append((desc, ruta))
            print(f"  ✗ {desc}  ← OBLIGATORIO")
            print(f"      {ruta}")
            for alt in candidatas_unicas[1:]:
                print(f"      (alterno probado, tampoco existe: {alt})")
        else:
            faltan_opc.append((desc, ruta))
            print(f"  ⚠ {desc}  (opcional — su hoja saldrá vacía)")
            print(f"      {ruta}")

    print("─" * 60)
    print(f"  {len(ok)} OK · {len(faltan_opc)} opcionales ausentes · {len(faltan_obl)} obligatorios ausentes")
    if faltan_obl:
        _log.error("Insumos obligatorios ausentes: %s", [d for d, _ in faltan_obl])
    print("─" * 60)
    return {"ok": ok, "faltan_obligatorios": faltan_obl, "faltan_opcionales": faltan_opc}


def main() -> dict:
    """
    Ejecuta el cálculo completo conectado a la nube usando paths.py.
    Devuelve dict para alertas_telegram.py y alertas_email.py:
      {df_detalle, df_resumen, rutas_adjuntos, mes, anio}
    """
    print("\n" + "═" * 55)
    print("  CALCULAR CUMPLIMIENTOS — Eficacia (Cloud via paths.py)")
    print("═" * 55)
    _log.info("Iniciando cálculo de cumplimientos en la nube con paths.py")

    # NOTA DE ORDEN: los archivos de DETALLE se leen DESPUÉS de los KPIs,
    # porque su nombre lleva el periodo (`..._AGOSTO_2026.xlsx`) y el periodo
    # sale de los KPIs. Antes se leían aquí, de primeras, con un nombre fijo
    # sin periodo — ver el comentario de `ruta_*_detalle()`.

    # ── Cargar Base cupos (tabla maestra de personas, llave ACRONIMO) ────
    print("\nCargando tabla maestra de personas (Base cupos):")
    df_bc   = bcm.cargar()
    universo = bcm.universo_personas(df_bc)
    sups_bc  = bcm.supervisores(df_bc)
    nombres_sups_bc = set(sups_bc["NOMBRE"].astype(str).str.upper())
    
    if "ES_GDD" in df_bc.columns:
        nombres_sups_bc.update(
            df_bc[df_bc["ES_GDD"] == True]["NOMBRE"].astype(str).str.upper()
        )
    if "ES_LIDER" in df_bc.columns:
        nombres_sups_bc.update(
            df_bc[df_bc["ES_LIDER"] == True]["NOMBRE"].astype(str).str.upper()
        )
    idx_bc = bcm.construir_indices(df_bc)
    print(f"  ✓ Universo: {len(universo)} personas activas | {len(sups_bc)} supervisores")

    # ── KPIs V3 ──────────────────────────────────────────────────────────
    print("\nLeyendo KPIs V3 pre-calculados por los ETLs:")
    kpis_v3 = cargar_kpis_v3(idx_bc, nombres_sups_bc)
    df_cif_gest = kpis_v3["cif_gest"]
    df_np_gest  = kpis_v3["np_gest"]
    df_pr_gest  = kpis_v3["pr_gest"]
    df_sos_gest = kpis_v3["sos_gest"]
    df_exp_gest = kpis_v3.get("exp_gest", pd.DataFrame())
    df_egr_gest = kpis_v3.get("egr_gest", pd.DataFrame())

    # ── Periodo activo ────────────────────────────────────────────────────
    # Se resuelve aquí, apenas se tienen los KPIs, porque de él dependen
    # tanto los nombres de los archivos de detalle como las rutas por año.
    ahora = datetime.now()
    if not df_cif_gest.empty and "MES" in df_cif_gest.columns:
        mes  = int(pd.to_numeric(df_cif_gest["MES"], errors="coerce").dropna().mode().iloc[0])
        anio = int(pd.to_numeric(df_cif_gest["AÑO"], errors="coerce").dropna().mode().iloc[0])
    else:
        mes, anio = ahora.month, ahora.year
    print(f"  Periodo activo: {mes:02d}/{anio}")

    global _ANIO_ACTIVO
    _ANIO_ACTIVO = int(anio)

    # ── Candidatos sin cruzar: opcionalmente agregarlos a Base cupos ──────
    # Solo se ejecuta si se corre con --agregar-faltantes (nunca por
    # defecto, para no modificar Base cupos sin que alguien lo pida
    # explícitamente en esa corrida puntual).
    candidatos_faltantes = kpis_v3.get("candidatos_faltantes", [])
    if candidatos_faltantes:
        if "--agregar-faltantes" in sys.argv:
            print(f"\n--agregar-faltantes activo: agregando {len(candidatos_faltantes)} "
                  f"candidato(s) (con duplicados entre módulos) a Base_cupos.xlsx...")
            agregar_candidatos_base_cupos(candidatos_faltantes, mes, anio)
        else:
            print(f"\n  ℹ️  {len(set((c['CEDULA'] or c['NOMBRE']) for c in candidatos_faltantes))} "
                  f"persona(s) sin cruzar quedaron pendientes. Corre con --agregar-faltantes para "
                  f"agregarlas a Base cupos (solo NOMBRE+CEDULA, para completar ACRONIMO/ROL a mano).")
    
    df_cif_sup = pd.DataFrame()
    df_np_sup  = pd.DataFrame()
    df_pr_sup  = pd.DataFrame()
    df_sos_sup = pd.DataFrame()
    print(f"  ✓ CIF        — {len(df_cif_gest)} personas")
    print(f"  ✓ NP         — {len(df_np_gest)} personas")
    print(f"  ✓ Precios    — {len(df_pr_gest)} personas")
    print(f"  ✓ SOS        — {len(df_sos_gest)} personas")
    print(f"  ✓ Exh PAG    — {len(df_exp_gest)} empleados")
    print(f"  ✓ Exh GRATIS — {len(df_egr_gest)} empleados")

    # Verificación temprana de rutas: se hace aquí (y no al inicio) porque
    # varias rutas dependen del periodo, que sale de los KPIs.
    if "--sin-verificar" not in sys.argv:
        verificar_insumos(mes, anio)

    # ── Cargar archivos de DETALLE del periodo desde la nube ─────────────
    print(f"\nCargando archivos de detalle del periodo {mes:02d}/{anio}:")
    df_cif = _leer_excel_cloud(ruta_cif_detalle(),          "CIF (CIF.xlsx)")
    df_sos = _leer_excel_cloud(ruta_sos_detalle(mes, anio), "SOS")
    df_np  = _leer_excel_cloud(ruta_np_detalle(mes, anio),  "No Presencia")
    df_pr  = _leer_excel_cloud(ruta_pr_detalle(mes, anio),  "Precios")

    # ── Maestra de supervisores ──────────────────────────────────────────
    print("\nVerificando maestra de supervisores:")
    generar_maestra(df_np, df_pr, df_sos)

    # ── Filtrar los archivos de DETALLE al periodo activo ────────────────
    # `CIF.xlsx` acumula histórico (varios meses en el mismo archivo), así que
    # este filtro sigue siendo indispensable para él. Los KPIs sí venían
    # filtrados por cargar_kpis_v3(), pero las hojas de detalle de los adjuntos
    # leían el archivo completo y mostraban N meses de PDVs, sin cuadrar con el
    # resumen. SOS/NP/Precios ya llegan del periodo (su archivo lo lleva en el
    # nombre); para ellos el filtro solo confirma que así sea.
    print("\nFiltrando bases de detalle al periodo activo:")

    def _filtrar_periodo_pdv(df: pd.DataFrame, nombre: str) -> pd.DataFrame:
        if df is None or df.empty:
            return df
        if "MES" not in df.columns or "AÑO" not in df.columns:
            print(f"  ⚠️  {nombre}: sin columnas MES/AÑO, no se puede filtrar por periodo.")
            return df
        n0 = len(df)
        out = df[
            (pd.to_numeric(df["MES"], errors="coerce") == mes)
            & (pd.to_numeric(df["AÑO"], errors="coerce") == anio)
        ].copy()
        if out.empty:
            print(f"  ⚠️  {nombre}: {n0} filas, NINGUNA del periodo {mes:02d}/{anio} "
                  f"(periodos presentes: "
                  f"{sorted(set(zip(pd.to_numeric(df['MES'], errors='coerce'), pd.to_numeric(df['AÑO'], errors='coerce'))))})")
        elif len(out) != n0:
            print(f"  ✓ {nombre}: {n0:,} filas → {len(out):,} del periodo {mes:02d}/{anio}")
        else:
            print(f"  ✓ {nombre}: {n0:,} filas (ya venía solo del periodo)")
        return out

    df_cif = _filtrar_periodo_pdv(df_cif, "Plan de trabajo (CIF)")
    df_sos = _filtrar_periodo_pdv(df_sos, "SOS")
    df_np  = _filtrar_periodo_pdv(df_np,  "No Presencia")
    df_pr  = _filtrar_periodo_pdv(df_pr,  "Precios")
    df_cif_pdv = df_cif

    # ── KPIs D&P ──────────────────────────────────────────────────────────
    kpis_dyp = cumplimiento_dyp.calcular_kpis_dyp(
        mes, anio,
        base_cupos_idx=idx_bc,
        nombres_supervisores_bc=nombres_sups_bc,
    )
    df_dyp_gest = kpis_dyp["gestor"]
    df_dyp_sup  = kpis_dyp["supervisor"]

    # ── Ensamblaje final ──────────────────────────────────────────────────
    print("\nEnsamblando detalle consolidado:")
    df_detalle = ensamblar_detalle(
        universo,
        df_cif_gest, df_cif_sup,
        df_np_gest,  df_np_sup,
        df_pr_gest,  df_pr_sup,
        df_sos_gest, df_sos_sup,
        df_dyp_gest=df_dyp_gest,
        df_dyp_sup=df_dyp_sup,
        df_exp_gest=df_exp_gest,
        df_egr_gest=df_egr_gest,
        mes=mes, anio=anio,
    )

    n_gest = (df_detalle["ES_SUPERVISOR"] == False).sum()
    n_sup  = (df_detalle["ES_SUPERVISOR"] == True).sum()
    print(f"  Personas en detalle: {len(df_detalle)} (gestores: {n_gest} | supervisores: {n_sup})")

    # ── Resumen ───────────────────────────────────────────────────────────
    print("\nCalculando resumen por supervisor:")
    df_resumen = calcular_resumen(df_detalle, universo)
    print(f"  Supervisores en resumen: {len(df_resumen)}")

    # ── Guardar archivos usando las rutas de `paths.py` ────────────────────
    print("\nGuardando archivos de salida en SharePoint:")
    ruta_salida_detalle = f"{paths._SALIDAS_ROOT}/Detalle_Cumplimientos_{anio}_{mes:02d}.xlsx"
    ruta_salida_resumen = f"{paths._SALIDAS_ROOT}/Resumen_Supervisores_{anio}_{mes:02d}.xlsx"
    
    _guardar_excel_cloud(df_detalle, ruta_salida_detalle, "Detalle Consolidado")
    _guardar_excel_cloud(df_resumen, ruta_salida_resumen, "Resumen por Supervisor")

    def _agregar_acr_sup(df, col_sup_origen="SUPERVISOR_LIDER"):
        if df is None or df.empty or col_sup_origen not in df.columns:
            return df
        df = df.copy()
        df["ACRONIMO_SUP"] = df[col_sup_origen].apply(
            lambda s: _resolver_acr_supervisor(s, nombres_sups_bc, idx_bc)
        )
        return df

    df_cif_pdv_e = _agregar_acr_sup(df_cif_pdv)
    df_np_e      = _agregar_acr_sup(df_np)
    df_pr_e      = _agregar_acr_sup(df_pr)
    df_sos_e     = _agregar_acr_sup(df_sos)

    print("\nPre-cargando insumos D&P para hojas de productos nuevos:")
    seg_data = _precargar_segmentos_nuevos(mes, anio, idx_bc, nombres_sups_bc)
    if seg_data.get("rutero") is not None:
        n_sups_con_pdvs = len(seg_data.get("pdvs_por_sup") or {})
        print(f"  ✓ Segmentos listos — supervisores con PDVs en periodo: {n_sups_con_pdvs}")

    rutas_adj = generar_adjuntos_por_supervisor(
        df_detalle, df_cif_pdv_e, df_np_e, df_pr_e, df_sos_e, mes, anio,
        seg_data=seg_data, idx_bc=idx_bc,
    )

    _log.info(
        f"Cálculo completado (Cloud vía paths) — periodo {mes:02d}/{anio} | "
        f"gestores={n_gest} supervisores={n_sup} adjuntos={len(rutas_adj)}"
    )
    print("\n" + "═" * 55)
    print(f"  ✅ Proceso completado en la Nube (SharePoint) — {mes:02d}/{anio}")
    print("═" * 55 + "\n")

    rango_periodo = detectar_rango_periodo(mes, anio)
    if rango_periodo.get("rango_legible"):
        print(f"\n  📆 Periodo de corte detectado: {rango_periodo['rango_legible']}"
              f" | avance esperado del mes: {rango_periodo['avance_esperado_pct']*100:.0f}%")

    return {
        "df_detalle"     : df_detalle,
        "df_resumen"     : df_resumen,
        "rutas_adjuntos" : rutas_adj,
        "mes"            : mes,
        "anio"           : anio,
        "rango_periodo"  : rango_periodo,
        "df_cif_pdv"     : df_cif_pdv_e,
        "df_np"          : df_np_e,
        "df_pr"          : df_pr_e,
        "df_sos"         : df_sos_e,
    }

##############################################################################################################################
##############################################################################################################################
##############################################################################################################################
##############################################################################################################################
##############################################################################################################################






def _ultimo_archivo(directorio: Path, patron: str) -> Path:
    """Devuelve el archivo más reciente que matchea `patron`."""
    candidatos = sorted(
        directorio.glob(patron),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    return candidatos[0] if candidatos else directorio / patron


# Umbrales y ponderación de negocio
UMBRAL_OK      = float(os.environ.get("UMBRAL_OK",      "0.90"))
UMBRAL_WARNING = float(os.environ.get("UMBRAL_WARNING", "0.70"))

# ─────────────────────────────────────────────────────────────────────────────
# UMBRALES POR KPI (Sprint 13.4)  — defaults D17
# ─────────────────────────────────────────────────────────────────────────────
def _umbral(kpi: str, default: float) -> float:
    return float(os.environ.get(f"UMBRAL_OK_{kpi}", str(default)))

UMBRALES_OK = {
    "CIF":                _umbral("CIF",                0.90),
    "NP":                 _umbral("NP",                 1.00),
    "PRECIOS":            _umbral("PRECIOS",            1.00),
    "SOS":                _umbral("SOS",                1.00),
    "EXHIB_PAG":          _umbral("EXHIB_PAG",          1.00),
    "EXHIB_PAG_CAPTURA":  _umbral("EXHIB_PAG_CAPTURA",  1.00),  # Sprint 17
    "EXHIB_GRA_ALTO":     _umbral("EXHIB_GRA_ALTO",     1.00),  # Sprint 17
    "EXHIB_GRA_MEDIO":    _umbral("EXHIB_GRA_MEDIO",    1.00),  # Sprint 17
    "EXHIB_GRATIS_PROM":  _umbral("EXHIB_GRATIS_PROM",  1.00),  # Sprint 17.13 — promedio alto+medio
    "VENTA":              _umbral("VENTA",              0.90),
    "IMPACTOS":           _umbral("IMPACTOS",           0.90),
    "MSL":                _umbral("MSL",                1.00),
    "PROD_NUEVOS":        _umbral("PROD_NUEVOS",        0.70),
    "GLOBAL":             _umbral("GLOBAL",             0.90),
}

# Mapeo de columna del detalle → KPI key en UMBRALES_OK
COL_A_KPI = {
    "CIF_%":              "CIF",
    "NP_%":               "NP",
    "PRECIOS_%":          "PRECIOS",
    "SOS_%":              "SOS",
    "EXHIB_PAG_%":        "EXHIB_PAG",
    "EXHIB_PAG_CAPTURA_%":"EXHIB_PAG_CAPTURA",    # Sprint 17
    "EXHIB_GRA_ALTO_%":   "EXHIB_GRA_ALTO",      # Sprint 17
    "EXHIB_GRA_MEDIO_%":  "EXHIB_GRA_MEDIO",     # Sprint 17
    "EXHIB_GRATIS_PROM_%":"EXHIB_GRATIS_PROM",    # Sprint 17.13
    "VENTA_%":            "VENTA",
    "IMPACTOS_%":         "IMPACTOS",
    "MSL_%":              "MSL",
    "PROD_NUEVOS_%":      "PROD_NUEVOS",
    "CUMPL_GLOBAL_%":     "GLOBAL",
}


def umbral_de(col_o_kpi: str) -> float:
    """Devuelve UMBRAL_OK para una columna del detalle o una KPI key."""
    kpi = COL_A_KPI.get(col_o_kpi, col_o_kpi)
    return UMBRALES_OK.get(kpi, UMBRAL_OK)


# Roles usados como filtros en adjuntos / detalle.
ROL_GESTOR     = "GESTOR"
ROL_SUPERVISOR = "SUPERVISOR"


# ─────────────────────────────────────────────────────────────────────────────
# RANGO DE FECHAS DEL PERIODO (Sprint 13.1 - Adaptado a Nube)
# ─────────────────────────────────────────────────────────────────────────────

def detectar_rango_periodo(mes: int, anio: int) -> dict:
    """
    Detecta el rango de fechas activas del periodo desde
    el reporte de visitas procesado en SharePoint (FECHA_VISITA).

    Devuelve dict con:
      fecha_inicio_dt, fecha_fin_dt    — datetime.date | None
      fecha_inicio, fecha_fin          — strings "DD de MES"
      rango_legible                    — "1 al 6 de abril" o "1 de marzo al 6 de abril"
      dias_transcurridos, dias_mes_total, avance_esperado_pct
    """
    from datetime import date
    import calendar

    rango = {
        "fecha_inicio_dt": None, "fecha_fin_dt": None,
        "fecha_inicio": "", "fecha_fin": "",
        "rango_legible": "",
        "dias_transcurridos": 0, "dias_mes_total": 0, "avance_esperado_pct": 0.0,
    }

    try:
        # FIX: el archivo se renombró a Involves_Procesado_<MM>_<AAAA>.csv y se
        # movió a BASES DE RESPUESTAS/CIF/<año>. OJO: paths.RUTA_CARPETA_BASES_CIF
        # YA incluye el año (con AÑO_ACTUAL, el del reloj), así que concatenarlo
        # de nuevo daría .../CIF/2026/2026/. _carpeta_periodo() normaliza y de
        # paso fuerza el año del PERIODO en vez del año del sistema.
        if hasattr(paths, "cloud_cif_visitas"):
            ruta_visitas_cloud = paths.cloud_cif_visitas(mes, anio)
        else:
            _base_cif = getattr(paths, "RUTA_CARPETA_BASES_CIF", f"{paths._BASES_ROOT}/CIF")
            ruta_visitas_cloud = (
                f"{_carpeta_periodo(_base_cif, anio)}/Involves_Procesado_{mes:02d}_{anio}.csv"
            )
        token = _obtener_token_azure()
        headers = {"Authorization": f"Bearer {token}"}
        drive_id = _obtener_default_drive_id(token)
        ruta_limpia = _limpiar_ruta_graph(ruta_visitas_cloud)
        
        url_codificada = urllib.parse.quote(ruta_limpia)
        url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/root:/{url_codificada}:/content"
        
        response = requests.get(url, headers=headers)
        if response.status_code != 200:
            print(f"  ⚠️  No pude leer el informe de visitas ({response.status_code}): {ruta_visitas_cloud}")
            print(f"      El correo saldrá sin rango de corte. Verifica nombre/ubicación del archivo.")
            _log.warning("Informe de visitas no encontrado (%s): %s",
                         response.status_code, ruta_visitas_cloud)
            return rango

        df = pd.read_csv(
            io.BytesIO(response.content),
            sep=";", encoding="utf-8-sig", usecols=["FECHA_VISITA"],
        )
        df["F"] = pd.to_datetime(df["FECHA_VISITA"], dayfirst=True, errors="coerce")
        # Filtrar al mes/año del periodo activo
        df = df[(df["F"].dt.month == mes) & (df["F"].dt.year == anio)]
        if df.empty:
            return rango
        f_ini = df["F"].min().date()
        f_fin = df["F"].max().date()
        rango["fecha_inicio_dt"] = f_ini
        rango["fecha_fin_dt"]    = f_fin
    except Exception as e:
        # FIX: antes era un `except Exception: return rango` mudo — cualquier
        # cambio de nombre de archivo, separador o columna se volvía invisible.
        print(f"  ⚠️  detectar_rango_periodo falló: {e}")
        _log.warning("detectar_rango_periodo falló: %s", e)
        return rango

    meses_es = {1:"enero", 2:"febrero", 3:"marzo", 4:"abril", 5:"mayo",
                6:"junio", 7:"julio", 8:"agosto", 9:"septiembre",
                10:"octubre", 11:"noviembre", 12:"diciembre"}
    rango["fecha_inicio"] = f"{f_ini.day} de {meses_es.get(f_ini.month, '')}"
    rango["fecha_fin"]    = f"{f_fin.day} de {meses_es.get(f_fin.month, '')}"

    if f_ini.month == f_fin.month and f_ini.year == f_fin.year:
        rango["rango_legible"] = (
            f"{f_ini.day} al {f_fin.day} de {meses_es.get(f_fin.month, '')}"
        )
    else:
        rango["rango_legible"] = f"{rango['fecha_inicio']} al {rango['fecha_fin']}"

    dias_mes = calendar.monthrange(anio, mes)[1]
    dias_trans = f_fin.day  
    rango["dias_mes_total"] = dias_mes
    rango["dias_transcurridos"] = dias_trans
    rango["avance_esperado_pct"] = (dias_trans / dias_mes) if dias_mes > 0 else 0.0
    return rango


# ─────────────────────────────────────────────────────────────────────────────
# COLORES PARA EL EXCEL DE SALIDA
# ─────────────────────────────────────────────────────────────────────────────
COLOR_HEADER     = "1F3864"
COLOR_HEADER_FNT = "FFFFFF"
COLOR_OK         = "C6EFCE"
COLOR_WARN       = "FFEB9C"
COLOR_ERROR      = "FFC7CE"
COLOR_FILA_PAR   = "F2F2F2"
COLOR_MAESTRA_HDR= "2E4057"

FMT_PCT   = "0.00%"            
FMT_VENTA = '"$"#,##0'
FMT_NUM   = "0.####"            


# ─────────────────────────────────────────────────────────────────────────────
# UTILIDADES (Adaptadas a Nube)
# ─────────────────────────────────────────────────────────────────────────────

def _semaforo(valor) -> str:
    if pd.isna(valor):
        return "—"
    if valor >= UMBRAL_OK:
        return "✅"
    if valor >= UMBRAL_WARNING:
        return "⚠️"
    return "❌"


def _color_celda(valor) -> str:
    """Color hex (sin #) según umbral."""
    if pd.isna(valor):
        return COLOR_FILA_PAR
    if valor >= UMBRAL_OK:
        return COLOR_OK
    if valor >= UMBRAL_WARNING:
        return COLOR_WARN
    return COLOR_ERROR


def _safe_div(num: pd.Series, den: pd.Series) -> pd.Series:
    """División segura: NaN donde denominador ≤ 0 o NaN."""
    return num.where(den > 0, other=np.nan) / den.where(den > 0, other=np.nan)


def _leer(ruta_sharepoint: str, nombre: str) -> pd.DataFrame:
    """Carga un Excel desde SharePoint utilizando Graph API con manejo descriptivo de error."""
    try:
        print(f"    ⏳ Leyendo [{nombre}] desde SharePoint ({ruta_sharepoint})...")
        token = _obtener_token_azure()
        headers = {"Authorization": f"Bearer {token}"}
        drive_id = _obtener_default_drive_id(token)
        ruta_limpia = _limpiar_ruta_graph(ruta_sharepoint)
        
        url_codificada = urllib.parse.quote(ruta_limpia)
        url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/root:/{url_codificada}:/content"
        
        response = requests.get(url, headers=headers)
        if response.status_code != 200:
            print(f"    ⚠️  [{nombre}] No se pudo encontrar el archivo en la ruta: {ruta_sharepoint}")
            print(f"        Este módulo no se incluirá en el cálculo.")
            return pd.DataFrame()
            
        df = pd.read_excel(io.BytesIO(response.content))
        df.columns = df.columns.str.strip()
        print(f"    ✓ [{nombre}] {len(df):,} filas cargadas desde la nube")
        return df
    except Exception as e:
        print(f"    ⚠️  [{nombre}] Error al cargar desde la nube: {e}")
        return pd.DataFrame()


def _normalizar_nombre(serie: pd.Series) -> pd.Series:
    """Strip + upper para evitar fallos de merge por espacios o casing."""
    return serie.astype(str).str.strip().str.upper()


def _w_avg(valores: pd.Series, pesos: pd.Series) -> float:
    """
    Promedio ponderado robusto: ignora pares donde el valor es NaN o el
    peso es NaN/0. Devuelve NaN si no hay pares válidos.
    """
    pesos = pesos.fillna(0)
    mask = valores.notna() & (pesos > 0)
    if not mask.any():
        return np.nan
    w = pesos[mask]
    v = valores[mask]
    s = w.sum()
    if s <= 0:
        return np.nan
    return float((v * w).sum() / s)


def _promedio_simple(*vals) -> float:
    """Promedio aritmético ignorando NaN."""
    valid = [v for v in vals if v is not None and not pd.isna(v)]
    return float(sum(valid) / len(valid)) if valid else np.nan


def _coalesce_periodo(df: pd.DataFrame, default_mes: int, default_anio: int) -> tuple[int, int]:
    """Detecta MES y AÑO predominante en un df, con fallback."""
    if "MES" in df.columns and df["MES"].notna().any():
        mes = int(pd.to_numeric(df["MES"], errors="coerce").dropna().mode().iloc[0])
    else:
        mes = default_mes
    if "AÑO" in df.columns and df["AÑO"].notna().any():
        anio = int(pd.to_numeric(df["AÑO"], errors="coerce").dropna().mode().iloc[0])
    else:
        anio = default_anio
    return mes, anio


# ─────────────────────────────────────────────────────────────────────────────
# RESOLUCIÓN DE ACRONIMOS (Base cupos como llave canónica)
# ─────────────────────────────────────────────────────────────────────────────
# Filler usado para armonización por palabras
_PALABRAS_FILLER = {"DE", "LA", "LAS", "LOS", "DEL", "Y"}


def _sin_tildes(s: str) -> str:
    """Quita tildes/diacríticos (á→a, ñ→n, etc.) para comparar nombres que
    llegan escritos con o sin tilde según la fuente (KPI vs Base cupos)."""
    return unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")


def _palabras(s: str) -> set:
    """Tokens significativos de un nombre (excluye fillers, ignora tildes)."""
    s_plano = _sin_tildes(str(s).upper())
    return {w for w in s_plano.split() if w and w not in _PALABRAS_FILLER}


def _resolver_acr_supervisor(nombre_sup: str, nombres_sups_bc: set, idx: dict) -> str:
    """
    Mapea un nombre de supervisor (puede venir truncado) al ACRONIMO
    canónico de Base cupos. Usa armonización por subset de palabras.
    Devuelve "" si no hay match.
    """
    if not nombre_sup or pd.isna(nombre_sup):
        return ""
    s = str(nombre_sup).strip().upper()
    nombre_a_acr = idx["nombre_a_acronimo"]
    if s in nombre_a_acr and s in nombres_sups_bc:
        return nombre_a_acr[s]
    pal_v = _palabras(s)
    if not pal_v:
        return ""
    for nombre_pt in nombres_sups_bc:
        pal_pt = _palabras(nombre_pt)
        if pal_v.issubset(pal_pt) or pal_pt.issubset(pal_v):
            return nombre_a_acr.get(nombre_pt, "")
    return ""


def _resolver_acr_fuzzy_general(nombre: str, idx: dict) -> str:
    """
    Respaldo del match exacto por nombre, para CUALQUIER persona (gestor
    o supervisor) — no solo supervisores, a diferencia de
    _resolver_acr_supervisor. Tolera nombres con un nombre/apellido de
    más o de menos (ej. "JOSE LIBARDO JOYA" en el KPI vs "LIBARDO JOYA"
    en Base cupos) comparando por subconjunto de palabras.

    Si el nombre matchea por palabras contra MÁS DE UNA persona distinta
    de Base cupos, no resuelve nada (devuelve "") — es preferible dejar
    el ACRONIMO vacío que asignarle a alguien el KPI de otra persona.
    """
    if not nombre or pd.isna(nombre):
        return ""
    pal_v = _palabras(str(nombre).strip().upper())
    if not pal_v:
        return ""
    nombre_a_acr = idx.get("nombre_a_acronimo", {})
    candidatos = set()
    for nombre_bc, acr in nombre_a_acr.items():
        pal_bc = _palabras(nombre_bc)
        if pal_bc and (pal_v.issubset(pal_bc) or pal_bc.issubset(pal_v)):
            # FIX: un ACRONIMO vacío (fila de Base cupos sin completar) no es
            # un match válido — antes se agregaba al set y podía "resolver" a "".
            if str(acr).strip():
                candidatos.add(str(acr).strip())
        if len(candidatos) > 1:
            return ""  # ambiguo — corta apenas encuentra un segundo candidato
    return candidatos.pop() if len(candidatos) == 1 else ""


_CACHE_BC_SIN_ACR: dict = {}

# Año del periodo que se está procesando. Lo fija main() al detectarlo; sirve
# para que utilidades de diagnóstico resuelvan la carpeta correcta sin tener
# que recibir el año por parámetro desde 6 sitios distintos.
_ANIO_ACTIVO: int = datetime.now().year


def _base_cupos_sin_acronimo() -> tuple[set, set]:
    """
    Devuelve (nombres, cédulas) de las filas de Base_cupos.xlsx que existen
    pero NO tienen ACRONIMO.

    Se lee el Excel crudo desde SharePoint en vez de pedírselo a
    `base_cupos.py`, porque ese módulo descarta esas filas al normalizar
    ("Se descartaron N filas por no tener ACRONIMO o NOMBRE") y por tanto no
    puede informarlas. Sin esto, el diagnóstico del cruce dice "probablemente
    no está en Base cupos" para gente que SÍ está — solo con la celda vacía —
    y manda a buscar el problema donde no está.

    Resultado cacheado: se descarga una sola vez por corrida.
    """
    if "data" in _CACHE_BC_SIN_ACR:
        return _CACHE_BC_SIN_ACR["data"]

    nombres, cedulas = set(), set()
    try:
        ruta_bc = f"{_carpeta_periodo(paths.RUTA_CARPETA_BASES_DYP, _ANIO_ACTIVO)}/Base_cupos.xlsx"
        df_bc = _leer(ruta_bc, "Base_cupos.xlsx (diagnóstico de filas sin ACRONIMO)")
        if not df_bc.empty:
            df_bc.columns = [str(c).strip().upper() for c in df_bc.columns]
            if "ACRONIMO" in df_bc.columns and "NOMBRE" in df_bc.columns:
                acr_vacio = (
                    df_bc["ACRONIMO"].isna()
                    | (df_bc["ACRONIMO"].astype(str).str.strip() == "")
                    | (df_bc["ACRONIMO"].astype(str).str.strip().str.lower() == "nan")
                )
                sin = df_bc[acr_vacio]
                nombres = set(sin["NOMBRE"].astype(str).str.strip().str.upper())
                if "CEDULA" in sin.columns:
                    cedulas = set(
                        sin["CEDULA"].astype(str).str.strip().str.replace(r"\.0$", "", regex=True)
                    )
    except Exception as e:
        _log.warning("No pude leer Base_cupos crudo para el diagnóstico: %s", e)

    _CACHE_BC_SIN_ACR["data"] = (nombres, cedulas)
    return nombres, cedulas


def _candidatos_parecidos(nombre: str, idx: dict, top: int = 3) -> list[str]:
    """
    Para diagnóstico: busca en Base cupos los nombres que comparten AL MENOS
    una palabra significativa con `nombre` (no exige subconjunto completo
    como _resolver_acr_fuzzy_general — esto es solo para mostrarle al
    usuario qué tan cerca/lejos está el nombre real de Base cupos, y así
    distinguir "es solo una tilde" de "esta persona no está en Base cupos").
    Devuelve hasta `top` nombres, ordenados por cantidad de palabras en común.
    """
    pal_v = _palabras(str(nombre).strip().upper())
    if not pal_v:
        return []
    puntajes = []
    for nombre_bc in idx.get("nombre_a_acronimo", {}):
        pal_bc = _palabras(nombre_bc)
        comunes = pal_v & pal_bc
        if comunes:
            puntajes.append((len(comunes), nombre_bc))
    puntajes.sort(key=lambda t: -t[0])
    return [n for _, n in puntajes[:top]]


def agregar_candidatos_base_cupos(candidatos: list[dict], mes: int, anio: int) -> int:
    """
    Agrega a Base_cupos.xlsx (en SharePoint) una fila nueva por cada
    persona en `candidatos` que no se logró cruzar (ni por cédula ni por
    nombre) en esta corrida. Solo llena NOMBRE y CEDULA — el resto de
    columnas (ACRONIMO, ROL EN ACTIVO, CANAL, COD PT, etc.) quedan en
    blanco a propósito, para que alguien las complete a mano.

    Como ACRONIMO queda vacío, estas filas nuevas NO entran al universo
    activo (bcm.cargar() las descarta vía el filtro sin_acronimo_o_nombre)
    hasta que alguien las complete — no afectan el cálculo de esta corrida
    ni de las siguientes mientras sigan incompletas.

    Deduplica por CEDULA (o por NOMBRE si no trae cédula) y no vuelve a
    agregar a alguien que ya está en Base cupos con esa cédula o nombre
    exacto (para no duplicar entradas ya completadas manualmente).

    Devuelve cuántas filas nuevas se agregaron.
    """
    if not candidatos:
        return 0

    # FIX: usaba paths.RUTA_CARPETA_BASES_DYP directamente. Esa constante trae
    # el año HORNEADO con AÑO_ACTUAL (el del reloj, evaluado una sola vez al
    # importar paths.py) — el mismo problema que ya se corrigió en las otras 4
    # rutas de BASES (CIF, Exhibiciones, D&P para Rutero/Listas/Ventas), pero
    # aquí se había quedado sin tocar. Si se corre este comando en enero para
    # cerrar diciembre, o en cualquier backfill de un periodo anterior, escribía
    # (o intentaba leer) Base_cupos.xlsx en la carpeta del año EQUIVOCADO.
    ruta_bc = f"{_carpeta_periodo(paths.RUTA_CARPETA_BASES_DYP, anio)}/Base_cupos.xlsx"
    df_bc = _leer(ruta_bc, "Base_cupos.xlsx (para agregar candidatos)")
    if df_bc.empty:
        print("  ⚠️  No se pudo leer Base_cupos.xlsx — no se agregó ningún candidato.")
        return 0

    cedulas_existentes = (
        set(df_bc["CEDULA"].astype(str).str.strip().replace({"nan": ""}))
        if "CEDULA" in df_bc.columns else set()
    )
    nombres_existentes = (
        set(df_bc["NOMBRE"].astype(str).str.strip().str.upper())
        if "NOMBRE" in df_bc.columns else set()
    )

    # Deduplicar candidatos entre sí — la misma persona puede salir "sin
    # match" en varios módulos a la vez (CIF, NP, Precios, SOS, Exhib...),
    # y en algunos de esos módulos trae cédula y en otros no (Exhibiciones
    # nunca tuvo esa columna). Por eso se agrupa SIEMPRE por NOMBRE primero
    # — nunca por cédula-o-nombre — y de todas las ocurrencias de ese
    # nombre se toma la primera cédula no vacía que aparezca, para no
    # terminar con dos filas para la misma persona (una con cédula y otra
    # "(sin cédula)").
    por_nombre: dict[str, str] = {}
    for c in candidatos:
        nombre = str(c.get("NOMBRE", "")).strip().upper()
        cedula = str(c.get("CEDULA", "")).strip()
        if not nombre:
            continue
        if nombre not in por_nombre or (not por_nombre[nombre] and cedula):
            por_nombre[nombre] = cedula

    nuevas_filas = []
    for nombre, cedula in por_nombre.items():
        if cedula and cedula in cedulas_existentes:
            continue  # ya está en Base cupos con esa cédula
        if not cedula and nombre in nombres_existentes:
            continue  # ya está con ese nombre exacto

        fila = {col: "" for col in df_bc.columns}
        if "NOMBRE" in fila:
            fila["NOMBRE"] = nombre
        if "CEDULA" in fila:
            fila["CEDULA"] = cedula
        nuevas_filas.append(fila)

    if not nuevas_filas:
        print("  ℹ️  Ningún candidato nuevo que agregar (todos ya estaban en Base cupos o se repetían).")
        return 0

    df_nuevas = pd.DataFrame(nuevas_filas)
    df_bc_actualizado = pd.concat([df_bc, df_nuevas], ignore_index=True)

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df_bc_actualizado.to_excel(writer, sheet_name="Tabla total roles", index=False)
    _subir_bytes_cloud(buffer.getvalue(), ruta_bc, "Base_cupos.xlsx (candidatos agregados)")

    print(f"  ✅ Se agregaron {len(nuevas_filas)} fila(s) nueva(s) a Base_cupos.xlsx "
          f"(solo NOMBRE + CEDULA — completa ACRONIMO/ROL/CANAL/etc. manualmente):")
    for f in nuevas_filas:
        print(f"     · {f.get('NOMBRE', '')} — cédula: {f.get('CEDULA') or '(sin cédula)'}")
    return len(nuevas_filas)


# Evita que _canonizar_nombre_gestor repita el mismo aviso una vez por cada
# adjunto (37 veces en una corrida típica) para la misma persona.
_AVISADOS_CANONIZAR: set = set()

# Personas ya reportadas como "está en Base cupos pero sin ACRONIMO". El
# diagnóstico corre una vez por módulo (CIF/SOS/NP/Precios/Exh), así que sin
# esto la misma persona se reporta hasta 6 veces por corrida.
_AVISADOS_SIN_ACRONIMO: set = set()


def _canonizar_nombre_gestor(df: pd.DataFrame, idx_bc: dict, col: str = "NOMBRE_GESTOR") -> pd.DataFrame:
    """
    Reemplaza el nombre "crudo" de cada gestor (tal como quedó escrito en
    el detalle fuente — puede variar entre filas para la MISMA persona,
    ej. "LIBARDO JOYA" en una visita y "LIBARDO ANTONIO JOYA TEJADA" en
    otra) por su nombre canónico en Base cupos, para que las hojas de
    detalle de los adjuntos siempre muestren un único nombre consistente
    por persona, sin importar cómo se haya capturado cada fila.

    Prioridad para resolver de quién se trata: CEDULA (si la columna
    existe) → nombre exacto → nombre por palabras (fuzzy, sin tildes).
    Si no logra resolver a nadie, deja el nombre crudo tal cual (no borra
    ni deja vacío — mejor mostrar el nombre que tenga, aunque no esté
    unificado, que perder la fila).
    """
    if df is None or df.empty or col not in df.columns:
        return df
    df = df.copy()

    cedula_a_acr = idx_bc.get("cedula_a_acronimo", {})
    nombre_a_acr = idx_bc.get("nombre_a_acronimo", {})
    acr_a_nombre = idx_bc.get("acronimo_a_nombre", {})

    if "CEDULA" in df.columns and cedula_a_acr:
        cedula_norm = df["CEDULA"].astype(str).str.strip().str.replace(r"\.0$", "", regex=True)
        acr = cedula_norm.map(cedula_a_acr).fillna("")
    else:
        acr = pd.Series([""] * len(df), index=df.index)

    falta = acr == ""
    if falta.any():
        acr.loc[falta] = df.loc[falta, col].astype(str).str.strip().str.upper().map(nombre_a_acr).fillna("")

    falta = acr == ""
    if falta.any():
        acr.loc[falta] = df.loc[falta, col].apply(lambda n: _resolver_acr_fuzzy_general(n, idx_bc))

    falta = acr == ""
    if falta.any():
        nombres_sin_resolver = df.loc[falta, col].astype(str).str.strip().unique().tolist()
        # FIX: esta función corre una vez POR ADJUNTO (37 veces en una corrida
        # típica), y antes repetía el mismo aviso para la misma persona cada
        # vez — en la corrida real llegó a imprimir el mismo nombre 7-8 veces
        # seguidas, ahogando el log. Con 8 personas sin cruzar eso son ~60
        # líneas de ruido idéntico. Se avisa una sola vez por corrida.
        for n in nombres_sin_resolver:
            if n in _AVISADOS_CANONIZAR:
                continue
            _AVISADOS_CANONIZAR.add(n)
            mask_n = df[col].astype(str).str.strip() == n
            if "CEDULA" in df.columns:
                cedulas_vistas = df.loc[mask_n, "CEDULA"].astype(str).str.strip().unique().tolist()
            else:
                cedulas_vistas = ["(sin columna CEDULA)"]
            candidatos = _candidatos_parecidos(n, idx_bc, top=3)
            print(f"    ⚠️  _canonizar_nombre_gestor no pudo resolver '{n}' "
                  f"(cédula(s) vista(s) en sus filas: {cedulas_vistas}) — "
                  f"candidato(s) parecido(s) en Base cupos: {candidatos or '(ninguno)'}")

    nombre_canonico = acr.map(acr_a_nombre)
    df[col] = nombre_canonico.fillna(df[col])
    return df


def _enriquecer_acronimo_pt(
    df_calc: pd.DataFrame,
    df_pdv: pd.DataFrame,
    idx: dict,
    nombres_sups_bc: set,
) -> pd.DataFrame:
    """
    Enriquece un DataFrame de cálculo (CIF/NP/Precios/SOS por persona)
    con dos columnas:
        ACRONIMO     → ACRONIMO canónico de Base cupos para esta persona.
        ACRONIMO_SUP → ACRONIMO canónico del supervisor a cargo.

    df_calc: salida de cif_por_gestor / calcular_modulo_simple (gestor side).
             Tiene NOMBRE (y opcional SUPERVISOR_LIDER).
    df_pdv:  el DataFrame fuente que tiene ACRONIMO + CEDULA + NOMBRE.
    """
    if df_calc.empty:
        out = df_calc.copy()
        out["ACRONIMO"]     = ""
        out["ACRONIMO_SUP"] = ""
        return out

    # Recuperar ACRONIMO + CEDULA originales desde la fuente, una fila por NOMBRE
    cols_src = [c for c in ["NOMBRE", "CEDULA", "ACRONIMO"] if c in df_pdv.columns]
    src = df_pdv[cols_src].copy()
    src["NOMBRE"] = src["NOMBRE"].astype(str).str.strip().str.upper()
    if "CEDULA" in src.columns:
        src["CEDULA"] = (
            src["CEDULA"].astype(str).str.strip().str.replace(r"\.0$", "", regex=True)
        )
    if "ACRONIMO" in src.columns:
        src["ACRONIMO"] = src["ACRONIMO"].astype(str).str.strip().str.upper()
    src = src.dropna(subset=["NOMBRE"]).drop_duplicates(subset=["NOMBRE"], keep="first")

    out = df_calc.merge(src, on="NOMBRE", how="left", suffixes=("", "_src"))

    # Resolver ACRONIMO canónico (cascada de fallbacks):
    #   1) ACRONIMO del PT si existe y está en la maestra
    #   2) CEDULA del PT si está en la maestra
    #   3) NOMBRE del PT (último recurso, exact match contra Base cupos)
    if "ACRONIMO" in out.columns:
        cedula_serie = out["CEDULA"] if "CEDULA" in out.columns else None
        out["ACRONIMO"] = bcm.resolver_acronimo(
            out["ACRONIMO"].fillna(""),
            idx,
            "acronimo+cedula",
            df_pt_cedula=cedula_serie if cedula_serie is not None else None,
        )
    else:
        out["ACRONIMO"] = ""

    # Fallback por NOMBRE para filas que aún no resolvieron (típico de NP,
    # cuyo reporte no incluye ACRONIMO ni CEDULA).
    if "NOMBRE" in out.columns:
        nombre_a_acr = idx["nombre_a_acronimo"]
        falta = out["ACRONIMO"].fillna("").astype(str) == ""
        if falta.any():
            out.loc[falta, "ACRONIMO"] = (
                out.loc[falta, "NOMBRE"].astype(str).str.strip().str.upper()
                   .map(nombre_a_acr).fillna("")
            )

    # ACRONIMO_SUP: armonizar SUPERVISOR_LIDER contra supervisores Base cupos
    if "SUPERVISOR_LIDER" in out.columns:
        out["ACRONIMO_SUP"] = out["SUPERVISOR_LIDER"].apply(
            lambda s: _resolver_acr_supervisor(s, nombres_sups_bc, idx)
        )
    else:
        out["ACRONIMO_SUP"] = ""

    return out


def _enriquecer_acronimo_supervisor(
    df_calc: pd.DataFrame,
    idx: dict,
    nombres_sups_bc: set,
) -> pd.DataFrame:
    """
    Para los DataFrames "supervisor side" (cif_por_supervisor,
    calcular_modulo_simple._sup), donde NOMBRE es el nombre del supervisor.
    Resuelve ACRONIMO mediante armonización contra Base cupos.
    """
    if df_calc.empty:
        out = df_calc.copy()
        out["ACRONIMO"] = ""
        return out
    out = df_calc.copy()
    out["ACRONIMO"] = out["NOMBRE"].apply(
        lambda s: _resolver_acr_supervisor(s, nombres_sups_bc, idx)
    )
    return out


# ─────────────────────────────────────────────────────────────────────────────
# A.5) CARGA DE KPIS V3 (Sprint 10 - Adaptado a Nube)
# ─────────────────────────────────────────────────────────────────────────────
# Los ETLs ahora producen *_KPIS.xlsx con los cumplimientos por gestor
# ya pre-calculados (Sprints 7/8/9). Esta función reemplaza el cálculo
# que antes hacía calcular_cumplimientos a partir del detalle crudo.
#
# Formato de salida (consumido por ensamblar_detalle):
#   • df_cif_gest    → ACRONIMO, NOMBRE, SUPERVISOR_LIDER, MES, AÑO,
#                      CIF_COB_%, CIF_FREC_%, CIF_HRS_%, CIF_%,
#                      N_PDVS, VENTA_TOTAL
#   • df_np_gest     → ACRONIMO, NOMBRE, NP_%
#   • df_pr_gest     → ACRONIMO, NOMBRE, PRECIOS_%
#   • df_sos_gest    → ACRONIMO, NOMBRE, SOS_%
#
# La distinción "gestor vs supervisor" se preserva en `ensamblar_detalle`
# mediante el universo de Base cupos. Los KPIs V3 no separan ambos —
# cada persona en KPIS_CIF.xlsx se evalúa por su propio detalle.

def cargar_kpis_v3(
    idx_bc: dict,
    nombres_sups_bc: set,
    mes_filtro: int | None = None,
    anio_filtro: int | None = None,
) -> dict:
    """
    Lee los archivos *_KPIS.xlsx (CIF/SOS/NP/Precios/Exhibiciones) generados
    por los ETLs y los devuelve en el formato que espera ensamblar_detalle.
    Filtra opcionalmente por (mes_filtro, anio_filtro).

    Los ETLs suben estos archivos *_KPIS.xlsx directamente a SharePoint
    (SALIDAS/<módulo>/), así que se leen SIEMPRE de ahí — nunca de disco
    local. Antes se intentaba primero un archivo local (paths.CIF_OUT_KPIS,
    etc.) y solo se caía a SharePoint si no existía; eso causaba que, si
    quedaba un archivo local viejo de una corrida anterior en el PC, este
    paso lo usara en vez de la versión más reciente de la nube, sin avisar
    que estaba usando datos desactualizados. Ahora se ignora el disco local
    por completo, para no arriesgarse a eso.
    """
    def _read_kpi(ruta_local, nombre_kpi: str, ruta_cloud: str | None = None,
                  alternativos: list[str] | None = None) -> pd.DataFrame:
        """
        Lee el KPI desde SharePoint.

        El nombre CANÓNICO es el que declara `paths.py` (`<MOD>_OUT_KPIS.name`)
        y se prueba primero. `alternativos` es solo una red de seguridad: si el
        ETL de origen llegara a subir el archivo con otro nombre, aquí se
        detecta y se avisa, en vez de lo que pasaba antes — `_leer()` devuelve
        un DataFrame vacío ante un 404, así que el KPI quedaba en blanco para
        TODA la gente sin que nada fallara ni quedara registrado.
        """
        if ruta_cloud:
            df = _leer(ruta_cloud, nombre_kpi)
            if not df.empty:
                return df
            carpeta  = ruta_cloud.rsplit("/", 1)[0]
            canonico = ruta_cloud.rsplit("/", 1)[-1]
            for alt in (alternativos or []):
                if alt == canonico:
                    continue
                df = _leer(f"{carpeta}/{alt}", f"{nombre_kpi} (alterno)")
                if not df.empty:
                    print(f"    ⚠️  [{nombre_kpi}] No existe '{canonico}' (el nombre canónico de "
                          f"paths.py), pero SÍ existe '{alt}'. Se usó el segundo para no perder "
                          f"el KPI. Unifica el nombre: el ETL de origen y paths.py deben coincidir.")
                    _log.warning("KPI %s: canónico=%s ausente, se usó alterno=%s",
                                 nombre_kpi, canonico, alt)
                    return df

        print(f"    ⚠️  KPI [{nombre_kpi}] No se encontró en SharePoint (¿ya corriste el ETL de este módulo para el periodo?): {ruta_cloud}")
        return pd.DataFrame()

    def _filtrar_periodo(df, col_mes="MES", col_anio="AÑO"):
        if df.empty or col_mes not in df.columns or col_anio not in df.columns:
            return df
        df = df.copy()
        df[col_mes]  = pd.to_numeric(df[col_mes],  errors="coerce").fillna(0).astype(int)
        df[col_anio] = pd.to_numeric(df[col_anio], errors="coerce").fillna(0).astype(int)
        if mes_filtro is not None:
            df = df[df[col_mes] == int(mes_filtro)]
        if anio_filtro is not None:
            df = df[df[col_anio] == int(anio_filtro)]
        return df

    # ── CIF (V3): COBERTURA / INTENSIDAD / FRECUENCIA / TOTAL ────────────
    ruta_cif_kpis = paths.CIF_OUT_KPIS
    df_cif_kpi = _filtrar_periodo(_read_kpi(
        ruta_cif_kpis, "CIF",
        ruta_cloud=f"{paths.RUTA_CARPETA_SALIDAS_CIF}/{ruta_cif_kpis.name}",
        alternativos=["KPIS_CIF.xlsx", "CIF_KPIS.xlsx"],
    ))
    if not df_cif_kpi.empty:
        df_cif_kpi = df_cif_kpi.rename(columns={
            "CUMPLIMIENTO COBERTURA":  "CIF_COB_%",
            "CUMPLIMIENTO INTENSIDAD": "CIF_HRS_%",
            "CUMPLIMIENTO FRECUENCIA": "CIF_FREC_%",
            "TOTAL":                   "CIF_%",
        })
        df_cif_kpi["NOMBRE"] = df_cif_kpi["NOMBRE"].astype(str).str.strip().str.upper()
    df_cif_gest = df_cif_kpi.copy()

    # ── SOS ──────────────────────────────────────────────────────────────
    ruta_sos_kpis = paths.SOS_OUT_KPIS
    df_sos_kpi = _filtrar_periodo(_read_kpi(
        ruta_sos_kpis, "SOS",
        ruta_cloud=f"{paths.RUTA_CARPETA_SALIDAS_SOS}/{ruta_sos_kpis.name}",
        alternativos=["SOS_KPIS.xlsx", "KPIS_SOS.xlsx"],
    ))
    if not df_sos_kpi.empty:
        df_sos_kpi = df_sos_kpi.rename(columns={"CUMPLIMIENTO": "SOS_%"})
        df_sos_kpi["NOMBRE"] = df_sos_kpi["NOMBRE"].astype(str).str.strip().str.upper()
    df_sos_gest = df_sos_kpi.copy()

    # ── NP (col EJECUCION) ───────────────────────────────────────────────
    ruta_np_kpis = paths.NP_OUT_KPIS
    df_np_kpi = _filtrar_periodo(_read_kpi(
        ruta_np_kpis, "NP",
        ruta_cloud=f"{paths.RUTA_CARPETA_SALIDAS_NP}/{ruta_np_kpis.name}",
        alternativos=["NP_KPIS.xlsx", "NO_PRESENCIA_KPIS.xlsx"],
    ))
    if not df_np_kpi.empty:
        df_np_kpi = df_np_kpi.rename(columns={"EJECUCION": "NP_%"})
        df_np_kpi["NOMBRE"] = df_np_kpi["NOMBRE"].astype(str).str.strip().str.upper()
    df_np_gest = df_np_kpi.copy()

    # ── PRECIOS ──────────────────────────────────────────────────────────
    ruta_pr_kpis = paths.PR_OUT_KPIS
    df_pr_kpi = _filtrar_periodo(_read_kpi(
        ruta_pr_kpis, "PRECIOS",
        ruta_cloud=f"{paths.RUTA_CARPETA_SALIDAS_PRECIOS}/{ruta_pr_kpis.name}",
        alternativos=["PRECIOS_KPIS.xlsx", "KPIS_PRECIOS.xlsx"],
    ))
    if not df_pr_kpi.empty:
        df_pr_kpi = df_pr_kpi.rename(columns={"CUMPLIMIENTO": "PRECIOS_%"})
        df_pr_kpi["NOMBRE"] = df_pr_kpi["NOMBRE"].astype(str).str.strip().str.upper()
    df_pr_gest = df_pr_kpi.copy()

    # ── EXHIBICIONES PAGADAS (V3 — Sprint 9 / Sprint 17 captura) ─────────
    ruta_exp_kpis = paths.EXHIB_PAG_OUT_KPIS
    df_exp_kpi = _filtrar_periodo(_read_kpi(
        ruta_exp_kpis, "EXHIB_PAG",
        ruta_cloud=f"{paths.RUTA_CARPETA_SALIDAS_EXHIB}/{ruta_exp_kpis.name}",
        alternativos=["EXHIBICIONES_PAGADAS_KPIS.xlsx", "KPI_Exhibiciones_Pagadas.xlsx"],
    ))
    if not df_exp_kpi.empty:
        df_exp_kpi = df_exp_kpi.rename(columns={
            "CUMPLIMIENTO":        "EXHIB_PAG_%",
            "CUMPLIMIENTO_CAPTURA": "EXHIB_PAG_CAPTURA_%",
            "EMPLEADO":            "NOMBRE",
        })
        df_exp_kpi["NOMBRE"] = df_exp_kpi["NOMBRE"].astype(str).str.strip().str.upper()
    df_exp_gest = df_exp_kpi.copy()

    # ── EXHIBICIONES GRATIS (V3 — Sprint 9 / Sprint 17 cumplimiento) ────
    ruta_egr_kpis = paths.EXHIB_GRA_OUT_KPIS
    df_egr_kpi = _filtrar_periodo(_read_kpi(
        ruta_egr_kpis, "EXHIB_GRA",
        ruta_cloud=f"{paths.RUTA_CARPETA_SALIDAS_EXHIB}/{ruta_egr_kpis.name}",
        alternativos=["KPI_Exhibiciones_Gratis.xlsx", "EXHIBICIONES_GRATIS_KPIS.xlsx"],
    ), col_mes="Mes", col_anio="Año")
    if not df_egr_kpi.empty:
        df_egr_kpi = df_egr_kpi.rename(columns={
            "Mes":           "MES",
            "Año":           "AÑO",
            "Empleado":      "NOMBRE",
            "ALTO IMPACTO":  "EXHIB_GRATIS_ALTO",
            "MEDIO IMPACTO": "EXHIB_GRATIS_MEDIO",
            "TOTAL":         "EXHIB_GRATIS_TOTAL",
            "CUMP_ALTO":     "EXHIB_GRA_ALTO_%",
            "CUMP_MEDIO":    "EXHIB_GRA_MEDIO_%",
        })
        df_egr_kpi["NOMBRE"] = df_egr_kpi["NOMBRE"].astype(str).str.strip().str.upper()
        # FIX: CUMP_ALTO/CUMP_MEDIO vienen del ETL como ALTO/TARGET sin techo,
        # así que pueden valer 6.375 (=637%) o 30.0 (=3000%). Al entrar como
        # "_%" al CUMPL_GLOBAL inflaban el global de forma masiva. Se capan a
        # 1.0 igual que ya hace _filas_gdd_y_lider() y construir_equipo_lider()
        # para el HTML — así el detalle, el resumen y el correo coinciden.
        # NOTA: el sobrecumplimiento de CIF_% (llega a ~110%) SÍ es válido
        # según el criterio de negocio, así que ese KPI NO se capa. El cap de
        # abajo aplica solo a exhibiciones gratis, donde los valores llegaban
        # a 3.000% por ser una división sin techo.
        for _c in ("EXHIB_GRA_ALTO_%", "EXHIB_GRA_MEDIO_%"):
            if _c in df_egr_kpi.columns:
                _v = pd.to_numeric(df_egr_kpi[_c], errors="coerce")
                n_cap = int((_v > 1.0).sum())
                if n_cap:
                    print(f"    ℹ️  {_c}: {n_cap} fila(s) con sobrecumplimiento "
                          f"(máx {_v.max():.2f}) capadas a 100%.")
                df_egr_kpi[_c] = _v.clip(upper=1.0)
    df_egr_gest = df_egr_kpi.copy()

    # ── Enriquecer cada uno con ACRONIMO usando Base cupos ────────────────
    nombre_a_acr = idx_bc.get("nombre_a_acronimo", {})
    cedula_a_acr = idx_bc.get("cedula_a_acronimo", {})
    candidatos_faltantes: list[dict] = []

    def _add_acronimo(df: pd.DataFrame) -> pd.DataFrame:
        if df.empty:
            df["ACRONIMO"] = ""
            return df
        df = df.copy()

        # 1) CEDULA primero — es el identificador más confiable (no le
        #    afectan tildes, nombres truncados ni orden de apellidos). Solo
        #    algunos módulos (CIF, NP, Precios, SOS) traen CEDULA; Exhib.
        #    Pagadas/Gratis no, porque su fuente (encuestas Involves) nunca
        #    tuvo ese dato — para esos dos, se salta directo al paso 2.
        if "CEDULA" in df.columns and cedula_a_acr:
            cedula_norm = df["CEDULA"].astype(str).str.strip().str.replace(r"\.0$", "", regex=True)
            df["ACRONIMO"] = cedula_norm.map(cedula_a_acr).fillna("")
        else:
            df["ACRONIMO"] = ""

        # 2) Nombre exacto, para lo que no resolvió por cédula.
        falta_cedula = df["ACRONIMO"] == ""
        if falta_cedula.any():
            df.loc[falta_cedula, "ACRONIMO"] = df.loc[falta_cedula, "NOMBRE"].map(nombre_a_acr).fillna("")

        falta = df["ACRONIMO"] == ""
        if falta.any():
            n_antes = falta.sum()
            df.loc[falta, "ACRONIMO"] = df.loc[falta, "NOMBRE"].apply(
                lambda n: _resolver_acr_fuzzy_general(n, idx_bc)
            )
            n_resueltos = (df.loc[falta, "ACRONIMO"] != "").sum()
            n_sin_resolver = n_antes - n_resueltos
            if n_sin_resolver > 0:
                aun_falta = falta & (df["ACRONIMO"] == "")
                nombres_sin_resolver = df.loc[aun_falta, "NOMBRE"].unique().tolist()
                # El listado completo se imprime solo la primera vez; en los
                # módulos siguientes se resume, porque son casi siempre las
                # MISMAS personas y repetir la lista entera 6 veces (una por
                # módulo) no aporta información nueva.
                nuevos = [n for n in nombres_sin_resolver
                          if str(n).strip().upper() not in _AVISADOS_SIN_ACRONIMO]
                if nuevos:
                    print(f"    ⚠️  {n_sin_resolver} persona(s) sin ACRONIMO, se excluyen del cruce "
                          f"({len(nombres_sin_resolver) - len(nuevos)} ya reportada(s) antes): {nuevos}")
                else:
                    print(f"    ⚠️  {n_sin_resolver} persona(s) sin ACRONIMO, se excluyen del cruce "
                          f"(todas ya reportadas arriba).")
                # FIX: separar los dos diagnósticos. Una persona puede ESTAR en
                # Base cupos y aun así no resolver, si su fila no tiene ACRONIMO
                # (típico de las filas que inserta --agregar-faltantes y nadie
                # completó). Antes el mensaje decía "probablemente no está en
                # Base cupos", que es falso y manda a buscar el problema donde
                # no está.
                nombres_bc_todos    = set(idx_bc.get("nombre_a_acronimo", {}).keys())
                nombres_bc_sin_acr  = set(idx_bc.get("nombres_sin_acronimo", set()))
                # FIX 2: `base_cupos.py` DESCARTA las filas sin ACRONIMO antes
                # de construir los índices (en la corrida real: "Se descartaron
                # 10 filas por no tener ACRONIMO o NOMBRE"). Por eso el idx no
                # las conoce y este diagnóstico caía siempre en el "no está en
                # Base cupos", que es engañoso: la persona SÍ está en el Excel,
                # solo le falta llenar la celda. Se lee Base_cupos.xlsx crudo
                # (una sola vez por corrida) para dar el mensaje correcto.
                # Si `base_cupos.py` ya expone las filas sin ACRONIMO (versión
                # nueva), se usan y NO se vuelve a descargar el Excel. La
                # relectura queda solo como respaldo para versiones viejas del
                # módulo que descartan esas filas sin informarlas.
                ced_sin_acr_crudo = set(idx_bc.get("cedulas_sin_acronimo", set()))
                if not nombres_bc_sin_acr and not ced_sin_acr_crudo:
                    sin_acr_crudo, ced_sin_acr_crudo = _base_cupos_sin_acronimo()
                    nombres_bc_sin_acr |= sin_acr_crudo

                for n in nombres_sin_resolver:
                    n_up = str(n).strip().upper()
                    # También se compara por cédula: el nombre puede venir
                    # escrito distinto entre el KPI y Base cupos.
                    ced_n = ""
                    if "CEDULA" in df.columns:
                        _f = df.loc[aun_falta & (df["NOMBRE"] == n), "CEDULA"]
                        if not _f.empty:
                            ced_n = str(_f.iloc[0]).strip().replace(".0", "")
                    if (n_up in nombres_bc_sin_acr
                            or (ced_n and ced_n in ced_sin_acr_crudo)
                            or (n_up in nombres_bc_todos
                                and not str(idx_bc["nombre_a_acronimo"].get(n_up, "")).strip())):
                        # Se avisa una sola vez por persona en toda la corrida:
                        # esta rama se ejecuta una vez por MÓDULO (CIF, SOS, NP,
                        # Precios, Exh Pag, Exh Gratis), así que sin este guard
                        # la misma persona aparecía hasta 6 veces con el mismo
                        # texto largo, ahogando el log.
                        if n_up not in _AVISADOS_SIN_ACRONIMO:
                            _AVISADOS_SIN_ACRONIMO.add(n_up)
                            print(f"       · '{n}' — ⚠️ SÍ ESTÁ en Base cupos (cédula {ced_n or 's/d'}), "
                                  f"pero su fila NO TIENE ACRONIMO. Complétalo en Base_cupos.xlsx "
                                  f"(ACRONIMO + ROL EN ACTIVO); hasta entonces se excluye del cruce.")
                        continue
                    # Los que de verdad NO están en Base cupos también se
                    # avisan una sola vez (antes WILMER VERA ARANGO y los
                    # demás se repetían en cada módulo).
                    if n_up in _AVISADOS_SIN_ACRONIMO:
                        continue
                    _AVISADOS_SIN_ACRONIMO.add(n_up)
                    candidatos = _candidatos_parecidos(n, idx_bc, top=3)
                    if candidatos:
                        print(f"       · '{n}' — NO está en Base cupos. Más parecido(s): {candidatos}")
                    else:
                        print(f"       · '{n}' — NINGÚN nombre de Base cupos comparte ni una palabra con este "
                              f"(probablemente no está en Base cupos, no es solo un tema de tildes)")
                # Recolectar nombre + cédula (si la trae el módulo) para el
                # reporte de "candidatos a agregar a Base cupos".
                filas_sin_resolver = df.loc[aun_falta]
                for n in nombres_sin_resolver:
                    fila_n = filas_sin_resolver[filas_sin_resolver["NOMBRE"] == n]
                    cedula_val = ""
                    if "CEDULA" in fila_n.columns:
                        ceds = (
                            fila_n["CEDULA"].astype(str).str.strip()
                            .replace({"": None, "nan": None, "NAN": None})
                            .dropna().unique().tolist()
                        )
                        cedula_val = ceds[0] if ceds else ""
                    candidatos_faltantes.append({"NOMBRE": n, "CEDULA": cedula_val})
        return df

    df_cif_gest = _add_acronimo(df_cif_gest)
    df_sos_gest = _add_acronimo(df_sos_gest)
    df_np_gest  = _add_acronimo(df_np_gest)
    df_pr_gest  = _add_acronimo(df_pr_gest)
    df_exp_gest = _add_acronimo(df_exp_gest)
    df_egr_gest = _add_acronimo(df_egr_gest)

    # ── ACRONIMO_SUP: armonizar SUPERVISOR_LIDER vs Base cupos ───────────
    def _add_acr_sup(df: pd.DataFrame) -> pd.DataFrame:
        if df.empty or "SUPERVISOR_LIDER" not in df.columns:
            if not df.empty:
                df["ACRONIMO_SUP"] = ""
            return df
        df = df.copy()
        df["ACRONIMO_SUP"] = df["SUPERVISOR_LIDER"].apply(
            lambda s: _resolver_acr_supervisor(s, nombres_sups_bc, idx_bc)
        )
        return df

    df_cif_gest = _add_acr_sup(df_cif_gest)
    df_sos_gest = _add_acr_sup(df_sos_gest)
    df_np_gest  = _add_acr_sup(df_np_gest)
    df_pr_gest  = _add_acr_sup(df_pr_gest)
    
    acr_a_sup: dict = {}
    for src in (df_cif_gest, df_np_gest, df_pr_gest, df_sos_gest):
        if src is None or src.empty:
            continue
        if "ACRONIMO" not in src.columns or "ACRONIMO_SUP" not in src.columns:
            continue
        for _, r in src[["ACRONIMO", "ACRONIMO_SUP"]].iterrows():
            acr, asu = r["ACRONIMO"], r["ACRONIMO_SUP"]
            if acr and asu and acr not in acr_a_sup:
                acr_a_sup[acr] = asu

    def _add_acr_sup_via_acr(df: pd.DataFrame) -> pd.DataFrame:
        if df.empty:
            return df
        df = df.copy()
        df["ACRONIMO_SUP"] = df["ACRONIMO"].map(acr_a_sup).fillna("")
        return df

    df_exp_gest = _add_acr_sup_via_acr(df_exp_gest)
    df_egr_gest = _add_acr_sup_via_acr(df_egr_gest)

    return {
        "cif_gest": df_cif_gest,
        "sos_gest": df_sos_gest,
        "np_gest":  df_np_gest,
        "pr_gest":  df_pr_gest,
        "exp_gest": df_exp_gest,   
        "egr_gest": df_egr_gest,   
        "candidatos_faltantes": candidatos_faltantes,
    }


# ─────────────────────────────────────────────────────────────────────────────
# A) GENERAR / VERIFICAR MAESTRA DE SUPERVISORES (Adaptado a Nube)
# ─────────────────────────────────────────────────────────────────────────────

def generar_maestra(df_np: pd.DataFrame, df_pr: pd.DataFrame, df_sos: pd.DataFrame) -> None:
    """
    Genera o actualiza MAESTRO_SUPERVISORES.xlsx desde Base cupos en la nube.

    FIX: creaba la carpeta local de ALERTAS aunque la maestra se lee y se
    escribe 100% en SharePoint.
    """
    df_bc   = bcm.cargar()
    sups_bc = bcm.supervisores(df_bc)
    # Destinatarios de la maestra = supervisores + GDD + líderes de ejecución
    # (los 37 destinatarios reales de Telegram/Email, no solo supervisores —
    # antes un GDD/Líder nuevo en Base cupos nunca recibía fila aquí y quedaba
    # sin CORREO/TELEGRAM_CHAT_ID resoluble, en silencio).
    nombres_destinatarios = set(sups_bc["NOMBRE"].dropna().astype(str).str.upper())
    if "ES_GDD" in df_bc.columns:
        nombres_destinatarios.update(
            df_bc[df_bc["ES_GDD"] == True]["NOMBRE"].dropna().astype(str).str.upper()
        )
    if "ES_LIDER" in df_bc.columns:
        nombres_destinatarios.update(
            df_bc[df_bc["ES_LIDER"] == True]["NOMBRE"].dropna().astype(str).str.upper()
        )
    nombres_bc = sorted(nombres_destinatarios)

    existente = pd.DataFrame(columns=["NOMBRE_SUPERVISOR", "CORREO", "TELEGRAM_CHAT_ID"])

    # Ruta de la maestra en SharePoint — definida SIEMPRE antes del try para
    # que esté disponible más abajo incluso si la lectura falla.
    ruta_maestra_cloud = RUTA_MAESTRA_CLOUD

    # Lectura de la maestra desde SharePoint en lugar de disco local
    try:
        print(f"    ⏳ Leyendo maestra de supervisores desde SharePoint...")
        token = _obtener_token_azure()
        headers = {"Authorization": f"Bearer {token}"}
        drive_id = _obtener_default_drive_id(token)
        ruta_limpia = _limpiar_ruta_graph(ruta_maestra_cloud)
        
        url_codificada = urllib.parse.quote(ruta_limpia)
        url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/root:/{url_codificada}:/content"
        
        response = requests.get(url, headers=headers)
        if response.status_code == 200:
            existente = pd.read_excel(io.BytesIO(response.content))
            existente.columns = existente.columns.str.strip().str.upper()
            existente["NOMBRE_SUPERVISOR"] = (
                existente["NOMBRE_SUPERVISOR"].astype(str).str.strip().str.upper()
            )
            existente = existente[
                existente["NOMBRE_SUPERVISOR"].notna()
                & (existente["NOMBRE_SUPERVISOR"] != "")
                & (existente["NOMBRE_SUPERVISOR"] != "NAN")
                & (~existente["NOMBRE_SUPERVISOR"].str.contains(r"[⚠✓✗]", na=False))
                & (~existente["NOMBRE_SUPERVISOR"].str.contains(r"^COMPLETA", na=False))
            ].reset_index(drop=True)
    except Exception as e:
        print(f"    ⚠️  No pude leer la maestra existente desde la nube ({e}); recreando")
        existente = pd.DataFrame(columns=["NOMBRE_SUPERVISOR", "CORREO", "TELEGRAM_CHAT_ID"])

    nombres_existentes = set(existente["NOMBRE_SUPERVISOR"]) if not existente.empty else set()
    nombres_nuevos = [n for n in nombres_bc if n not in nombres_existentes]
    nombres_obsoletos = [
        n for n in nombres_existentes
        if n and n != "NAN" and n not in set(nombres_bc)
    ]

    try:
        crudo_n = len(existente)
    except Exception:
        crudo_n = len(existente)
    hay_basura = crudo_n > len(existente)

    if not nombres_nuevos and not hay_basura and not existente.empty:
        if nombres_obsoletos:
            print(
                f"  ⚠️  {len(nombres_obsoletos)} destinatario(s) en la maestra ya no "
                f"figuran como activos en Base cupos:"
            )
            for n in nombres_obsoletos[:5]:
                print(f"     · {n}")
            print(f"     (Se conservan en la maestra; revisa si fueron desactivados.)")
        print(f"  ✓ Maestra al día: {len(existente)} destinatarios (supervisores + GDD + líderes)")
        return

    if not existente.empty:
        existente = existente[["NOMBRE_SUPERVISOR", "CORREO", "TELEGRAM_CHAT_ID"]].copy()
        nuevos_df = pd.DataFrame({
            "NOMBRE_SUPERVISOR": nombres_nuevos,
            "CORREO":            [""] * len(nombres_nuevos),
            "TELEGRAM_CHAT_ID":  [""] * len(nombres_nuevos),
        })
        df_maestra = pd.concat([existente, nuevos_df], ignore_index=True)
        df_maestra = df_maestra.sort_values("NOMBRE_SUPERVISOR").reset_index(drop=True)
        accion = (
            f"✏️  Maestra actualizada: {len(existente)} existentes + "
            f"{len(nombres_nuevos)} nuevos (correo/chat_id vacíos)"
        )
    else:
        df_maestra = pd.DataFrame({
            "NOMBRE_SUPERVISOR": nombres_bc,
            "CORREO":            [""] * len(nombres_bc),
            "TELEGRAM_CHAT_ID":  [""] * len(nombres_bc),
        })
        accion = f"📋 Maestra creada con {len(nombres_bc)} destinatarios (supervisores + GDD + líderes)"

    print(f"\n  {accion}")
    if nombres_nuevos:
        print(f"     Nuevos:")
        for n in nombres_nuevos[:10]:
            print(f"       · {n}")
        if len(nombres_nuevos) > 10:
            print(f"       · ... ({len(nombres_nuevos) - 10} más)")

    wb = Workbook()
    ws = wb.active
    ws.title = "Supervisores"

    headers = ["NOMBRE_SUPERVISOR", "CORREO", "TELEGRAM_CHAT_ID"]
    anchos  = [40, 40, 20]

    hdr_font  = Font(name="Arial", bold=True, color=COLOR_HEADER_FNT, size=11)
    hdr_fill  = PatternFill("solid", fgColor=COLOR_MAESTRA_HDR)
    hdr_align = Alignment(horizontal="center", vertical="center")
    thin      = Side(style="thin", color="CCCCCC")
    borde     = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, (hdr, ancho) in enumerate(zip(headers, anchos), start=1):
        cell = ws.cell(row=1, column=col_idx, value=hdr)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = hdr_align
        cell.border    = borde
        ws.column_dimensions[get_column_letter(col_idx)].width = ancho
    ws.row_dimensions[1].height = 22

    fill_par = PatternFill("solid", fgColor=COLOR_FILA_PAR)
    for row_idx, row in enumerate(df_maestra.itertuples(index=False), start=2):
        fill = fill_par if row_idx % 2 == 0 else PatternFill()
        for col_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font      = Font(name="Arial", size=10)
            cell.fill      = fill
            cell.alignment = Alignment(vertical="center")
            cell.border    = borde

    ws.freeze_panes = "A2"
    # El nombre de una pestaña de Excel tiene un límite DURO de 31
    # caracteres — "Supervisores - Completa CORREO y TELEGRAM_CHAT_ID"
    # (50 caracteres) lo excede y produce un .xlsx inválido que Excel
    # reporta como dañado/corrupto al abrirlo. Se deja el título corto y
    # la instrucción se agrega como nota en una celda en vez de en el título.
    ws.cell(row=1, column=5,
            value="👉 Completa CORREO y TELEGRAM_CHAT_ID para los supervisores nuevos")

    # Guardar en la ruta local temporal y luego subir a SharePoint (o guardar directo si se usa BytesIO)
    output_io = io.BytesIO()
    wb.save(output_io)
    output_io.seek(0)
    
    # Subida a SharePoint usando Graph API
    try:
        token = _obtener_token_azure()
        headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}
        drive_id = _obtener_default_drive_id(token)
        ruta_limpia = _limpiar_ruta_graph(ruta_maestra_cloud)
        
        url_codificada = urllib.parse.quote(ruta_limpia)
        url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/root:/{url_codificada}:/content"
        
        response = requests.put(url, headers=headers, data=output_io.getvalue())
        if response.status_code in [200, 201]:
            print(f"  ✅ Maestra guardada en SharePoint con {len(df_maestra)} supervisores: {ruta_maestra_cloud}")
        else:
            print(f"  ⚠️ Error al subir la maestra a SharePoint: {response.status_code} - {response.text}")
    except Exception as e:
        print(f"  ⚠️ Excepción al subir la maestra a SharePoint: {e}")

    if nombres_nuevos:
        print(f"     → Completa CORREO y TELEGRAM_CHAT_ID para los {len(nombres_nuevos)} nuevos.\n")


# ─────────────────────────────────────────────────────────────────────────────
# B) CÁLCULO DE KPIS  →  movido a los ETLs (V3, Sprint 7/8/9).
#    Antes vivía aquí (calcular_cif_componentes_por_pdv, _cif_agregar_por_clave,
#    cif_por_gestor, cif_por_supervisor, calcular_modulo_simple) — ~310 líneas.
#    Hoy cada ETL escribe su <MOD>_KPIS.xlsx y este módulo solo los ensambla
#    via cargar_kpis_v3() (definido arriba en la sección A.5).
# ─────────────────────────────────────────────────────────────────────────────


# ─────────────────────────────────────────────────────────────────────────────
# C) ENSAMBLAJE FINAL — df_detalle y df_resumen
# ─────────────────────────────────────────────────────────────────────────────

def ensamblar_detalle(
    universo: pd.DataFrame,
    df_cif_gest: pd.DataFrame,
    df_cif_sup:  pd.DataFrame,
    df_np_gest:  pd.DataFrame, df_np_sup:  pd.DataFrame,
    df_pr_gest:  pd.DataFrame, df_pr_sup:  pd.DataFrame,
    df_sos_gest: pd.DataFrame, df_sos_sup: pd.DataFrame,
    df_dyp_gest: pd.DataFrame | None = None,
    df_dyp_sup:  pd.DataFrame | None = None,
    df_exp_gest: pd.DataFrame | None = None,
    df_egr_gest: pd.DataFrame | None = None,
    mes: int | None = None,
    anio: int | None = None,
) -> pd.DataFrame:
    """
    Una fila por persona del UNIVERSO (Base cupos), keyed por ACRONIMO.

    Columnas:
      ACRONIMO, NOMBRE, CEDULA, ROL, CANAL, ES_SUPERVISOR,
      ACRONIMO_SUP, NOMBRE_SUPERVISOR,
      N_PDVS, VENTA_TOTAL,
      CIF_COB_%, CIF_FREC_%, CIF_HRS_%, CIF_%,
      NP_%, PRECIOS_%, SOS_%,
      VENTA_%, IMPACTOS_%, MSL_%, PROD_NUEVOS_%,
      MES, AÑO

    Reglas de fusión:
      • Universo = `bcm.universo_personas(df_bc)` (200 personas activas).
      • Para cada DataFrame de cálculo (gestor/supervisor), merge por ACRONIMO.
      • `ACRONIMO_SUP` por persona se toma con prioridad: D&P → CIF →
        otros (NP/Precios/SOS), reflejando que D&P es la fuente más
        completa para transferencistas puros.
      • `NOMBRE_SUPERVISOR` se deriva del ACRONIMO_SUP cruzando con el
        universo.
    """
    # ── BASE: universo de personas (de Base cupos) ───────────────────────
    base = universo.copy()
    # Renombrar para usar nombres canónicos del PT en el output
    # (NOMBRE ya es UPPER por bcm.cargar)
    if mes is not None:
        base["MES"] = mes
    if anio is not None:
        base["AÑO"] = anio

    # Helper local: agregar/upsert columnas desde un df_calc cruzando por ACRONIMO
    def _merge_por_acr(target, df_calc, cols_traer):
        if df_calc is None or df_calc.empty:
            for c in cols_traer:
                if c not in target.columns:
                    target[c] = np.nan
            return target
        cols_disp = [c for c in cols_traer if c in df_calc.columns]
        if not cols_disp or "ACRONIMO" not in df_calc.columns:
            for c in cols_traer:
                if c not in target.columns:
                    target[c] = np.nan
            return target
        cand = df_calc[df_calc["ACRONIMO"].astype(str).str.strip() != ""]
        # FIX: antes se hacía drop_duplicates a ciegas y se perdía el dato de
        # la 2a fila sin avisar (caso real: una persona con dos supervisores
        # reportada 2 veces en SOS/NP — su cumplimiento salía a la mitad).
        # Ahora se denuncia y, si son valores numéricos, se promedian en vez
        # de descartarse.
        dups = cand[cand.duplicated("ACRONIMO", keep=False)]
        if not dups.empty:
            for acr, grupo in dups.groupby("ACRONIMO"):
                nombres = grupo["NOMBRE"].astype(str).unique().tolist() if "NOMBRE" in grupo.columns else []
                msg = (f"ACRONIMO {acr} aparece {len(grupo)} veces en el KPI de origen "
                       f"{cols_disp} — se consolida promediando. Nombre(s): {nombres}")
                print(f"    ⚠️  {msg}")
                _log.warning(msg)
            num_cols = [c for c in cols_disp
                        if pd.api.types.is_numeric_dtype(cand[c])]
            no_num   = [c for c in cols_disp if c not in num_cols]
            agg_map  = {c: "mean" for c in num_cols}
            agg_map.update({c: "first" for c in no_num})
            merged = cand[["ACRONIMO"] + cols_disp].groupby("ACRONIMO", as_index=False).agg(agg_map)
        else:
            merged = cand[["ACRONIMO"] + cols_disp].drop_duplicates(subset=["ACRONIMO"])
        return target.merge(merged, on="ACRONIMO", how="left")

    # ── CIF (gestor + supervisor) ────────────────────────────────────────
    cif_g_ren = pd.DataFrame()
    if not df_cif_gest.empty:
        cif_g_ren = df_cif_gest.rename(columns={
            "COB":  "CIF_COB_%", "FREC": "CIF_FREC_%",
            "HRS":  "CIF_HRS_%", "CIF":  "CIF_%",
        })
    cif_s_ren = pd.DataFrame()
    if not df_cif_sup.empty:
        cif_s_ren = df_cif_sup.rename(columns={
            "COB":  "CIF_COB_%", "FREC": "CIF_FREC_%",
            "HRS":  "CIF_HRS_%", "CIF":  "CIF_%",
        })
    # Concatenar gestor + supervisor (cada persona aparece en uno u otro)
    cif_unif = pd.concat([cif_g_ren, cif_s_ren], ignore_index=True) \
                  if (not cif_g_ren.empty or not cif_s_ren.empty) else pd.DataFrame()
    cols_cif = ["CIF_COB_%", "CIF_FREC_%", "CIF_HRS_%", "CIF_%",
                "N_PDVS", "VENTA_TOTAL"]
    base = _merge_por_acr(base, cif_unif, cols_cif)

    # Si N_PDVS/VENTA_TOTAL no llegaron, asegurar columnas
    for c in ["N_PDVS", "VENTA_TOTAL"]:
        if c not in base.columns:
            base[c] = np.nan

    # ── NP, Precios, SOS — unificar gestor+supervisor de cada uno ────────
    for df_g, df_s, col in [
        (df_np_gest,  df_np_sup,  "NP_%"),
        (df_pr_gest,  df_pr_sup,  "PRECIOS_%"),
        (df_sos_gest, df_sos_sup, "SOS_%"),
    ]:
        unif = pd.concat(
            [df for df in (df_g, df_s) if df is not None and not df.empty],
            ignore_index=True,
        ) if any(df is not None and not df.empty for df in (df_g, df_s)) else pd.DataFrame()
        base = _merge_por_acr(base, unif, [col])

    # ── EXHIBICIONES (Sprint 12 / Sprint 17) ─────────────────────────────
    # KPIs que entran al CUMPL_GLOBAL: EXHIB_PAG_% (ejecución), EXHIB_PAG_CAPTURA_%,
    # EXHIB_GRA_ALTO_% (vs target 3), EXHIB_GRA_MEDIO_% (vs target 5).
    # Conteos absolutos informativos (no entran al global):
    # EXHIB_GRATIS_TOTAL / ALTO / MEDIO.
    if df_exp_gest is not None:
        base = _merge_por_acr(base, df_exp_gest,
                              ["EXHIB_PAG_%", "EXHIB_PAG_CAPTURA_%"])
    if df_egr_gest is not None:
        base = _merge_por_acr(
            base, df_egr_gest,
            ["EXHIB_GRATIS_TOTAL", "EXHIB_GRATIS_ALTO", "EXHIB_GRATIS_MEDIO",
             "EXHIB_GRA_ALTO_%", "EXHIB_GRA_MEDIO_%"],
        )

    # ── D&P (gestor + supervisor) ────────────────────────────────────────
    cols_dyp = ["VENTA_%", "IMPACTOS_%", "MSL_%", "PROD_NUEVOS_%"]
    dyp_unif = pd.concat(
        [df for df in (df_dyp_gest, df_dyp_sup) if df is not None and not df.empty],
        ignore_index=True,
    ) if (df_dyp_gest is not None and not df_dyp_gest.empty) or \
         (df_dyp_sup  is not None and not df_dyp_sup.empty) else pd.DataFrame()
    base = _merge_por_acr(base, dyp_unif, cols_dyp)

    # ── ACRONIMO_SUP por persona (prioridad: D&P > CIF > otros) ──────────
    # Construimos un mapa ACRONIMO → ACRONIMO_SUP a partir de cada fuente.
    sup_map = {}

    def _acumular(df, prioridad):
        if df is None or df.empty:
            return
        if "ACRONIMO" not in df.columns or "ACRONIMO_SUP" not in df.columns:
            return
        for _, r in df.iterrows():
            acr = r.get("ACRONIMO", "")
            asu = r.get("ACRONIMO_SUP", "")
            if not acr or not asu:
                continue
            existente = sup_map.get(acr)
            if existente is None or existente[0] > prioridad:
                sup_map[acr] = (prioridad, asu)

    # Prioridad menor = más prioritario
    _acumular(df_dyp_gest, prioridad=1)        # D&P primero (gestores)
    _acumular(df_cif_gest if not df_cif_gest.empty else None, prioridad=2)
    _acumular(df_np_gest  if not df_np_gest.empty  else None, prioridad=3)
    _acumular(df_pr_gest  if not df_pr_gest.empty  else None, prioridad=4)
    _acumular(df_sos_gest if not df_sos_gest.empty else None, prioridad=5)

    base["ACRONIMO_SUP"] = base["ACRONIMO"].map(
        lambda a: sup_map[a][1] if a in sup_map else ""
    )
    # NOMBRE_SUPERVISOR derivado del universo
    nombre_por_acr = dict(zip(universo["ACRONIMO"], universo["NOMBRE"]))
    base["NOMBRE_SUPERVISOR"] = base["ACRONIMO_SUP"].map(nombre_por_acr).fillna("")

    # ── Reordenar columnas finales ───────────────────────────────────────
    cols_finales = [
        "ACRONIMO", "NOMBRE", "CEDULA", "ROL", "CANAL", "CIUDAD_CUPO",
        "ES_SUPERVISOR", "ES_GDD", "ES_LIDER",         # Sprint 17.15
        "ACRONIMO_SUP", "NOMBRE_SUPERVISOR",
        "N_PDVS", "VENTA_TOTAL",
        "CIF_COB_%", "CIF_FREC_%", "CIF_HRS_%", "CIF_%",
        "NP_%", "PRECIOS_%", "SOS_%",
        # Exhibiciones (Sprint 12 / Sprint 17)
        "EXHIB_PAG_%", "EXHIB_PAG_CAPTURA_%",
        "EXHIB_GRATIS_TOTAL", "EXHIB_GRATIS_ALTO", "EXHIB_GRATIS_MEDIO",
        "EXHIB_GRA_ALTO_%", "EXHIB_GRA_MEDIO_%",
        "VENTA_%", "IMPACTOS_%", "MSL_%", "PROD_NUEVOS_%",
        "MES", "AÑO",
    ]
    for c in cols_finales:
        if c not in base.columns:
            base[c] = np.nan
    base = base[cols_finales]

    # Orden: supervisores primero por ACRONIMO; gestores agrupados por su sup.
    base = base.sort_values(
        ["ES_SUPERVISOR", "ACRONIMO_SUP", "NOMBRE"],
        ascending=[False, True, True],
    ).reset_index(drop=True)

    return base


def calcular_resumen(df_detalle: pd.DataFrame, universo: pd.DataFrame) -> pd.DataFrame:
    """
    Una fila por supervisor:
      ACRONIMO_SUP, SUPERVISOR_LIDER (= NOMBRE del supervisor),
      N_GESTORES, VENTA_EQUIPO,
      CIF_%, NP_%, PRECIOS_%, SOS_%,
      VENTA_%, IMPACTOS_%, MSL_%, PROD_NUEVOS_%,
      CUMPL_GLOBAL_%,
      MES, AÑO

    Los KPIs del supervisor se toman de la fila propia (ES_SUPERVISOR=True)
    en df_detalle. N_GESTORES y VENTA_EQUIPO se calculan agrupando por
    ACRONIMO_SUP las filas no-supervisor.
    """
    if df_detalle.empty:
        return pd.DataFrame()

    # ── Stats del equipo: agrupar gestores por ACRONIMO_SUP ──────────────
    no_sup = df_detalle[df_detalle["ES_SUPERVISOR"] == False].copy()
    if not no_sup.empty:
        equipo_stats = (
            no_sup[no_sup["ACRONIMO_SUP"].fillna("").astype(str) != ""]
                  .groupby("ACRONIMO_SUP", dropna=False)
                  .agg(
                      N_GESTORES   = ("ACRONIMO",    "nunique"),
                      VENTA_EQUIPO = ("VENTA_TOTAL", "sum"),
                  )
                  .reset_index()
                  .rename(columns={"ACRONIMO_SUP": "ACRONIMO"})
        )
    else:
        equipo_stats = pd.DataFrame(columns=["ACRONIMO", "N_GESTORES", "VENTA_EQUIPO"])

    # ── Filas SUPERVISOR de df_detalle (sus propios KPIs) ────────────────
    sups = df_detalle[df_detalle["ES_SUPERVISOR"] == True].copy()

    # Sprint 12 / Sprint 17: entran al CUMPL_GLOBAL_% (D11=Sí):
    #   EXHIB_PAG_% (ejecución), EXHIB_PAG_CAPTURA_% (captura del planning),
    #   EXHIB_GRA_ALTO_% (vs target 3), EXHIB_GRA_MEDIO_% (vs target 5).
    # Conteos informativos (no global): EXHIB_GRATIS_TOTAL/ALTO/MEDIO.
    cols_pct = ["CIF_%", "NP_%", "PRECIOS_%", "SOS_%",
                "EXHIB_PAG_%", "EXHIB_PAG_CAPTURA_%",
                "EXHIB_GRA_ALTO_%", "EXHIB_GRA_MEDIO_%",
                "VENTA_%", "IMPACTOS_%", "MSL_%", "PROD_NUEVOS_%"]
    cols_info = ["EXHIB_GRATIS_TOTAL", "EXHIB_GRATIS_ALTO", "EXHIB_GRATIS_MEDIO"]
    cols_sup = ["ACRONIMO", "NOMBRE", "MES", "AÑO"] + cols_pct + cols_info
    sup_metrics = sups[[c for c in cols_sup if c in sups.columns]].copy()

    # ── KPI agregado por equipo (Sprint 10) ──────────────────────────────
    # Los KPIs V3 solo poblan al supervisor cuando él mismo tiene ruta
    # (típico en CIF). Para NP/Precios/SOS/D&P/EXHIB, el supervisor casi
    # nunca aparece como persona con KPI propio. Tomamos el promedio del
    # equipo de sus gestores como fallback razonable (suma para conteos
    # informativos como EXHIB_GRATIS_TOTAL).
    if not no_sup.empty:
        equipo_filtered = no_sup[no_sup["ACRONIMO_SUP"].fillna("").astype(str) != ""]
        team_kpis_pct = (
            equipo_filtered.groupby("ACRONIMO_SUP", dropna=False)[cols_pct]
                           .mean()
                           .reset_index()
        )
        cols_info_disp = [c for c in cols_info if c in no_sup.columns]
        if cols_info_disp:
            team_kpis_info = (
                equipo_filtered.groupby("ACRONIMO_SUP", dropna=False)[cols_info_disp]
                               .sum(min_count=1)
                               .reset_index()
            )
            team_kpis = team_kpis_pct.merge(team_kpis_info, on="ACRONIMO_SUP")
        else:
            team_kpis = team_kpis_pct
        team_kpis = team_kpis.rename(columns={"ACRONIMO_SUP": "ACRONIMO"})
    else:
        team_kpis = pd.DataFrame(columns=["ACRONIMO"] + cols_pct + cols_info)

    # Combinar: KPI propio del supervisor > promedio del equipo
    sup_metrics_combinado = sup_metrics.merge(
        team_kpis, on="ACRONIMO", how="outer", suffixes=("", "_eq")
    )
    # Para %s, prevalece KPI propio sobre el de equipo
    for c in cols_pct + cols_info:
        col_eq = f"{c}_eq"
        if col_eq in sup_metrics_combinado.columns:
            sup_metrics_combinado[c] = (
                sup_metrics_combinado[c].where(
                    sup_metrics_combinado[c].notna(),
                    sup_metrics_combinado[col_eq],
                )
            )
            sup_metrics_combinado = sup_metrics_combinado.drop(columns=[col_eq])
    sup_metrics = sup_metrics_combinado

    # Merge externo: para no perder supervisores sin equipo ni equipos cuyo
    # supervisor no esté en el universo.
    resumen = pd.merge(sup_metrics, equipo_stats, on="ACRONIMO", how="outer")

    # Para filas que vinieron del equipo_stats (sup sin métricas propias),
    # rellenar NOMBRE desde el universo.
    nombre_por_acr = dict(zip(universo["ACRONIMO"], universo["NOMBRE"]))
    if "NOMBRE" in resumen.columns:
        resumen["NOMBRE"] = resumen["NOMBRE"].fillna(
            resumen["ACRONIMO"].map(nombre_por_acr)
        )
    else:
        resumen["NOMBRE"] = resumen["ACRONIMO"].map(nombre_por_acr)

    # Defensivo: completar columnas faltantes
    for c in cols_pct + cols_info + ["N_GESTORES", "VENTA_EQUIPO", "MES", "AÑO"]:
        if c not in resumen.columns:
            resumen[c] = np.nan
    resumen["N_GESTORES"]   = resumen["N_GESTORES"].fillna(0).astype(int)
    resumen["VENTA_EQUIPO"] = resumen["VENTA_EQUIPO"].fillna(0.0)

    # Cumplimiento global (promedio simple de los KPIs % disponibles).
    # EXHIB_PAG_% SÍ entra (es un KPI), EXHIB_GRATIS_TOTAL NO (es conteo).
    resumen["CUMPL_GLOBAL_%"] = resumen[cols_pct].mean(axis=1, skipna=True)

    # Renombrar para output legible
    resumen = resumen.rename(columns={
        "ACRONIMO": "ACRONIMO_SUP",
        "NOMBRE":   "SUPERVISOR_LIDER",
    })

    cols_finales = [
        "ACRONIMO_SUP", "SUPERVISOR_LIDER", "N_GESTORES", "VENTA_EQUIPO",
        "CIF_%", "NP_%", "PRECIOS_%", "SOS_%",
        "EXHIB_PAG_%", "EXHIB_PAG_CAPTURA_%",
        "EXHIB_GRA_ALTO_%", "EXHIB_GRA_MEDIO_%",
        "VENTA_%", "IMPACTOS_%", "MSL_%", "PROD_NUEVOS_%",
        "EXHIB_GRATIS_TOTAL", "EXHIB_GRATIS_ALTO", "EXHIB_GRATIS_MEDIO",
        "CUMPL_GLOBAL_%",
        "MES", "AÑO",
    ]
    for c in cols_finales:
        if c not in resumen.columns:
            resumen[c] = np.nan
    resumen = resumen[cols_finales]

    # Sprint 17.15 — agregar destinatarios GDD y Líder de Ejecución.
    resumen["_ROL_DEST"] = "SUPERVISOR"
    resumen_gdd_lid = _filas_gdd_y_lider(df_detalle, universo, cols_finales)
    if not resumen_gdd_lid.empty:
        resumen = pd.concat([resumen, resumen_gdd_lid], ignore_index=True)

    return resumen.sort_values("SUPERVISOR_LIDER").reset_index(drop=True)


# ─────────────────────────────────────────────────────────────────────────────
# Sprint 17.15 — destinatarios adicionales: GDD y Líder de Ejecución
# ─────────────────────────────────────────────────────────────────────────────

# Mapping ciudad → ACRONIMO del líder (canal DIRECTO y combinados).
LIDER_POR_CIUDAD: dict[str, str] = {
    "CALI": "LE001", "PEREIRA": "LE001", "PASTO": "LE001",
    "BUCARAMANGA": "LE001", "CUCUTA": "LE001",
    "BOGOTA": "LE002", "NEIVA": "LE002", "SOGAMOSO": "LE002",
    "MEDELLIN": "LE003", "BARRANQUILLA": "LE003",
    "CARTAGENA": "LE003", "VALLEDUPAR": "LE003", "MONTERIA": "LE003",
}


def _filas_gdd_y_lider(
    df_detalle: pd.DataFrame,
    universo: pd.DataFrame,
    cols_finales: list[str],
) -> pd.DataFrame:
    """
    Genera filas de "resumen" adicionales para:
      • GDDs (1 fila c/u con sus mediciones propias del df_detalle).
      • Líderes (1 fila c/u con sus mediciones propias + promedios de los
        supervisores asignados según LIDER_POR_CIUDAD).
    Esquema idéntico al resumen de supervisor.
    """
    cols_pct = ["CIF_%", "NP_%", "PRECIOS_%", "SOS_%",
                "EXHIB_PAG_%", "EXHIB_PAG_CAPTURA_%",
                "EXHIB_GRA_ALTO_%", "EXHIB_GRA_MEDIO_%",
                "VENTA_%", "IMPACTOS_%", "MSL_%", "PROD_NUEVOS_%"]
    cols_info = ["EXHIB_GRATIS_TOTAL", "EXHIB_GRATIS_ALTO", "EXHIB_GRATIS_MEDIO"]

    if df_detalle.empty or universo.empty:
        return pd.DataFrame(columns=cols_finales)

    mes_v  = df_detalle["MES"].dropna().iloc[0]  if "MES" in df_detalle.columns and df_detalle["MES"].notna().any() else np.nan
    anio_v = df_detalle["AÑO"].dropna().iloc[0]  if "AÑO" in df_detalle.columns and df_detalle["AÑO"].notna().any() else np.nan

    # Mapa rápido ACRONIMO → fila del detalle (KPIs propios).
    det_idx = df_detalle.set_index("ACRONIMO", drop=False)

    def _fila_persona(acr: str, n_gestores: int = 0, venta_eq: float = 0.0) -> dict:
        """Construye una fila estilo resumen con los KPIs de la persona."""
        fila = {c: np.nan for c in cols_finales}
        per = (universo[universo["ACRONIMO"] == acr].iloc[0]
               if (universo["ACRONIMO"] == acr).any() else None)
        fila["ACRONIMO_SUP"]     = acr
        fila["SUPERVISOR_LIDER"] = per["NOMBRE"] if per is not None else acr
        fila["N_GESTORES"]       = int(n_gestores)
        fila["VENTA_EQUIPO"]     = float(venta_eq)
        fila["MES"]  = mes_v
        fila["AÑO"]  = anio_v
        if acr in det_idx.index:
            row = det_idx.loc[acr]
            if isinstance(row, pd.DataFrame):
                row = row.iloc[0]
            for c in cols_pct + cols_info:
                if c in row.index and pd.notna(row[c]):
                    fila[c] = row[c]
        # CUMPL_GLOBAL_% = promedio de KPIs disponibles.
        vals = [fila[c] for c in cols_pct if pd.notna(fila.get(c))]
        fila["CUMPL_GLOBAL_%"] = float(np.mean(vals)) if vals else np.nan
        return fila

    filas: list[dict] = []

    # ── GDDs ──────────────────────────────────────────────────────────────
    gdds = universo[universo["ES_GDD"] == True]
    for _, p in gdds.iterrows():
        f = _fila_persona(p["ACRONIMO"])
        f["_ROL_DEST"] = "GDD"
        filas.append(f)

    # ── Líderes ───────────────────────────────────────────────────────────
    lideres = universo[universo["ES_LIDER"] == True]
    # Sprint 17.18 — solo supervisores cuyo CANAL incluya DIRECTO entran al
    # equipo de un líder (DIRECTO, DIRECTO-DROGUERIAS, DIRECTO-PROXIMITY,
    # DIRECTO-DROGUERIAS-PROXIMITY). DROGUERIAS, PROXIMITY, PROXIMITY-TAT,
    # PROXIMITY-DROGUERIAS quedan FUERA.
    sups_univ = universo[universo["ES_SUPERVISOR"] == True].copy()
    sups_univ["CIUDAD_U"] = sups_univ["CIUDAD_CUPO"].astype(str).str.upper().str.strip()
    sups_univ["LIDER_ACR"] = sups_univ["CIUDAD_U"].map(LIDER_POR_CIUDAD).fillna("")
    sups_univ["CANAL_U"]  = sups_univ["CANAL"].astype(str).str.upper()
    sups_univ = sups_univ[sups_univ["CANAL_U"].str.contains("DIRECTO", na=False)]

    for _, lid in lideres.iterrows():
        acr_lid = lid["ACRONIMO"]
        # Supervisores asignados a este líder (ya filtrados a los que tienen DIRECTO).
        sups_asig = sups_univ[sups_univ["LIDER_ACR"] == acr_lid]["ACRONIMO"].tolist()
        fila = _fila_persona(acr_lid)
        # Sprint 17.18 — para el LIDER, las métricas son el promedio del
        # EQUIPO de cada supervisor asignado (gestores + supervisor mismo),
        # no solo los KPIs propios del supervisor.
        if sups_asig:
            # Equipo completo: gestores cuyo ACRONIMO_SUP esté en sups_asig
            # + las filas propias de los supervisores asignados.
            mask_eq = (
                df_detalle["ACRONIMO_SUP"].isin(sups_asig) |
                df_detalle["ACRONIMO"].isin(sups_asig)
            )
            equipo_filas = df_detalle[mask_eq]
            if not equipo_filas.empty:
                for c in cols_pct:
                    if c in equipo_filas.columns:
                        v = pd.to_numeric(equipo_filas[c], errors="coerce")
                        if c in ("EXHIB_GRA_ALTO_%", "EXHIB_GRA_MEDIO_%"):
                            v = v.clip(upper=1.0)   # cap consistencia HTML
                        v = v.dropna()
                        if not v.empty:
                            fila[c] = float(v.mean())
                for c in cols_info:
                    if c in equipo_filas.columns:
                        v = equipo_filas[c].dropna()
                        if not v.empty:
                            fila[c] = float(v.sum())
                # N_GESTORES para LIDER = # supervisores a cargo (textos
                # "n supervisores a cargo").
                fila["N_GESTORES"] = int(len(sups_asig))
                vals = [fila[c] for c in cols_pct if pd.notna(fila.get(c))]
                fila["CUMPL_GLOBAL_%"] = float(np.mean(vals)) if vals else np.nan
        fila["_ROL_DEST"] = "LIDER"
        filas.append(fila)

    if not filas:
        return pd.DataFrame(columns=cols_finales + ["_ROL_DEST"])
    return pd.DataFrame(filas)[cols_finales + ["_ROL_DEST"]]


# ─────────────────────────────────────────────────────────────────────────────
# D) ESCRITURA DE EXCEL CON FORMATO
# ─────────────────────────────────────────────────────────────────────────────

def _aplicar_formato_tabla(
    ws,
    df: pd.DataFrame,
    cols_pct: list[str],
    cols_venta: list[str] | None = None,
    fila_inicio_datos: int = 2,
) -> None:
    """
    Aplica formato profesional a una hoja:
      • encabezado azul oscuro
      • filas alternas blanco/gris
      • columnas en `cols_pct` con formato 0.0% y color de semáforo
      • columnas en `cols_venta` con formato $#,##0
    """
    cols_venta = cols_venta or []
    thin   = Side(style="thin", color="CCCCCC")
    borde  = Border(left=thin, right=thin, top=thin, bottom=thin)

    hdr_font  = Font(name="Arial", bold=True, color=COLOR_HEADER_FNT, size=10)
    hdr_fill  = PatternFill("solid", fgColor=COLOR_HEADER)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Encabezados (en fila inmediatamente anterior a fila_inicio_datos)
    fila_hdr = fila_inicio_datos - 1
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=fila_hdr, column=col_idx, value=col_name)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = hdr_align
        cell.border    = borde
    ws.row_dimensions[fila_hdr].height = 30

    # Datos
    # Sprint 17.17 — cap a 100% en columnas Exh Gratis Alto/Medio para
    # consistencia con el HTML del cuerpo del correo.
    cols_cap_a_100 = {
        "Cump Exhibiciones Gratis Alto Impacto",
        "Cump Exhibiciones Gratis Medio Impacto",
        "CUMPLIMIENTO EXH. GRATIS ALTO",      # nombres legacy por si quedan
        "CUMPLIMIENTO EXH. GRATIS MEDIO",
        "EXHIB_GRA_ALTO_%",
        "EXHIB_GRA_MEDIO_%",
    }

    fill_par = PatternFill("solid", fgColor=COLOR_FILA_PAR)
    for offset, row in enumerate(df.itertuples(index=False)):
        row_idx = fila_inicio_datos + offset
        fill_base = fill_par if row_idx % 2 == 0 else PatternFill()
        for col_idx, (col_name, val) in enumerate(zip(df.columns, row), start=1):
            # openpyxl no puede serializar pd.NA ni pd.NaT, y NaN lo escribe
            # como cadena rara. Convertimos cualquier "missing" a None.
            if val is pd.NA:
                val = None
            else:
                try:
                    if pd.isna(val):
                        val = None
                except (TypeError, ValueError):
                    pass

            # Sprint 17.17 — cap a 100% en columnas seleccionadas.
            if col_name in cols_cap_a_100 and isinstance(val, (int, float)) and val is not None:
                val = min(float(val), 1.0)

            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font      = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical="center", horizontal="center")
            cell.border    = borde

            if col_name in cols_pct and isinstance(val, (int, float)) and not pd.isna(val):
                cell.number_format = FMT_PCT
                cell.fill = PatternFill("solid", fgColor=_color_celda(val))
            elif col_name in cols_venta and isinstance(val, (int, float)) and not pd.isna(val):
                cell.number_format = FMT_VENTA
                cell.fill = fill_base
            elif isinstance(val, (int, float)) and not pd.isna(val) and not isinstance(val, bool):
                # Sprint 17.17 — numéricos genéricos: máx 4 decimales.
                cell.number_format = FMT_NUM
                cell.fill = fill_base
            else:
                cell.fill = fill_base

    # Anchos de columna automáticos (mín 12, máx 40)
    for col_idx, col_name in enumerate(df.columns, start=1):
        max_len = max(len(str(col_name)), 12)
        for row in df[col_name].astype(str).values:
            max_len = max(max_len, len(str(row)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)

    # Freeze + auto_filter
    ws.freeze_panes = ws.cell(row=fila_inicio_datos, column=1).coordinate
    ws.auto_filter.ref = ws.dimensions


# NOTA: aquí vivían `guardar_detalle()` y `guardar_resumen()`, que escribían
# DETALLE_CUMPLIMIENTO_<mm>_<aaaa>.xlsx y RESUMEN_CUMPLIMIENTO_<mm>_<aaaa>.xlsx
# en la carpeta LOCAL de alertas. Eran código muerto (nadie las llamaba: main()
# publica ambos en SharePoint con _guardar_excel_cloud). Se eliminan por ser
# los últimos puntos del archivo que escribían en disco.

_MENSAJE_VACIO      = "Sin incumplimientos para este periodo"
_MENSAJE_NO_APLICA  = "No aplica para este rol"


def _acronimos_equipo_cascada(df_detalle: pd.DataFrame, acr_sup: str) -> set:
    """
    Conjunto de ACRONIMO_SUP a incluir al filtrar hojas de detalle por PDV
    (CIF/NP/Precios/SOS — Detalle). Para un SUPERVISOR normal es solo su
    propio acrónimo. Para un LÍDER (que tiene supervisores a cargo, no
    gestores directos), se agrega también el acrónimo de cada supervisor
    que le reporta — mismo criterio de LIDER_POR_CIUDAD que ya usa
    construir_equipo_lider() para la hoja "Resumen Equipo" — para que el
    detalle no salga vacío solo porque el líder no tiene gestores directos.
    """
    acrs = {acr_sup}
    propio = df_detalle[df_detalle["ACRONIMO"] == acr_sup]
    if propio.empty or not bool(propio.iloc[0].get("ES_LIDER", False)):
        return acrs

    sups = df_detalle[df_detalle["ES_SUPERVISOR"] == True].copy()
    sups["_CIUDAD"]  = sups.get("CIUDAD_CUPO", "").astype(str).str.upper().str.strip()
    sups["_LIDER"]   = sups["_CIUDAD"].map(LIDER_POR_CIUDAD).fillna("")
    sups["_CANAL_U"] = sups.get("CANAL", "").astype(str).str.upper()
    sups = sups[sups["_CANAL_U"].str.contains("DIRECTO", na=False)]
    acrs |= set(sups[sups["_LIDER"] == acr_sup]["ACRONIMO"].tolist())
    return acrs


def construir_equipo_lider(df_detalle: pd.DataFrame, acr_lider: str) -> pd.DataFrame:
    """
    Para un LIDER, devuelve un DataFrame con (a) su propia fila y (b) una
    fila SINTÉTICA por cada supervisor asignado (filtrados a canal DIRECTO),
    donde los KPIs son el PROMEDIO del equipo del supervisor (gestores +
    supervisor) y N_PDVS es la SUMA. Usado por la hoja "Resumen Equipo"
    del adjunto Excel y por el cuerpo HTML del correo del líder.
    """
    propio = df_detalle[df_detalle["ACRONIMO"] == acr_lider]
    sups = df_detalle[df_detalle["ES_SUPERVISOR"] == True].copy()
    sups["_CIUDAD"]  = sups.get("CIUDAD_CUPO", "").astype(str).str.upper().str.strip()
    sups["_LIDER"]   = sups["_CIUDAD"].map(LIDER_POR_CIUDAD).fillna("")
    sups["_CANAL_U"] = sups.get("CANAL", "").astype(str).str.upper()
    sups = sups[sups["_CANAL_U"].str.contains("DIRECTO", na=False)]
    sups_acr = sups[sups["_LIDER"] == acr_lider]["ACRONIMO"].tolist()

    cols_pct_eq = ["CIF_%", "NP_%", "PRECIOS_%", "SOS_%",
                   "EXHIB_PAG_%", "EXHIB_PAG_CAPTURA_%",
                   "EXHIB_GRA_ALTO_%", "EXHIB_GRA_MEDIO_%",
                   "VENTA_%", "IMPACTOS_%", "MSL_%", "PROD_NUEVOS_%"]
    filas_sup = []
    for acr_s in sups_acr:
        mask = (df_detalle["ACRONIMO_SUP"] == acr_s) | (df_detalle["ACRONIMO"] == acr_s)
        eq_sup = df_detalle[mask]
        if eq_sup.empty:
            continue
        base_row = df_detalle[df_detalle["ACRONIMO"] == acr_s].iloc[0].copy()
        for c in cols_pct_eq:
            if c in eq_sup.columns:
                v = pd.to_numeric(eq_sup[c], errors="coerce")
                if c in ("EXHIB_GRA_ALTO_%", "EXHIB_GRA_MEDIO_%"):
                    v = v.clip(upper=1.0)
                v = v.dropna()
                base_row[c] = float(v.mean()) if not v.empty else np.nan
        if "N_PDVS" in eq_sup.columns:
            base_row["N_PDVS"] = int(
                pd.to_numeric(eq_sup["N_PDVS"], errors="coerce").fillna(0).sum()
            )
        filas_sup.append(base_row)

    partes = [propio] + ([pd.DataFrame(filas_sup)] if filas_sup else [])
    return pd.concat([p for p in partes if not p.empty], ignore_index=True)


def _hoja_resumen_equipo(ws, df_detalle: pd.DataFrame, acr_sup: str) -> None:
    """
    Hoja 1: filas a mostrar al destinatario. Sprint 17.15:
      - SUPERVISOR: sus gestores a cargo + él mismo (si tiene KPIs propios).
      - GDD:        él mismo.
      - LIDER:      él mismo + supervisores que le asigne LIDER_POR_CIUDAD.
    """
    # Detectar rol del destinatario.
    propio = df_detalle[df_detalle["ACRONIMO"] == acr_sup]
    if propio.empty:
        eq = pd.DataFrame(columns=df_detalle.columns)
    else:
        info = propio.iloc[0]
        es_gdd   = bool(info.get("ES_GDD",   False))
        es_lider = bool(info.get("ES_LIDER", False))
        if es_gdd:
            eq = propio.copy()
        elif es_lider:
            eq = construir_equipo_lider(df_detalle, acr_sup)
        else:
            # SUPERVISOR (caso histórico): gestores a cargo + supervisor.
            gestores = df_detalle[
                (df_detalle["ES_SUPERVISOR"] == False)
                & (df_detalle["ACRONIMO_SUP"] == acr_sup)
            ]
            eq = pd.concat([gestores, propio], ignore_index=True)

    cols_pct_calc = ["CIF_%", "NP_%", "PRECIOS_%", "SOS_%",
                     "EXHIB_PAG_%", "EXHIB_PAG_CAPTURA_%",
                     "EXHIB_GRA_ALTO_%", "EXHIB_GRA_MEDIO_%",
                     "VENTA_%", "IMPACTOS_%", "MSL_%", "PROD_NUEVOS_%"]
    # Columnas que efectivamente se muestran (sin VENTA_TOTAL ni MSL_%)
    cols_visibles_origen = [
        "NOMBRE", "ROL", "N_PDVS",
        "CIF_%", "NP_%", "PRECIOS_%", "SOS_%",
        "EXHIB_PAG_CAPTURA_%", "EXHIB_PAG_%",
        "EXHIB_GRA_ALTO_%", "EXHIB_GRA_MEDIO_%",
        "VENTA_%", "IMPACTOS_%", "PROD_NUEVOS_%",
        "CUMPL_GLOBAL_%",
    ]
    # LIDER de ejecución no es responsable de canal D&P; omitir esas columnas
    # del adjunto y del cálculo de CUMPL_GLOBAL.
    if not propio.empty and bool(propio.iloc[0].get("ES_LIDER", False)):
        cols_dyp = {"VENTA_%", "IMPACTOS_%", "MSL_%", "PROD_NUEVOS_%"}
        cols_pct_calc       = [c for c in cols_pct_calc       if c not in cols_dyp]
        cols_visibles_origen = [c for c in cols_visibles_origen if c not in cols_dyp]

    # Mapeo a los nombres del entregable
    RENOMBRES = {
        "NOMBRE":               "NOMBRE",
        "ROL":                  "ROL",
        "N_PDVS":               "POS ASIGNADOS",
        "CIF_%":                "CUMPLIMIENTO CIF",
        "NP_%":                 "CUMPLIMIENTO CAPTURA AGOTADOS",
        "PRECIOS_%":            "CUMPLIMIENTO CAPTURA PRECIOS",
        "SOS_%":                "CUMPLIMIENTO CAPTURA SOS",
        "EXHIB_PAG_CAPTURA_%":  "CUMPLIMIENTO CAPTURA EXH. PAGADAS",
        "EXHIB_PAG_%":          "CUMPLIMIENTO EJECUCIÓN EXH. PAGADAS",
        "EXHIB_GRA_ALTO_%":     "Cump Exhibiciones Gratis Alto Impacto",
        "EXHIB_GRA_MEDIO_%":    "Cump Exhibiciones Gratis Medio Impacto",
        "VENTA_%":              "CUMPLIMIENTO CUOTA D&P",
        "IMPACTOS_%":           "CUMPLIMIENTO IMPACTOS GENERALES D&P",
        "PROD_NUEVOS_%":        "CUMPLIMIENTO IMPACTOS SKU NUEVOS",
        "CUMPL_GLOBAL_%":       "CUMPLIMIENTO GLOBAL",
    }
    cols_finales = [RENOMBRES[c] for c in cols_visibles_origen]

    if eq.empty:
        df_out = pd.DataFrame(columns=cols_finales)
        df_out.loc[0] = [_MENSAJE_NO_APLICA] + [None] * (len(cols_finales) - 1)
    else:
        # Sprint 17.17 — capar Exh Gratis ALTO/MEDIO a 100% antes de promediar
        # (consistencia con el cuerpo HTML del correo). Sobrecumplimientos
        # extremos no inflan el CUMPL_GLOBAL.
        eq = eq.copy()
        for _c in ("EXHIB_GRA_ALTO_%", "EXHIB_GRA_MEDIO_%"):
            if _c in eq.columns:
                eq[_c] = pd.to_numeric(eq[_c], errors="coerce").clip(upper=1.0)
        # Calcular CUMPL_GLOBAL_% por gestor (promedio simple de los KPIs)
        eq["CUMPL_GLOBAL_%"] = eq[cols_pct_calc].mean(axis=1, skipna=True)
        for c in cols_visibles_origen:
            if c not in eq.columns:
                eq[c] = np.nan
        df_out = (
            eq[cols_visibles_origen]
              .sort_values("NOMBRE")
              .reset_index(drop=True)
              .rename(columns=RENOMBRES)
        )

    cols_pct_render_origen = ["CIF_%", "NP_%", "PRECIOS_%", "SOS_%",
                              "EXHIB_PAG_CAPTURA_%", "EXHIB_PAG_%",          # Sprint 17.17
                              "EXHIB_GRA_ALTO_%", "EXHIB_GRA_MEDIO_%",       # Sprint 17.17
                              "VENTA_%", "IMPACTOS_%", "PROD_NUEVOS_%",
                              "CUMPL_GLOBAL_%"]
    cols_pct_render = [RENOMBRES[c] for c in cols_pct_render_origen
                       if RENOMBRES[c] in df_out.columns]
    _aplicar_formato_tabla(
        ws, df_out,
        cols_pct=cols_pct_render,
        cols_venta=[],
    )


def _hoja_cif_detalle(ws, df_cif_pdv: pd.DataFrame, acr_sup: str, idx_bc: dict | None = None, acrs_equipo: set | None = None) -> None:
    """Hoja 2: PDVs del equipo con incumplimiento en algún componente CIF."""
    acrs = acrs_equipo if acrs_equipo else {acr_sup}
    if "ACRONIMO_SUP" not in df_cif_pdv.columns:
        pdvs = pd.DataFrame()
    else:
        pdvs = df_cif_pdv[
            (df_cif_pdv["ROL"] == ROL_GESTOR)
            & (df_cif_pdv["ACRONIMO_SUP"].isin(acrs))
        ].copy()

    # FIX: el rename NOMBRE → NOMBRE_GESTOR se hacía ~45 líneas más abajo, así
    # que _canonizar_nombre_gestor(col="NOMBRE_GESTOR") entraba a su guarda
    # `if col not in df.columns: return df` y no hacía NADA (no-op silencioso).
    # Resultado: la hoja mostraba el nombre crudo de cada fila, con la misma
    # persona escrita de formas distintas. Ahora se renombra primero.
    if not pdvs.empty and "NOMBRE" in pdvs.columns:
        pdvs = pdvs.rename(columns={"NOMBRE": "NOMBRE_GESTOR"})
    if idx_bc and not pdvs.empty:
        pdvs = _canonizar_nombre_gestor(pdvs, idx_bc, col="NOMBRE_GESTOR")

    # Sprint 10: V3 cambió los nombres COB/FREC/HRS → COBERTURA/INTENSIDAD/FRECUENCIA.
    # Mantenemos compat con los nombres legacy para no tocar el resto del flujo.
    if not pdvs.empty:
        rename_v3 = {}
        if "COBERTURA"  in pdvs.columns and "COB"  not in pdvs.columns: rename_v3["COBERTURA"]  = "COB"
        if "INTENSIDAD" in pdvs.columns and "HRS"  not in pdvs.columns: rename_v3["INTENSIDAD"] = "HRS"
        if "FRECUENCIA" in pdvs.columns and "FREC" not in pdvs.columns: rename_v3["FRECUENCIA"] = "FREC"
        if rename_v3:
            pdvs = pdvs.rename(columns=rename_v3)

    if not pdvs.empty:
        # Incumplimiento = COB == 0 OR FREC < 1 OR HRS < 1
        incumple = (
            pdvs["COB"].fillna(0).eq(0)
            | pdvs["FREC"].fillna(1).lt(1)
            | pdvs["HRS"].fillna(1).lt(1)
        )
        pdvs = pdvs[incumple].copy()

    cols_origen = ["NOMBRE_GESTOR", "NOMBRE_PDV", "VENTAS_PROMEDIO_MES",
                   "CANTIDAD_VISITAS", "VISITAS_REAL",
                   "COB", "FREC", "HRS",
                   "HORAS_MES_PLANEADO", "SUMA_TIEMPO_SERVICIO_HORAS_REAL"]

    RENOMBRES = {
        "NOMBRE_GESTOR":                    "NOMBRE COLABORADOR",
        "NOMBRE_PDV":                       "NOMBRE PUNTO DE VENTA",
        "VENTAS_PROMEDIO_MES":              "VENTAS PROMEDIO PDV",
        "CANTIDAD_VISITAS":                 "FREC PLAN",
        "VISITAS_REAL":                     "FREC REAL",
        "COB":                              "CUMPL COBERTURA",
        "FREC":                             "CUMPL FRECUENCIA",
        "HRS":                              "CUMPL INTENSIDAD",
        "HORAS_MES_PLANEADO":               "INTENSIDAD PLAN",
        "SUMA_TIEMPO_SERVICIO_HORAS_REAL":  "INTENSIDAD REAL",
    }
    cols_finales = [RENOMBRES[c] for c in cols_origen]

    if pdvs.empty:
        df_out = pd.DataFrame(columns=cols_finales)
        df_out.loc[0] = [_MENSAJE_VACIO] + [None] * (len(cols_finales) - 1)
    else:
        # (el rename ya se hizo arriba, antes de canonizar)
        df_out = (
            pdvs[cols_origen]
                .sort_values(["NOMBRE_GESTOR", "VENTAS_PROMEDIO_MES"],
                             ascending=[True, False])
                .reset_index(drop=True)
                .rename(columns=RENOMBRES)
        )

    _aplicar_formato_tabla(
        ws, df_out,
        cols_pct=[RENOMBRES["COB"], RENOMBRES["FREC"], RENOMBRES["HRS"]],
        cols_venta=[RENOMBRES["VENTAS_PROMEDIO_MES"]],
    )


def _hoja_np_detalle(ws, df_np: pd.DataFrame, df_cif_pdv: pd.DataFrame, acr_sup: str, idx_bc: dict | None = None, acrs_equipo: set | None = None) -> None:
    """Hoja 3: TODOS los PDVs del equipo del supervisor en NP (capturados y
    no capturados) — refleja el detalle completo que calculó run_all, no
    solo los que quedaron por debajo del umbral.

    Nota: VENTAS_PROMEDIO_MES se eliminó del entregable; sólo se conserva
    internamente como criterio de orden secundario.
    """
    acrs = acrs_equipo if acrs_equipo else {acr_sup}
    cols_origen = ["NOMBRE_GESTOR", "NOMBRE_PDV",
                   "PLANEADO_MES", "CAPTURA_MES", "%_CUMPLIMIENTO_MES"]
    RENOMBRES = {
        "NOMBRE_GESTOR":      "NOMBRE COLABORADOR",
        "NOMBRE_PDV":         "NOMBRE PUNTO DE VENTA",
        "PLANEADO_MES":       "POS PLANEADOS MEDICION",
        "CAPTURA_MES":        "POS MEDIDOS REAL",
        "%_CUMPLIMIENTO_MES": "CUMP CAPTURA",
    }
    cols_finales = [RENOMBRES[c] for c in cols_origen]

    def _vacio(mensaje):
        df_out = pd.DataFrame(columns=cols_finales)
        df_out.loc[0] = [mensaje] + [None] * (len(cols_finales) - 1)
        _aplicar_formato_tabla(
            ws, df_out,
            cols_pct=[RENOMBRES["%_CUMPLIMIENTO_MES"]],
            cols_venta=[],
        )

    if df_np.empty:
        _vacio(_MENSAJE_VACIO)
        return

    df = df_np.copy()
    df["NOMBRE"]             = _normalizar_nombre(df["NOMBRE"])
    df["%_CUMPLIMIENTO_MES"] = pd.to_numeric(df["%_CUMPLIMIENTO_MES"], errors="coerce")

    # Antes filtraba solo %_CUMPLIMIENTO_MES < umbral (solo pendientes).
    # Ahora se conserva el equipo/PLANEADO_MES>=1 pero se incluyen TODOS
    # los PDVs, capturados o no.
    if "ACRONIMO_SUP" not in df.columns:
        df = df.head(0)  # vacío
    else:
        df = df[
            (df["ACRONIMO_SUP"].isin(acrs))
            & (pd.to_numeric(df["PLANEADO_MES"], errors="coerce").fillna(0) >= 1)
        ].copy()

    if df.empty:
        _vacio(_MENSAJE_VACIO)
        return

    # Cruce con CIF para usar VENTAS_PROMEDIO_MES como criterio de orden
    ventas_pdv = (
        df_cif_pdv[["ID_PDV_INVOLVES", "VENTAS_PROMEDIO_MES"]]
        .drop_duplicates(subset=["ID_PDV_INVOLVES"])
    )
    df = df.merge(ventas_pdv, on="ID_PDV_INVOLVES", how="left")
    df = df.rename(columns={"NOMBRE": "NOMBRE_GESTOR"})
    if idx_bc:
        df = _canonizar_nombre_gestor(df, idx_bc, col="NOMBRE_GESTOR")

    df_out = (
        df.sort_values(["%_CUMPLIMIENTO_MES", "VENTAS_PROMEDIO_MES"],
                       ascending=[True, False])
          [cols_origen]
          .reset_index(drop=True)
          .rename(columns=RENOMBRES)
    )

    _aplicar_formato_tabla(
        ws, df_out,
        cols_pct=[RENOMBRES["%_CUMPLIMIENTO_MES"]],
        cols_venta=[],
    )


def _hoja_modulo_captura(
    ws,
    df: pd.DataFrame,
    acr_sup: str,
    umbral_ok: float = 1.0,
    idx_bc: dict | None = None,
    acrs_equipo: set | None = None,
) -> None:
    """Helper genérico para hojas Precios/SOS — Detalle.

    Muestra TODOS los PDVs planeados del equipo (capturados o no) — refleja
    el detalle completo que calculó run_all, no solo los que quedaron por
    debajo del umbral.
    """
    acrs = acrs_equipo if acrs_equipo else {acr_sup}
    cols_origen = ["NOMBRE_GESTOR", "NOMBRE_PDV",
                   "CAPTURA_PLANEADA", "CAPTURA_EJECUTADA",
                   "CATEGORIAS_FALTANTES"]
    RENOMBRES = {
        "NOMBRE_GESTOR":        "NOMBRE COLABORADOR",
        "NOMBRE_PDV":           "NOMBRE PUNTO DE VENTA",
        "CAPTURA_PLANEADA":     "POS PLANEADOS MEDICION",
        "CAPTURA_EJECUTADA":    "POS MEDIDOS REAL",
        "CATEGORIAS_FALTANTES": "DETALLE CATEGORIAS FALTANTES",
    }
    cols_finales = [RENOMBRES[c] for c in cols_origen]

    if df is None or df.empty or "ACRONIMO_SUP" not in df.columns:
        df_out = pd.DataFrame(columns=cols_finales)
        df_out.loc[0] = [_MENSAJE_VACIO] + [None] * (len(cols_finales) - 1)
        _aplicar_formato_tabla(ws, df_out, cols_pct=[], cols_venta=[])
        return

    d = df.copy()
    d["NOMBRE"] = _normalizar_nombre(d["NOMBRE"])
    d["CAPTURA_PLANEADA"]  = pd.to_numeric(d["CAPTURA_PLANEADA"],  errors="coerce").fillna(0)
    d["CAPTURA_EJECUTADA"] = pd.to_numeric(d["CAPTURA_EJECUTADA"], errors="coerce").fillna(0)

    # Cumplimiento por fila = ejecutada/planeada (se conserva solo para
    # ordenar — ya no se usa para filtrar quién aparece en la hoja).
    cumpl = np.where(
        d["CAPTURA_PLANEADA"] > 0,
        d["CAPTURA_EJECUTADA"] / d["CAPTURA_PLANEADA"],
        np.nan,
    )
    d["_cumpl"] = cumpl

    d = d[
        (d["ACRONIMO_SUP"].isin(acrs))
        & (d["CAPTURA_PLANEADA"] >= 1)
    ].copy()

    if d.empty:
        df_out = pd.DataFrame(columns=cols_finales)
        df_out.loc[0] = [_MENSAJE_VACIO] + [None] * (len(cols_finales) - 1)
    else:
        d = d.rename(columns={"NOMBRE": "NOMBRE_GESTOR"})
        if idx_bc:
            d = _canonizar_nombre_gestor(d, idx_bc, col="NOMBRE_GESTOR")
        if "CATEGORIAS_FALTANTES" not in d.columns:
            d["CATEGORIAS_FALTANTES"] = ""
        if "VENTAS_PROMEDIO_MES" not in d.columns:
            d["VENTAS_PROMEDIO_MES"] = np.nan
        df_out = (
            d.sort_values(["_cumpl", "VENTAS_PROMEDIO_MES"],
                           ascending=[True, False])
             [cols_origen]
             .reset_index(drop=True)
             .rename(columns=RENOMBRES)
        )

    _aplicar_formato_tabla(ws, df_out, cols_pct=[], cols_venta=[])


# Caché del Consolidado de Impactos: la hoja se arma una vez por supervisor,
# pero el archivo del periodo es único.
_CACHE_IMPACTOS: dict = {}


def _hoja_impactos_detalle(
    ws,
    df_detalle: pd.DataFrame,
    acr_sup: str,
    mes: int,
    anio: int,
) -> None:
    """Hoja "Impactos — Detalle" (Sprint 13.6).

    Lee Consolidado_Impactos.csv y muestra el detalle por gestor del equipo
    del supervisor con su cumplimiento de impactos (Clientes Impactados /
    Total Clientes Impactar). Solo aparece si el supervisor tiene gestores
    con datos D&P; sino, mensaje "No aplica para este rol".
    """
    cols_origen = ["NOMBRE_GESTOR", "CIUDAD",
                   "CLIENTES_IMPACTAR", "CLIENTES_IMPACTADOS", "IMPACTOS_%"]
    RENOMBRES = {
        "NOMBRE_GESTOR":       "NOMBRE COLABORADOR",
        "CIUDAD":              "CIUDAD",
        "CLIENTES_IMPACTAR":   "CLIENTES A IMPACTAR",
        "CLIENTES_IMPACTADOS": "CLIENTES IMPACTADOS",
        "IMPACTOS_%":          "CUMP IMPACTOS",
    }
    cols_finales = [RENOMBRES[c] for c in cols_origen]

    def _vacio(mensaje):
        df_out = pd.DataFrame(columns=cols_finales)
        df_out.loc[0] = [mensaje] + [None] * (len(cols_finales) - 1)
        _aplicar_formato_tabla(
            ws, df_out,
            cols_pct=[RENOMBRES["IMPACTOS_%"]],
            cols_venta=[],
        )

    # FIX: el Consolidado de Impactos se movió a la nube (SALIDAS/DYP) y ahora
    # lleva mes/año en el nombre. El `paths.DYP_OUT_IMPACTOS.is_file()` local
    # daba False siempre, así que esta hoja salía "No aplica" en TODOS los
    # adjuntos sin ningún aviso.
    from cumplimiento_dyp import MESES_INT_A_STR, parsear_nombre_asesor
    mes_str  = MESES_INT_A_STR.get(int(mes), "")
    anio_str = str(int(anio))

    # Esta hoja se genera una vez POR SUPERVISOR, pero el consolidado del
    # periodo es el mismo para todos. Sin caché se descargaba y parseaba una
    # vez por cada destinatario (con 37 supervisores, 37 descargas evitables
    # del mismo CSV). Mismo criterio que ya usa _CACHE_EXH_PLANNING.
    cache_key = (int(mes), int(anio))
    if cache_key in _CACHE_IMPACTOS:
        cached = _CACHE_IMPACTOS[cache_key]
        if cached is None:
            _vacio(_MENSAJE_NO_APLICA)
            return
        df_imp = cached.copy()
    else:
        ruta_imp = None
        _salidas_dyp = getattr(paths, "RUTA_CARPETA_SALIDAS_DYP", f"{paths._SALIDAS_ROOT}/DYP")
        # El consolidado es un archivo único acumulado. Antes había además
        # candidatos con el periodo en el nombre; se quitaron porque esos
        # archivos ya no se generan y, si quedara alguno viejo dando vueltas,
        # el reporte saldría con datos de otro mes sin avisar.
        _cands_imp = [paths.cloud_dyp_impactos(mes, anio)]
        for _ruta in dict.fromkeys(_cands_imp):
            ruta_imp = _descargar_a_temp(_ruta, f"Impactos {mes_str} {anio}")
            if ruta_imp is not None:
                break
        if ruta_imp is None:
            _CACHE_IMPACTOS[cache_key] = None
            _vacio(_MENSAJE_NO_APLICA)
            return

        try:
            df_imp = pd.read_csv(
                str(ruta_imp),
                sep="|", decimal=",", encoding="utf-8",
                dtype=str, low_memory=False,
            )
            df_imp.columns = [str(c).strip() for c in df_imp.columns]
        except Exception as e:
            print(f"    ⚠️  No pude leer el Consolidado de Impactos: {e}")
            _CACHE_IMPACTOS[cache_key] = None
            _vacio(_MENSAJE_NO_APLICA)
            return
        _CACHE_IMPACTOS[cache_key] = df_imp.copy()

    # FIX: el CSV real trae las columnas como 'Mes', 'Año' y
    # 'Total Clientes a Impactar' (con la "a"). El código pedía 'MES', 'AÑO' y
    # 'Total Clientes Impactar' → KeyError, y como el try/except solo cubría el
    # read_csv, la excepción tumbaba TODO main(). Ahora se resuelven por
    # detección, tolerando ambas variantes.
    col_mes  = next((c for c in df_imp.columns if c.strip().upper() == "MES"), None)
    col_anio = next((c for c in df_imp.columns if c.strip().upper() in ("AÑO", "ANO", "ANIO")), None)
    col_ases = next((c for c in df_imp.columns if "asesor" in c.lower()), None)
    col_ciud = next((c for c in df_imp.columns if "ciudad" in c.lower()), None)
    col_plan = next((c for c in df_imp.columns
                     if "clientes" in c.lower() and "impactar" in c.lower()), None)
    col_real = next((c for c in df_imp.columns
                     if "clientes" in c.lower() and "impactados" in c.lower()), None)
    faltan = [n for n, c in [("MES", col_mes), ("AÑO", col_anio), ("Asesor", col_ases),
                             ("Ciudad", col_ciud), ("Clientes a Impactar", col_plan),
                             ("Clientes Impactados", col_real)] if c is None]
    if faltan:
        print(f"    ⚠️  Consolidado de Impactos sin columnas esperadas: {faltan}")
        print(f"        Columnas encontradas: {list(df_imp.columns)}")
        _vacio(_MENSAJE_NO_APLICA)
        return

    df_imp[col_mes]  = df_imp[col_mes].astype(str).str.strip().str.capitalize()
    df_imp[col_anio] = df_imp[col_anio].astype(str).str.strip()
    df_imp = df_imp[(df_imp[col_mes] == mes_str) & (df_imp[col_anio] == anio_str)].copy()
    if df_imp.empty:
        _vacio(_MENSAJE_NO_APLICA)
        return

    def _num(serie):
        # El CSV usa coma decimal y separador de miles con coma en algunos
        # campos; se limpia antes de convertir.
        return pd.to_numeric(
            serie.astype(str).str.replace(".", "", regex=False).str.replace(",", ".", regex=False),
            errors="coerce",
        ).fillna(0)

    df_imp["NOMBRE_GESTOR"]       = df_imp[col_ases].apply(parsear_nombre_asesor)
    df_imp["CIUDAD"]              = df_imp[col_ciud].astype(str).str.strip()
    df_imp["CLIENTES_IMPACTAR"]   = _num(df_imp[col_plan])
    df_imp["CLIENTES_IMPACTADOS"] = _num(df_imp[col_real])

    # Agregar por (NOMBRE_GESTOR, CIUDAD)
    agg = (
        df_imp.groupby(["NOMBRE_GESTOR", "CIUDAD"], as_index=False)
              .agg(CLIENTES_IMPACTAR=("CLIENTES_IMPACTAR", "sum"),
                   CLIENTES_IMPACTADOS=("CLIENTES_IMPACTADOS", "sum"))
    )
    agg["IMPACTOS_%"] = np.where(
        agg["CLIENTES_IMPACTAR"] > 0,
        agg["CLIENTES_IMPACTADOS"] / agg["CLIENTES_IMPACTAR"],
        np.nan,
    )

    # Filtrar al equipo del supervisor (ACRONIMO_SUP=acr_sup): cruzar
    # NOMBRE_GESTOR con df_detalle para conseguir ACRONIMO_SUP.
    # Sprint 17.16 — armonización por subset de palabras: el CSV de impactos
    # trae nombres truncados ("CLEOFELINA TORRADO") que no matchean con el
    # universo ("CLEOFELINA TORRADO ALVAREZ") por igualdad literal.
    equipo = df_detalle[
        (df_detalle["ES_SUPERVISOR"] == False)
        & (df_detalle["ACRONIMO_SUP"] == acr_sup)
    ][["NOMBRE"]].copy()
    equipo["NOMBRE"] = equipo["NOMBRE"].astype(str).str.strip().str.upper()
    nombres_equipo = set(equipo["NOMBRE"])

    def _match_subset(nombre_csv: str) -> bool:
        pal_csv = _palabras(str(nombre_csv))
        if not pal_csv:
            return False
        for n_eq in nombres_equipo:
            pal_eq = _palabras(n_eq)
            if pal_csv.issubset(pal_eq) or pal_eq.issubset(pal_csv):
                return True
        return False

    agg = agg[agg["NOMBRE_GESTOR"].apply(_match_subset)].copy()

    # Filtrar solo incumplimientos (cumplimiento < UMBRAL_OK_IMPACTOS)
    umb_imp = umbral_de("IMPACTOS_%")
    agg = agg[agg["IMPACTOS_%"].fillna(0) < umb_imp].copy()

    if agg.empty:
        _vacio(_MENSAJE_VACIO)
        return

    df_out = (
        agg.sort_values(["IMPACTOS_%", "CLIENTES_IMPACTAR"],
                         ascending=[True, False])
            [cols_origen]
            .reset_index(drop=True)
            .rename(columns=RENOMBRES)
    )

    _aplicar_formato_tabla(
        ws, df_out,
        cols_pct=[RENOMBRES["IMPACTOS_%"]],
        cols_venta=[],
    )


def _hoja_precios_detalle(ws, df_pr: pd.DataFrame, acr_sup: str, idx_bc: dict | None = None, acrs_equipo: set | None = None) -> None:
    """Hoja Precios — Detalle."""
    _hoja_modulo_captura(ws, df_pr, acr_sup, umbral_ok=umbral_de("PRECIOS_%"), idx_bc=idx_bc, acrs_equipo=acrs_equipo)


def _hoja_sos_detalle(ws, df_sos: pd.DataFrame, acr_sup: str, idx_bc: dict | None = None, acrs_equipo: set | None = None) -> None:
    """Hoja SOS — Detalle."""
    _hoja_modulo_captura(ws, df_sos, acr_sup, umbral_ok=umbral_de("SOS_%"), idx_bc=idx_bc, acrs_equipo=acrs_equipo)


# Versión legacy mantenida por compatibilidad con código antiguo si alguien
# todavía la llama. Sprint 13.5 oficializó las funciones separadas.
def _hoja_precios_sos_detalle(
    ws, df_pr: pd.DataFrame, df_sos: pd.DataFrame, acr_sup: str,
) -> None:
    """[Deprecated Sprint 13.5] Use _hoja_precios_detalle + _hoja_sos_detalle."""
    _hoja_modulo_captura(ws, df_pr, acr_sup, umbral_ok=umbral_de("PRECIOS_%"))


# ─────────────────────────────────────────────────────────────────────────────
# E.5)  HOJAS DE PRODUCTOS NUEVOS POR SUPERVISOR
#       Una hoja por segmento (ListSant, DoyPackBaby, CremasBaby) — MSL queda
#       fuera por petición del cliente. La estructura es la misma que la de
#       `etl_impactos_segmentos.construir_bd`, restringiendo el universo de
#       PDVs a los del supervisor (intersección con los PDVs target del
#       segmento).
# ─────────────────────────────────────────────────────────────────────────────

# Segmentos a generar (excluye MustStock = MSL, según indicación del cliente)
# Sprint 14.5: 4 segmentos V2. Importamos la lista canónica del ETL para
# que el único punto de verdad sea etl_impactos_segmentos.SEGMENTOS_PRODNUEVOS.
try:
    from etl_impactos_segmentos import SEGMENTOS_PRODNUEVOS as _SEGMENTOS_NUEVOS
except Exception:
    # FIX: el fallback traía ["Listerine_Sandia", "DoyPack_J&J_Baby",
    # "Liserine_Kids", "Visine"] — ninguno de esos nombres existe como hoja en
    # listas_referencia.xlsx, que trae ListSant / DoyPackBaby / CremasBaby
    # (+ MSL, excluido por indicación del cliente). Como `_hoja_segmento_nuevo`
    # hace `listas.get(nombre_segmento, {})`, con los nombres viejos el set de
    # EANs salía SIEMPRE vacío y las hojas quedaban en blanco aunque el
    # insumo estuviera bien. Nota: "Liserine_Kids" además venía con typo.
    # Estos nombres deben coincidir EXACTAMENTE con las hojas del archivo de
    # listas; si el cliente agrega un segmento, se agrega la hoja y aquí.
    _SEGMENTOS_NUEVOS = ["ListSant", "DoyPackBaby", "CremasBaby"]


def _precargar_segmentos_nuevos(
    mes: int,
    anio: int,
    idx_bc: dict,
    nombres_sups_bc: set,
) -> dict:
    """
    Carga UNA SOLA VEZ los insumos para las hojas de productos nuevos:
      • rutero (PDVs canónicos)
      • listas de referencia (segmentos → EANs + PDVs target)
      • ventas consolidadas (todas las del CSV)
      • universo de PDVs por supervisor (vendidos por su equipo en el periodo)

    Devuelve dict con claves: rutero, listas, ventas, mes_actual, pdvs_por_sup.
    Si falta cualquier insumo, devuelve dict vacío (las hojas quedarán con
    el mensaje "No aplica para este rol").
    """
    out = {
        "rutero": None, "listas": None, "ventas": None,
        "mes_actual": "", "pdvs_por_sup": {},
    }

    try:
        import etl_impactos_segmentos as eis
    except Exception as e:
        print(f"  ⚠️  No pude importar etl_impactos_segmentos: {e}")
        return out

    # FIX: rutero, listas y consolidado de ventas se movieron a la nube:
    #   BASES DE RESPUESTAS/DYP/<año>/Rutero, .../Listas y SALIDAS/DYP.
    # Los `is_file()` locales daban False y las 3 hojas de productos nuevos
    # salían "No aplica" en todos los adjuntos, sin aviso. Se descargan a
    # temporal para seguir usando etl_impactos_segmentos sin modificarlo.
    import periodo_resolver as _pr_mod
    _spec = _pr_mod.resolver(mes, anio)
    # paths.RUTA_CARPETA_BASES_DYP YA trae el año (AÑO_ACTUAL). _carpeta_periodo
    # lo normaliza al año del periodo sin duplicarlo.
    _bases_dyp   = _carpeta_periodo(
        getattr(paths, "RUTA_CARPETA_BASES_DYP", f"{paths._BASES_ROOT}/DYP"), anio)
    _salidas_dyp = getattr(paths, "RUTA_CARPETA_SALIDAS_DYP", f"{paths._SALIDAS_ROOT}/DYP")

    def _primer_disponible(candidatos: list[tuple[str, str]]) -> Path | None:
        vistos = set()
        for ruta, desc in candidatos:
            if ruta in vistos:
                continue          # evita pedir la misma URL dos veces
            vistos.add(ruta)
            p = _descargar_a_temp(ruta, desc)
            if p is not None:
                return p
        return None

    def _rutero_del_periodo(carpeta: str) -> str | None:
        """
        Busca en la carpeta de Rutero un archivo que corresponda AL PERIODO
        (debe contener el mes y el año en el nombre).

        FIX: la versión anterior tomaba el archivo *más reciente* de la
        carpeta, y en la corrida real eso eligió `Rutero_Droguerias.xlsx`
        (un genérico de 19 MB) en vez del rutero del mes — silenciosamente,
        y los segmentos se calcularon contra el archivo equivocado. El
        nombre correcto lleva mes y año ("RUTERO JULIO 2026 D&P.xlsx") y
        cambia cada mes, así que se exige que ambos aparezcan; solo si no
        hay ninguno con periodo se recurre al genérico, y avisando.
        """
        try:
            hijos = _listar_hijos_cloud(carpeta)
        except Exception:
            return None
        mes_up  = _spec.mes_str_upper
        anio_s  = str(anio)
        con_periodo, genericos = [], []
        for h in hijos:
            nom = h.get("name", "")
            if not nom.lower().endswith(".xlsx") or nom.startswith("~$"):
                continue
            if not nom.upper().startswith("RUTERO"):
                continue
            nom_norm = _sin_tildes(nom).upper()
            if mes_up in nom_norm and anio_s in nom_norm:
                con_periodo.append(h)
            else:
                genericos.append(h)

        if con_periodo:
            con_periodo.sort(key=lambda h: h.get("lastModifiedDateTime", ""), reverse=True)
            return f"{carpeta}/{con_periodo[0]['name']}"
        if genericos:
            genericos.sort(key=lambda h: h.get("lastModifiedDateTime", ""), reverse=True)
            elegido = genericos[0]["name"]
            print(f"    ⚠️  No hay rutero de {mes_up} {anio} en la carpeta; se usará "
                  f"'{elegido}'. Verifica que sea el del periodo — si no lo es, "
                  f"las hojas de productos nuevos saldrán con datos de otro mes.")
            _log.warning("Rutero del periodo ausente; se usó el genérico %s", elegido)
            return f"{carpeta}/{elegido}"
        return None

    _carpeta_rutero = f"{_bases_dyp}/Rutero"
    _r_rutero = []
    if hasattr(paths, "cloud_dyp_rutero"):
        _r_rutero.append((paths.cloud_dyp_rutero(mes, anio), "Rutero D&P"))
    _ruta_listada = _rutero_del_periodo(_carpeta_rutero)
    if _ruta_listada:
        _r_rutero.append((_ruta_listada, "Rutero D&P"))
    ruta_rutero = _primer_disponible(_r_rutero + [
        (f"{_carpeta_rutero}/RUTERO {_spec.mes_str_upper} {anio} D&P.xlsx", "Rutero D&P"),
        (f"{_carpeta_rutero}/RUTERO_{_spec.mes_str_upper}_{anio}_D_P.xlsx", "Rutero D&P"),
    ])
    # Listas D&P: el nombre canónico es "MSL & Listas Target Catman.xlsx"
    # (el mismo que ya declara paths.DYP_LISTAS_FILE). El "&" se codifica como
    # %26 en la URL de Graph, así que no hay que escaparlo aparte.
    _r_listas = [(paths.cloud_dyp_listas(anio), "Listas D&P")] if hasattr(paths, "cloud_dyp_listas") else []
    ruta_listas = _primer_disponible(_r_listas + [
        (f"{_bases_dyp}/Listas/{Path(paths.DYP_LISTAS_FILE).name}", "Listas D&P"),
        (f"{_bases_dyp}/Listas/MSL & Listas Target Catman.xlsx",    "Listas D&P"),
        (f"{_bases_dyp}/Listas/listas_referencia.xlsx",             "Listas D&P (nombre alterno)"),
    ])
    # Archivo único acumulado. Se quitaron los candidatos con el periodo en el
    # nombre: ya no se generan, y un archivo viejo que quedara suelto haría que
    # el reporte saliera con otro mes sin que nada lo advirtiera.
    ruta_ventas = _primer_disponible([
        (paths.cloud_dyp_ventas(mes, anio), "Consolidado Ventas D&P"),
    ])

    # Sin respaldo local: estos insumos viven solo en SharePoint. Si no están
    # ahí hay que resolverlo en la nube, no taparlo con una copia vieja en
    # disco que produciría un reporte con datos de otro periodo.

    if not (ruta_rutero and ruta_listas and ruta_ventas):
        faltan = [n for n, r in [("rutero", ruta_rutero), ("listas", ruta_listas),
                                 ("consolidado de ventas", ruta_ventas)] if r is None]
        print(f"  ⚠️  Faltan insumos D&P en la nube: {faltan}")
        print("       Las hojas de productos nuevos quedarán como 'No aplica'.")
        return out

    try:
        rutero = eis.cargar_rutero(ruta_rutero)
        listas = eis.cargar_listas(ruta_listas)
        # V2 (Sprint 14.3): cargar_ventas_desde_csv devuelve (ventas, desc_map)
        ventas_result = eis.cargar_ventas_desde_csv(ruta_ventas)
        ventas = ventas_result[0] if isinstance(ventas_result, tuple) else ventas_result
    except Exception as e:
        print(f"  ⚠️  Error cargando insumos de segmentos: {e}")
        return out

    if ventas.empty:
        print("  ⚠️  Consolidado de ventas vacío — hojas de productos nuevos no se llenarán.")
        return out

    # Período más reciente — el mismo criterio que usa etl_impactos_segmentos
    ventas["_sort"] = ventas.apply(
        lambda r: eis.periodo_sort_key(f"{r['mes']}-{r['año']}"), axis=1
    )
    fila_max   = ventas.loc[ventas["_sort"].idxmax()]
    mes_actual = f"{fila_max['mes']}-{fila_max['año']}"
    ventas.drop(columns=["_sort"], inplace=True)

    # Universo de PDVs del supervisor — PDVs donde el equipo vendió en el mes
    # activo (según el campo Supervisor del CSV de ventas). Resolvemos cada
    # nombre crudo a su ACRONIMO de Base cupos.
    df_periodo = cumplimiento_dyp._cargar_ventas_periodo(mes, anio)
    pdvs_por_sup: dict[str, set] = {}
    if not df_periodo.empty:
        df_v = df_periodo[df_periodo["cant_total"] > 0].copy()
        df_v["acr_sup"] = df_v["supervisor"].apply(
            lambda s: _resolver_acr_supervisor(s, nombres_sups_bc, idx_bc)
        )
        pdvs_por_sup = (
            df_v[df_v["acr_sup"] != ""]
              .groupby("acr_sup")["cod_cliente"]
              .apply(lambda s: set(s.dropna().astype(str).str.strip()))
              .to_dict()
        )

    out.update({
        "rutero": rutero,
        "listas": listas,
        "ventas": ventas,
        "mes_actual": mes_actual,
        "pdvs_por_sup": pdvs_por_sup,
    })
    return out


def _hoja_segmento_nuevo(
    ws,
    nombre_segmento: str,
    seg_data: dict,
    acr_sup: str,
) -> None:
    """
    Hoja por segmento de producto nuevo (ListSant / DoyPackBaby / CremasBaby).

    Universo: intersección de los PDVs target del segmento con los PDVs
    donde el equipo del supervisor vendió en el periodo. Si esa intersección
    es vacía → "No aplica para este rol".
    """
    # Guard: ¿se pudieron precargar los insumos?
    if not seg_data or seg_data.get("rutero") is None:
        ws.cell(row=1, column=1, value=_MENSAJE_NO_APLICA)
        ws.column_dimensions["A"].width = 50
        return

    rutero      = seg_data["rutero"]
    listas      = seg_data["listas"]
    ventas      = seg_data["ventas"]
    mes_actual  = seg_data["mes_actual"]
    pdvs_sup    = seg_data["pdvs_por_sup"].get(acr_sup, set())

    pdvs_target_seg = listas.get(nombre_segmento, {}).get("pdvs") or set()
    eans_seg        = listas.get(nombre_segmento, {}).get("eans") or set()

    # Universo del supervisor para este segmento
    pdvs_target_para_sup = pdvs_target_seg & pdvs_sup if pdvs_sup else set()

    if not pdvs_target_para_sup:
        ws.cell(row=1, column=1, value=_MENSAJE_NO_APLICA)
        ws.column_dimensions["A"].width = 50
        return

    # Reusamos construir_bd del ETL — misma estructura/columnas que la salida
    # global de impacto_segmentos.xlsx.
    import etl_impactos_segmentos as eis
    df_seg = eis.construir_bd(
        rutero        = rutero,
        ventas        = ventas,
        eans_segmento = eans_seg,
        pdvs_target   = pdvs_target_para_sup,
        mes_actual    = mes_actual,
    )

    if df_seg.empty:
        ws.cell(row=1, column=1, value=_MENSAJE_NO_APLICA)
        ws.column_dimensions["A"].width = 50
        return

    # Si TODOS los PDVs target del supervisor tuvieron impacto → no hay nada
    # que reportar como incumplimiento.
    if "tiene_impacto" in df_seg.columns and (df_seg["tiene_impacto"] == "Sí").all():
        ws.cell(row=1, column=1, value=_MENSAJE_VACIO)
        ws.column_dimensions["A"].width = 50
        return

    # Volcado a la hoja con el mismo formato visual que el ETL global
    # (encabezado azul, filas verde/rojo según impacto).
    for col_idx, col_name in enumerate(df_seg.columns, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)
    for row_idx, row in enumerate(df_seg.itertuples(index=False), start=2):
        for col_idx, val in enumerate(row, start=1):
            try:
                if pd.isna(val):
                    val = None
            except (TypeError, ValueError):
                pass
            ws.cell(row=row_idx, column=col_idx, value=val)
    eis.aplicar_formato(ws, df_seg)


# ─────────────────────────────────────────────────────────────────────────────
# HOJAS ADICIONALES SPRINT 17.10 — Exhibiciones (no capturadas / fuera de regla)
# ─────────────────────────────────────────────────────────────────────────────

_CACHE_EXH_PLANNING: dict = {}
_CACHE_BASE_CUPOS: dict = {}


def _hoja_exh_pagadas_no_capturadas(ws, acr_sup: str, mes: int, anio: int,
                                     df_cif_pdv: pd.DataFrame | None = None) -> None:
    """
    Lista los PDVs del PLANNING de exhibiciones pagadas asignados al equipo
    del supervisor que NO respondieron la encuesta para el periodo.
    Granularidad: (PDV, gestor, marca, tipo). Filtrado al equipo del supervisor.

    Los archivos de planning se localizan y leen directamente desde
    SharePoint (Graph API), replicando en la nube la convención de nombres
    de `periodo_resolver.exh_planning_master` / `exh_base_planning`.

    Esta hoja se genera una vez POR SUPERVISOR (llamada desde
    `generar_adjuntos_por_supervisor`), pero el planning del periodo es el
    mismo para todos — se cachea en `_CACHE_EXH_PLANNING` para no volver a
    descargar los mismos 2 archivos de SharePoint una vez por cada
    supervisor (con 30+ supervisores, eso son 60+ descargas evitables).
    """
    cache_key = (mes, anio)
    if cache_key in _CACHE_EXH_PLANNING:
        cached = _CACHE_EXH_PLANNING[cache_key]
        if isinstance(cached, Exception):
            ws.cell(row=1, column=1, value=f"No disponible: {cached}")
            ws.column_dimensions["A"].width = 80
            return
        df_plan, df_enc = cached[0].copy(), cached[1].copy()
    else:
        import periodo_resolver as pr_mod
        spec = pr_mod.resolver(mes, anio)
        try:
            carpeta_exhib = _resolver_exhib_data_dir_cloud(anio)
            ruta_plan_cloud = _buscar_archivo_cloud(
                carpeta_exhib,
                rf"^PLANNING DE {re.escape(spec.mes_str_upper)} {spec.anio}\.xlsx$",
                contexto=f"Exh/Planning maestro {spec.etiqueta}",
            )
            ruta_enc_cloud = _buscar_archivo_cloud(
                carpeta_exhib,
                rf"^Base Exhibiciones Planning {re.escape(spec.mes_str)} {spec.anio}\.xlsx$",
                contexto=f"Exh/Base Planning {spec.etiqueta}",
            )
        except (FileNotFoundError, RuntimeError) as e:
            _CACHE_EXH_PLANNING[cache_key] = e
            ws.cell(row=1, column=1, value=f"No disponible: {e}")
            ws.column_dimensions["A"].width = 80
            return

        df_plan = _leer_excel_cloud(ruta_plan_cloud, "Exh Planning Maestro")
        df_enc  = _leer_excel_cloud(ruta_enc_cloud,  "Exh Base Planning")
        _CACHE_EXH_PLANNING[cache_key] = (df_plan, df_enc)
        df_plan, df_enc = df_plan.copy(), df_enc.copy()

    col_pdv_p = next((c for c in df_plan.columns if "punto de venta" in str(c).lower()), None)
    col_emp_p = next((c for c in df_plan.columns if "empleado"       in str(c).lower()), None)
    if not (col_pdv_p and col_emp_p):
        ws.cell(row=1, column=1, value="Planning sin columnas reconocidas.")
        return

    df_plan["PDV_N"]      = df_plan[col_pdv_p].astype(str).str.strip().str.upper()
    df_plan["EMPLEADO_N"] = df_plan[col_emp_p].astype(str).str.strip().str.upper()
    col_pdv_e = "PDV" if "PDV" in df_enc.columns else next(
        (c for c in df_enc.columns if str(c).strip().upper() == "PDV"), None)
    pdvs_capturados = (
        set(df_enc[col_pdv_e].astype(str).str.strip().str.upper().unique())
        if col_pdv_e else set()
    )

    # Mapa EMPLEADO_N → ACRONIMO_SUP via Base cupos.
    # Se usa bcm.cargar() (ya migrado a SharePoint) en vez de leer el Excel
    # local directamente. Se cachea a nivel de módulo porque esta hoja se
    # genera una vez por supervisor y bcm.cargar() es una descarga + proceso
    # completo (incluye la auditoría) — sin caché se repetía 30+ veces.
    if "bc" not in _CACHE_BASE_CUPOS:
        _CACHE_BASE_CUPOS["bc"] = bcm.cargar()
    bc = _CACHE_BASE_CUPOS["bc"].copy()
    bc["NOMBRE_N"] = (
        bc["NOMBRE"].astype(str).str.strip().str.upper()
        .str.replace(r"\s+", " ", regex=True)
    )
    # Si Base cupos no tiene ACRONIMO_SUP, derivamos del df_detalle global
    # (queda como tarea simple: filtramos por empleados cuyo nombre aparezca
    # en el equipo del supervisor según el resumen del cumplimiento).
    bc["CANAL"] = bc["CANAL"].astype(str).str.strip().str.upper()

    # Empleados del equipo desde df_cif_pdv (en memoria, tiene ACRONIMO_SUP).
    if df_cif_pdv is not None and not df_cif_pdv.empty and "ACRONIMO_SUP" in df_cif_pdv.columns:
        empleados_eq = set(
            df_cif_pdv.loc[df_cif_pdv["ACRONIMO_SUP"] == acr_sup, "NOMBRE"]
                .astype(str).str.strip().str.upper()
                .str.replace(r"\s+", " ", regex=True).unique()
        )
    else:
        empleados_eq = set()

    if not empleados_eq:
        ws.cell(row=1, column=1, value=_MENSAJE_NO_APLICA)
        ws.column_dimensions["A"].width = 50
        return

    # Sprint 17.16 — armonización por subset de palabras: el PLANNING puede
    # traer nombres truncados que no matchean por igualdad literal.
    pal_eq_set = [(_palabras(n), n) for n in empleados_eq]

    def _match_emp(nombre_csv: str) -> bool:
        pal_csv = _palabras(str(nombre_csv))
        if not pal_csv:
            return False
        for pal_eq, _n in pal_eq_set:
            if pal_csv.issubset(pal_eq) or pal_eq.issubset(pal_csv):
                return True
        return False

    plan_eq = df_plan[df_plan["EMPLEADO_N"].apply(_match_emp)].copy()
    no_cap = plan_eq[~plan_eq["PDV_N"].isin(pdvs_capturados)].copy()
    if no_cap.empty:
        ws.cell(row=1, column=1, value="Sin PDVs pendientes de captura — todos respondieron.")
        ws.column_dimensions["A"].width = 70
        return

    cols_out = [
        col_emp_p, col_pdv_p,
        "*Tipo - Pagadas" if "*Tipo - Pagadas" in no_cap.columns else None,
        "*MARCA - Pagadas" if "*MARCA - Pagadas" in no_cap.columns else None,
        "*Fecha de expiración" if "*Fecha de expiración" in no_cap.columns else None,
    ]
    cols_out = [c for c in cols_out if c]
    detalle = no_cap[cols_out].drop_duplicates().reset_index(drop=True)
    detalle.columns = ["EMPLEADO", "PDV", "TIPO", "MARCA", "FECHA EXP"][:len(cols_out)]
    detalle = detalle.sort_values(["EMPLEADO", "PDV"]).reset_index(drop=True)

    for c_idx, c_name in enumerate(detalle.columns, start=1):
        ws.cell(row=1, column=c_idx, value=c_name)
    for r_idx, row in enumerate(detalle.itertuples(index=False), start=2):
        for c_idx, val in enumerate(row, start=1):
            if pd.isna(val): val = None
            ws.cell(row=r_idx, column=c_idx, value=val)


_CACHE_EXH_GRATIS_FUERA_REGLA: dict = {}


def _hoja_exh_gratis_fuera_regla(ws, acr_sup: str, mes: int, anio: int,
                                  df_cif_pdv: pd.DataFrame | None = None) -> None:
    """
    Lista las exhibiciones gratis del equipo que NO cumplieron la regla de
    "≥2 semanas distintas" cuando frecuencia de referencia > 1. Persistido
    por el ETL en SharePoint, en
    RUTA_CARPETA_SALIDAS_EXHIB/Exh_Gratis_Fuera_de_Regla.xlsx.

    Se cachea a nivel de módulo (`_CACHE_EXH_GRATIS_FUERA_REGLA`) porque
    esta hoja se genera una vez por supervisor pero el archivo fuente es el
    mismo para todos — sin caché se re-descargaba desde SharePoint una vez
    por cada supervisor.
    """
    if "df" not in _CACHE_EXH_GRATIS_FUERA_REGLA:
        ruta_cloud = f"{paths.RUTA_CARPETA_SALIDAS_EXHIB}/Exh_Gratis_Fuera_de_Regla.xlsx"
        _CACHE_EXH_GRATIS_FUERA_REGLA["df"] = _leer(ruta_cloud, "Exh Gratis Fuera de Regla")
    df = _CACHE_EXH_GRATIS_FUERA_REGLA["df"]
    if df.empty:
        ws.cell(row=1, column=1, value="No hay archivo de fuera de regla (no se generó este periodo).")
        ws.column_dimensions["A"].width = 70
        return
    col_mes  = "Mes del año" if "Mes del año" in df.columns else "Mes"
    col_anio = "Año" if "Año" in df.columns else next((c for c in df.columns if "año" in str(c).lower()), None)
    if col_mes in df.columns and col_anio:
        df = df[(pd.to_numeric(df[col_mes], errors="coerce") == mes) &
                (pd.to_numeric(df[col_anio], errors="coerce") == anio)]
    if df.empty:
        ws.cell(row=1, column=1, value="Sin exhibiciones fuera de regla en este periodo.")
        ws.column_dimensions["A"].width = 60
        return

    # Empleados del equipo desde df_cif_pdv (memoria con ACRONIMO_SUP).
    if df_cif_pdv is not None and not df_cif_pdv.empty and "ACRONIMO_SUP" in df_cif_pdv.columns:
        empleados_eq = set(
            df_cif_pdv.loc[df_cif_pdv["ACRONIMO_SUP"] == acr_sup, "NOMBRE"]
                .astype(str).str.strip().str.upper()
                .str.replace(r"\s+", " ", regex=True).unique()
        )
    else:
        empleados_eq = set()

    if "Empleado" in df.columns:
        # Sprint 17.16 — armonización por subset de palabras.
        pal_eq_set = [(_palabras(n), n) for n in empleados_eq]

        def _match_emp(nombre_csv: str) -> bool:
            pal_csv = _palabras(str(nombre_csv))
            if not pal_csv:
                return False
            for pal_eq, _n in pal_eq_set:
                if pal_csv.issubset(pal_eq) or pal_eq.issubset(pal_csv):
                    return True
            return False

        df = df[df["Empleado"].apply(_match_emp)]
    if df.empty:
        ws.cell(row=1, column=1, value="Sin exhibiciones fuera de regla para el equipo.")
        ws.column_dimensions["A"].width = 60
        return

    cols_show = [
        "Empleado", "ID del PDV", "Categoría", "Marca",
        "Tipo Exhibición", "Cantidad", "_semanas_distintas", "_frec_ref",
    ]
    cols_show = [c for c in cols_show if c in df.columns]
    out = df[cols_show].rename(columns={
        "_semanas_distintas": "SEMANAS REGISTRADAS",
        "_frec_ref":          "FRECUENCIA REQUERIDA",
    }).sort_values(["Empleado", "ID del PDV"]).reset_index(drop=True)

    for c_idx, c_name in enumerate(out.columns, start=1):
        ws.cell(row=1, column=c_idx, value=c_name)
    for r_idx, row in enumerate(out.itertuples(index=False), start=2):
        for c_idx, val in enumerate(row, start=1):
            if pd.isna(val): val = None
            ws.cell(row=r_idx, column=c_idx, value=val)


def generar_adjuntos_por_supervisor(
    df_detalle: pd.DataFrame,
    df_cif_pdv: pd.DataFrame,
    df_np:      pd.DataFrame,
    df_pr:      pd.DataFrame,
    df_sos:     pd.DataFrame,
    mes: int,
    anio: int,
    seg_data: dict | None = None,
    idx_bc: dict | None = None,
) -> dict:
    """
    Genera un Excel por supervisor con 7 hojas:
      1. Resumen Equipo
      2. CIF Detalle PDVs
      3. NP Detalle PDVs
      4. Precios y SOS Detalle PDVs
      5. ListSant     (productos nuevos — segmento Sandia)
      6. DoyPackBaby  (productos nuevos — DoyPack Baby)
      7. CremasBaby   (productos nuevos — Cremas Baby)

    `seg_data` es el dict producido por `_precargar_segmentos_nuevos`. Si
    se pasa None o queda vacío, las 3 hojas de productos nuevos se llenan
    con el mensaje "No aplica para este rol".

    Devuelve un dict {NOMBRE_SUPERVISOR: ruta_archivo}.
    """
    # Los adjuntos definitivos viven en la NUBE:
    #   Equipo Información/BI/INVOLVES/SALIDAS/ALERTAS/ADJUNTOS
    # Localmente sólo se deja una copia temporal, necesaria para que
    # `alertas_email` pueda adjuntar el archivo al correo.
    dir_adj = Path(tempfile.gettempdir()) / "alertas_adjuntos"
    dir_adj.mkdir(parents=True, exist_ok=True)

    seg_data = seg_data or {}

    # Sprint 17.15 — destinatarios: SUPERVISOR + GDD + LIDER.
    # Cada uno recibe un adjunto propio.
    if df_detalle.empty:
        supervisores_iter = []
    else:
        sups_df = (
            df_detalle[
                (df_detalle.get("ES_SUPERVISOR", False) == True) |
                (df_detalle.get("ES_GDD",         False) == True) |
                (df_detalle.get("ES_LIDER",       False) == True)
            ][["ACRONIMO", "NOMBRE"]]
                     .drop_duplicates(subset=["ACRONIMO"])
                     .sort_values("NOMBRE")
        )
        supervisores_iter = list(sups_df.itertuples(index=False, name=None))

    rutas = {}
    for acr_sup, nombre_sup in supervisores_iter:
        if not acr_sup or not nombre_sup:
            continue
        nombre_archivo = f"Detalle_{nombre_sup.replace(' ', '_')}_{mes:02d}_{anio}.xlsx"
        ruta = dir_adj / nombre_archivo

        wb = Workbook()
        # Hoja 1 — Resumen del equipo
        ws1 = wb.active
        ws1.title = "Resumen Equipo"
        _hoja_resumen_equipo(ws1, df_detalle, acr_sup)

        # Conjunto de acrónimos a incluir en las hojas de detalle por PDV:
        # para un supervisor normal es solo él mismo; para un LÍDER (que
        # tiene supervisores a cargo, no gestores directos) también incluye
        # a sus supervisores, para que el detalle no salga vacío.
        acrs_equipo = _acronimos_equipo_cascada(df_detalle, acr_sup)

        # Sprint 13.5: nombres alineados con el resumen del correo
        # Hoja 2 — CIF
        ws2 = wb.create_sheet("CIF — Detalle")
        _hoja_cif_detalle(ws2, df_cif_pdv, acr_sup, idx_bc=idx_bc, acrs_equipo=acrs_equipo)

        # Hoja 3 — No Presencia
        ws3 = wb.create_sheet("No Presencia — Detalle")
        _hoja_np_detalle(ws3, df_np, df_cif_pdv, acr_sup, idx_bc=idx_bc, acrs_equipo=acrs_equipo)

        # Hojas 4 y 5 — Precios y SOS separados (D18=Sí)
        ws4 = wb.create_sheet("Precios — Detalle")
        _hoja_precios_detalle(ws4, df_pr, acr_sup, idx_bc=idx_bc, acrs_equipo=acrs_equipo)
        ws5 = wb.create_sheet("SOS — Detalle")
        _hoja_sos_detalle(ws5, df_sos, acr_sup, idx_bc=idx_bc, acrs_equipo=acrs_equipo)

        # Hoja 6 — Impactos D&P (Sprint 13.6)
        ws6 = wb.create_sheet("Impactos — Detalle")
        _hoja_impactos_detalle(ws6, df_detalle, acr_sup, mes, anio)

        # Sprint 17.10 — hojas nuevas de Exhibiciones detalle
        ws_excap = wb.create_sheet("Exh Pagadas — No Capturadas")
        _hoja_exh_pagadas_no_capturadas(ws_excap, acr_sup, mes, anio, df_cif_pdv)

        ws_exgfr = wb.create_sheet("Exh Gratis — Fuera de Regla")
        _hoja_exh_gratis_fuera_regla(ws_exgfr, acr_sup, mes, anio, df_cif_pdv)

        # Hojas finales — Productos nuevos (uno por segmento)
        for nombre_segmento in _SEGMENTOS_NUEVOS:
            ws_seg = wb.create_sheet(nombre_segmento)
            _hoja_segmento_nuevo(ws_seg, nombre_segmento, seg_data, acr_sup)

        # Copia temporal local (para adjuntar al correo)
        wb.save(ruta)

        # Subida a la nube — destino oficial de los adjuntos
        buffer = io.BytesIO()
        wb.save(buffer)
        try:
            _subir_bytes_cloud(
                buffer.getvalue(),
                f"{RUTA_CARPETA_ADJUNTOS_CLOUD}/{nombre_archivo}",
                f"adjunto {nombre_archivo}",
            )
        except Exception as e:
            _log.warning("Fallo al subir adjunto %s a la nube: %s", nombre_archivo, e)

        # Indexamos el dict por NOMBRE del supervisor (lo que usa alertas_email
        # para resolver el destinatario).
        rutas[nombre_sup.upper()] = str(ruta)

    # El conteo estaba fijo en "9 hojas" pero se generan 8 fijas + una por
    # segmento de productos nuevos.
    n_hojas = 8 + len(_SEGMENTOS_NUEVOS)
    print(f"  ✅ {len(rutas)} adjuntos generados ({n_hojas} hojas c/u)")
    return rutas


if __name__ == "__main__":
    main()