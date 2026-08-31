"""
etl_impactos_segmentos.py  (V2 — Sprint 14.3)
─────────────────────────────────────────────
Genera el Excel `impacto_segmentos.xlsx` con hojas:

  1. Resumen           — dashboard global (KPIs + tendencia + por supervisor)
  2. MustStock         — TODO el rutero, marcando si tuvo venta de algún EAN MSL.
  3. Listerine_Sandia  — PDVs target con/sin venta de EANs Listerine Sandia.
  4. DoyPack_J&J_Baby  — PDVs target con/sin venta de EANs DoyPack J&J Baby.
  5. Liserine_Kids     — PDVs target con/sin venta de EANs Listerine Kids.
  6. Visine            — PDVs target con/sin venta de EANs Visine.
  + Hojas de detalle por SKU para segmentos con ≤ SPLIT_SKU_MAX EANs.

Cambios V2 respecto a V1
────────────────────────
  • 4 segmentos productos nuevos (antes 3: ListSant/DoyPackBaby/CremasBaby).
  • Listas en archivo `MSL & Listas Target Catman.xlsx`.
  • Rutero `RUTERO <MES> <AÑO> D&P.xlsx` con `skiprows=13`. Incluye
    SUPERVISOR RESPONSABLE y CUPO GESTOR - TRANSF.
  • Pestaña "Resumen" con dashboard: KPIs globales, resumen por segmento,
    tendencia mensual y tabla por supervisor con columnas mensuales.
  • Split por SKU si segmento tiene ≤ SPLIT_SKU_MAX EANs (default 5).

Inputs requeridos (SharePoint, vía Microsoft Graph API)
────────────────────────────────────────────────────────
  • {RUTA_CARPETA_SALIDAS_DYP}/Consolidado_Ventas.csv — archivo único que ya
    trae todos los periodos (lo mantiene etl_ventas.py por upsert de mes)
  • {RUTA_CARPETA_BASES_DYP}/Listas/MSL & Listas Target Catman.xlsx
  • {RUTA_CARPETA_BASES_DYP}/Rutero/RUTERO <MES> <AÑO> D&P.xlsx (el más
    reciente por fecha de modificación), hoja "PLAN DE TRABAJO"

Output (SharePoint)
────────────────────
  • {RUTA_CARPETA_SALIDAS_DYP}/impacto_segmentos.xlsx
    (Excel con Resumen + segmentos formateados)
"""

from __future__ import annotations

import io
import os
import re
import sys
import urllib.parse

# UTF-8 stdout en Windows (cp1252 explota con ═/─/▶)
try:
    sys.stdout.reconfigure(encoding="utf-8")
except AttributeError:
    pass

from datetime import datetime

import requests
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

import paths


# ─────────────────────────────────────────────────────────────────────────────
# CONEXIÓN A LA NUBE (SharePoint vía Microsoft Graph API)
# ─────────────────────────────────────────────────────────────────────────────
# Los 3 insumos (Rutero, Listas, Consolidado_Ventas) y el Excel de salida
# (impacto_segmentos.xlsx) ya no viven en disco local — se leen/escriben
# directamente en SharePoint, mismo patrón que el resto del pipeline.
_DRIVE_ID_CACHE: dict = {}


def _obtener_default_drive_id(token: str) -> str:
    if "drive_id" in _DRIVE_ID_CACHE:
        return _DRIVE_ID_CACHE["drive_id"]
    headers = {"Authorization": f"Bearer {token}"}
    nombre_sitio = getattr(paths, "SHAREPOINT_SITE_NAME", "eficacia")
    res_site = requests.get(
        f"https://graph.microsoft.com/v1.0/sites?search={urllib.parse.quote(nombre_sitio)}",
        headers=headers,
    )
    if res_site.status_code == 200:
        sites = res_site.json().get("value", [])
        if sites:
            site = next(
                (s for s in sites
                 if nombre_sitio.lower() in (s.get("name", "").lower(), s.get("displayName", "").lower())),
                sites[0],
            )
            res_drive = requests.get(f"https://graph.microsoft.com/v1.0/sites/{site.get('id')}/drive", headers=headers)
            if res_drive.status_code == 200:
                _DRIVE_ID_CACHE["drive_id"] = res_drive.json().get("id")
                return _DRIVE_ID_CACHE["drive_id"]
    res_root = requests.get("https://graph.microsoft.com/v1.0/sites/root", headers=headers)
    if res_root.status_code == 200:
        site_id = res_root.json().get("id")
        res_drive = requests.get(f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive", headers=headers)
        if res_drive.status_code == 200:
            _DRIVE_ID_CACHE["drive_id"] = res_drive.json().get("id")
            return _DRIVE_ID_CACHE["drive_id"]
    raise Exception(f"No se pudo obtener el Drive ID del sitio SharePoint '{nombre_sitio}': {res_site.text}")


def _limpiar_ruta_graph(ruta_sharepoint: str) -> str:
    ruta_sharepoint = ruta_sharepoint.replace("\\", "/")
    for prefijo in ["Documentos compartidos/", "Shared Documents/"]:
        if ruta_sharepoint.startswith(prefijo):
            ruta_sharepoint = ruta_sharepoint.replace(prefijo, "", 1)
    return ruta_sharepoint.lstrip("/")


def _listar_hijos_cloud(ruta_carpeta: str) -> list:
    token = paths.obtener_token_azure()
    headers = {"Authorization": f"Bearer {token}"}
    drive_id = _obtener_default_drive_id(token)
    ruta_codificada = urllib.parse.quote(_limpiar_ruta_graph(ruta_carpeta))
    url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/root:/{ruta_codificada}:/children"
    items = []
    while url:
        response = requests.get(url, headers=headers)
        if response.status_code != 200:
            raise Exception(f"Error al listar carpeta '{ruta_carpeta}': {response.status_code} - {response.text}")
        data = response.json()
        items.extend(data.get("value", []))
        url = data.get("@odata.nextLink")
    return items


def _descargar_bytes_cloud(ruta_sharepoint: str, descripcion: str) -> bytes:
    token = paths.obtener_token_azure()
    headers = {"Authorization": f"Bearer {token}"}
    drive_id = _obtener_default_drive_id(token)
    ruta_limpia = _limpiar_ruta_graph(ruta_sharepoint)
    response = None
    for r in (ruta_limpia, f"Documentos compartidos/{ruta_limpia}"):
        url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/root:/{urllib.parse.quote(r)}:/content"
        response = requests.get(url, headers=headers)
        if response.status_code == 200:
            return response.content
    status = response.status_code if response is not None else "N/A"
    text = response.text if response is not None else "Sin respuesta"
    raise FileNotFoundError(f"No se pudo descargar {descripcion} desde SharePoint ({ruta_sharepoint}): {status} - {text}")


def _subir_bytes_cloud(contenido: bytes, ruta_sharepoint: str, content_type: str) -> None:
    token = paths.obtener_token_azure()
    headers = {"Authorization": f"Bearer {token}", "Content-Type": content_type}
    drive_id = _obtener_default_drive_id(token)
    ruta_limpia = _limpiar_ruta_graph(ruta_sharepoint)
    url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/root:/{urllib.parse.quote(ruta_limpia)}:/content"
    response = requests.put(url, headers=headers, data=contenido)
    if response.status_code not in (200, 201):
        raise Exception(f"Error al subir '{ruta_sharepoint}' a SharePoint: {response.status_code} - {response.text}")


def _descargar_rutero_cloud() -> bytes:
    """Toma el RUTERO*.xlsx / Rutero*.xlsx más reciente en
    {RUTA_CARPETA_BASES_DYP}/Rutero (mismo criterio que paths._resolver_rutero_dyp:
    el de fecha de modificación más reciente)."""
    carpeta = f"{paths.RUTA_CARPETA_BASES_DYP}/Rutero"
    hijos = _listar_hijos_cloud(carpeta)
    candidatos = [
        h for h in hijos
        if h.get("file") and re.match(r"^(RUTERO|Rutero).*\.xlsx$", h.get("name", ""))
        and not h["name"].startswith("~$")
    ]
    if not candidatos:
        raise FileNotFoundError(f"No se encontró ningún RUTERO*.xlsx en SharePoint ({carpeta})")
    candidatos.sort(key=lambda h: h.get("lastModifiedDateTime", ""), reverse=True)
    nombre = candidatos[0]["name"]
    return _descargar_bytes_cloud(f"{carpeta}/{nombre}", f"Rutero ({nombre})")


def _descargar_listas_cloud() -> bytes:
    ruta = f"{paths.RUTA_CARPETA_BASES_DYP}/Listas/MSL & Listas Target Catman.xlsx"
    return _descargar_bytes_cloud(ruta, "Listas D&P (MSL/Prod Nuevos)")


def _descargar_ventas_historico_cloud() -> list[bytes]:
    """
    Descarga `Consolidado_Ventas.csv`, que ya trae todos los meses.

    Antes esta función listaba TODOS los `Consolidado_Ventas_<MES>_<AÑO>.csv`
    y los concatenaba, porque etl_ventas.py subía uno por periodo. Pero el
    contenido de cada uno ya era el histórico entero, así que se bajaban ~800 MB
    para terminar con cada mes repetido seis veces — y con cifras distintas
    entre copias, porque cada archivo era una foto de un momento diferente.
    Ahora el consolidado es un archivo único y basta con leer ese.

    Se devuelve una lista de un elemento para no cambiarle la forma al caller.
    """
    ruta = paths.cloud_dyp_ventas()
    print(f"      Leyendo consolidado: {ruta}")
    try:
        return [_descargar_bytes_cloud(ruta, "Consolidado de ventas")]
    except Exception as e:
        raise FileNotFoundError(
            f"No se pudo leer {ruta}. Ejecuta primero etl_ventas.py. Detalle: {e}"
        ) from e


# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURACIÓN
# ─────────────────────────────────────────────────────────────────────────────
RUTERO_COLS_SALIDA = [
    "ID ECOM", "PUNTO DE VENTA", "CIUDAD", "DEPARTAMENTO",
    "REGION EFICACIA", "DECIL", "TAMAÑO", "CORPORATIVO",
    "SUPERVISOR RESPONSABLE", "CUPO GESTOR - TRANSF",
]

# Columna de descripción de producto en los archivos de ventas consolidados.
COL_DESC_PRODUCTO = "Producto"

# Segmentos con hasta este número de EANs generan pestañas de detalle por SKU
SPLIT_SKU_MAX = 5

# Lista canónica de segmentos V2 (orden estable para los outputs)
# El nombre interno coincide con el nombre de la hoja en MSL & Listas Target Catman.xlsx.
SEGMENTOS_PRODNUEVOS = [
    "Listerine_Sandia",
    "DoyPack_J&J_Baby",
    "Liserine_Kids",
    "Visine",
]

MESES_ORDEN = [
    "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
    "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre",
]

# Paleta Eficacia
COLOR_HEADER     = "0A1628"
COLOR_SECCION    = "1A3A6B"
COLOR_CYAN_ACC   = "00B5C8"
COLOR_IMPACTO_SI = "C6EFCE"
COLOR_IMPACTO_NO = "FFDCDC"
COLOR_ALTROW     = "F0F6FA"
COLOR_KPI_BG     = "D6EEF2"

# Nombres personalizados de pestañas de SKU (opcional — clave = nombre auto-generado)
CUSTOM_TAB_NAMES = {
    "CHAMPU BABY JJ ORIGINAL DOYPACK": "DoyPack_Shampoo_Baby",
    "BANO LIQUIDO BABY JJ CABEZA A P": "DoyPack_BañoLiquido_Baby",
}


# ─────────────────────────────────────────────────────────────────────────────
# 1. CARGA DE FUENTES
# ─────────────────────────────────────────────────────────────────────────────

def cargar_rutero(path) -> pd.DataFrame:
    """Carga el rutero. Auto-detecta la fila del header buscando "ID ECOM"
    en las primeras 20 filas (algunos exports vienen con 13 filas de
    basura arriba, otros con el header en la fila 1). Devuelve sólo las
    columnas listadas en RUTERO_COLS_SALIDA (con ID ECOM como string).

    Acepta una ruta de archivo local O un BytesIO/bytes (descargado de
    SharePoint). Si es BytesIO/bytes, se normaliza a bytes crudos una sola
    vez y se abre un BytesIO NUEVO en cada llamada a pd.read_excel — un
    mismo objeto BytesIO se agota (EOF) después de la primera lectura.
    """
    if isinstance(path, io.BytesIO):
        contenido = path.getvalue()
    elif isinstance(path, bytes):
        contenido = path
    else:
        contenido = None  # es una ruta de archivo local; se usa tal cual

    def _fuente():
        return io.BytesIO(contenido) if contenido is not None else path

    skip_detectado = None
    for skip in range(0, 20):
        try:
            preview = pd.read_excel(_fuente(), sheet_name="PLAN DE TRABAJO",
                                     skiprows=skip, nrows=0)
            cols_upper = {str(c).upper() for c in preview.columns}
            if "ID ECOM" in cols_upper:
                skip_detectado = skip
                break
        except Exception:
            continue
    if skip_detectado is None:
        raise KeyError(
            "No se encontró la columna 'ID ECOM' en el rutero "
            f"(busqué en skiprows 0..19). Archivo: {path}"
        )
    df = pd.read_excel(_fuente(), sheet_name="PLAN DE TRABAJO",
                        skiprows=skip_detectado,
                        dtype={"ID ECOM": str})
    # Asegurar que todas las columnas sean strings (algunas vienen como datetime)
    df.columns = [str(c) for c in df.columns]
    # Renombrado case-insensitive
    upper_map = {c.upper(): c for c in df.columns}
    renames = {
        upper_map[w.upper()]: w
        for w in RUTERO_COLS_SALIDA
        if w.upper() in upper_map and upper_map[w.upper()] != w
    }
    if renames:
        df = df.rename(columns=renames)
    df["ID ECOM"] = df["ID ECOM"].astype(str).str.strip()
    cols    = [c for c in RUTERO_COLS_SALIDA if c in df.columns]
    missing = [c for c in RUTERO_COLS_SALIDA if c not in df.columns]
    if missing:
        print(f"      [!] Columnas no encontradas en rutero: {missing}")
    return df[cols].drop_duplicates(subset=["ID ECOM"]).reset_index(drop=True)


def cargar_listas(path) -> dict:
    """Carga MSL + los 4 segmentos V2 desde `MSL & Listas Target Catman.xlsx`.

    Devuelve:
      { nombre_segmento: {"eans": set(str), "pdvs": set(str) | None} }
      • MSL → pdvs=None (todo el rutero)
      • Otros → pdvs=set de targets

    Acepta ruta local o BytesIO/bytes (ver nota en `cargar_rutero`: se
    normaliza a bytes una vez y se abre un BytesIO nuevo por cada hoja).
    """
    if isinstance(path, io.BytesIO):
        contenido = path.getvalue()
    elif isinstance(path, bytes):
        contenido = path
    else:
        contenido = None

    def _fuente():
        return io.BytesIO(contenido) if contenido is not None else path

    listas: dict[str, dict] = {}

    # MSL — solo EANs, sin target de PDVs
    df_msl = pd.read_excel(_fuente(), sheet_name="MSL", dtype=str)
    listas["MustStock"] = {
        "eans": set(df_msl["EAN"].dropna().str.strip()),
        "pdvs": None,
    }

    # 4 segmentos de productos nuevos
    for nombre_segmento in SEGMENTOS_PRODNUEVOS:
        df = pd.read_excel(_fuente(), sheet_name=nombre_segmento, dtype=str)
        eans: set[str] = set()
        for v in df["EAN PRODUCTO"].dropna():
            try:
                eans.add(str(int(float(v.strip()))))
            except Exception:
                eans.add(v.strip())
        listas[nombre_segmento] = {
            "eans": eans,
            "pdvs": set(df["ID ECOM"].dropna().str.strip()),
        }

    return listas


def cargar_ventas_desde_csv(csv_source) -> tuple[pd.DataFrame, dict]:
    """
    Lee el/los consolidado(s) de ventas. Acepta:
      - una ruta de archivo local (compatibilidad hacia atrás)
      - un único `bytes` (un CSV ya descargado de la nube)
      - una lista de `bytes` (varios CSV, uno por periodo — se concatenan;
        así se reconstruye el histórico completo para la tendencia mensual)
    Devuelve (ventas_df, desc_map).
    """
    cols_necesarias = [
        "Cod. Cliente", "Cod. EAN Producto",
        "Cantidades Totales", "Ventas Totales",
        "MES", "AÑO",
    ]
    cols_opt = [COL_DESC_PRODUCTO]

    def _leer_uno(fuente) -> pd.DataFrame:
        return pd.read_csv(
            fuente, sep="|", decimal=",", encoding="utf-8",
            dtype={"Cod. Cliente": str, "Cod. EAN Producto": str,
                   "MES": str, "AÑO": str, COL_DESC_PRODUCTO: str},
            low_memory=False,
        )

    if isinstance(csv_source, (list, tuple)):
        if not csv_source:
            raise FileNotFoundError("No se encontraron consolidados de ventas en la nube.")
        partes = [_leer_uno(io.BytesIO(b) if isinstance(b, bytes) else b) for b in csv_source]
        df = pd.concat(partes, ignore_index=True)
    elif isinstance(csv_source, bytes):
        df = _leer_uno(io.BytesIO(csv_source))
    elif isinstance(csv_source, (str, os.PathLike)) and os.path.isfile(csv_source):
        print(f"      Leyendo consolidado: {csv_source}")
        df = _leer_uno(csv_source)
    else:
        raise FileNotFoundError(
            f"No existe el consolidado de ventas: {csv_source}\n"
            f"Ejecuta primero `etl_ventas.py`."
        )

    # Usar solo las cols que existen
    cols_usar = cols_necesarias + [c for c in cols_opt if c in df.columns]
    df = df[cols_usar].copy()

    df["Cod. EAN Producto"] = df["Cod. EAN Producto"].fillna("").astype(str).str.strip()
    def _ean_clean(v: str) -> str:
        if v in ("", "nan"):
            return ""
        try:
            return str(int(float(v)))
        except Exception:
            return v
    df["Cod. EAN Producto"] = df["Cod. EAN Producto"].map(_ean_clean)

    df["Cod. Cliente"] = df["Cod. Cliente"].fillna("").astype(str).str.strip()
    df["MES"] = df["MES"].fillna("").astype(str).str.strip()
    df["AÑO"] = df["AÑO"].fillna("").astype(str).str.strip()
    df["Cantidades Totales"] = pd.to_numeric(df["Cantidades Totales"], errors="coerce").fillna(0)
    df["Ventas Totales"]     = pd.to_numeric(df["Ventas Totales"],     errors="coerce").fillna(0)

    # Descripción de producto (primera ocurrencia por EAN)
    desc_map: dict[str, str] = {}
    if COL_DESC_PRODUCTO in df.columns:
        for ean, sub in df.dropna(subset=[COL_DESC_PRODUCTO]).groupby("Cod. EAN Producto"):
            if ean:
                desc_map[ean] = str(sub.iloc[0][COL_DESC_PRODUCTO]).strip()

    df = df.rename(columns={
        "Cod. Cliente":         "cod_pdv",
        "Cod. EAN Producto":    "ean",
        "MES":                  "mes",
        "AÑO":                  "año",
        "Cantidades Totales":   "cant_total",
        "Ventas Totales":       "vta_total",
    })
    df = df[df["cant_total"] > 0].copy()
    df = (
        df.groupby(["cod_pdv", "ean", "mes", "año"], as_index=False)
          .agg(cant_total=("cant_total", "sum"),
               vta_total=("vta_total", "sum"))
    )
    return df, desc_map


# ─────────────────────────────────────────────────────────────────────────────
# 2. CONSTRUCCIÓN DE LA BD POR SEGMENTO
# ─────────────────────────────────────────────────────────────────────────────

def periodo_sort_key(mes_anio: str) -> int:
    """`Enero-2025` → 202501."""
    try:
        m, a = mes_anio.split("-")
        return int(a) * 100 + (MESES_ORDEN.index(m) + 1)
    except Exception:
        return 99999


def _make_tab_name(base: str, desc: str, ean: str) -> str:
    """Genera nombre de pestaña Excel válido (≤31 chars) a partir de la descripción."""
    raw = desc.strip() if desc.strip() else f"{base}_{ean[-4:]}"
    for ch in "/\\?*[]":
        raw = raw.replace(ch, " ")
    raw = raw.replace(":", " ")
    return raw.strip()[:31]


def construir_bd(
    rutero: pd.DataFrame,
    ventas: pd.DataFrame,
    eans_segmento: set,
    pdvs_target,
    mes_actual: str,
) -> pd.DataFrame:
    # Universo
    if pdvs_target is not None:
        universo = rutero[rutero["ID ECOM"].isin(pdvs_target)].copy()
    else:
        universo = rutero.copy()

    # Ventas del segmento
    ventas_seg = ventas[ventas["ean"].isin(eans_segmento)].copy()
    ventas_seg["periodo"] = ventas_seg["mes"] + "-" + ventas_seg["año"]

    if not ventas_seg.empty:
        metricas = (
            ventas_seg
            .groupby("cod_pdv")
            .agg(
                meses_lista     =("periodo", lambda x: sorted(x.unique(), key=periodo_sort_key)),
                eans_lista      =("ean",     lambda x: sorted(x.unique())),
                n_meses_impacto =("periodo", "nunique"),
            )
            .reset_index()
        )
        metricas["meses_con_venta"]  = metricas["meses_lista"].apply(", ".join)
        metricas["eans_vendidos"]    = metricas["eans_lista"].apply(", ".join)
        metricas["venta_mes_actual"] = metricas["meses_lista"].apply(
            lambda x: "Sí" if mes_actual in x else "No"
        )
        metricas = metricas[[
            "cod_pdv", "n_meses_impacto",
            "meses_con_venta", "venta_mes_actual", "eans_vendidos",
        ]]

        pivot = ventas_seg.pivot_table(
            index="cod_pdv", columns="periodo",
            values="cant_total", aggfunc="sum", fill_value=0,
        ).reset_index()
        periodos_cols = sorted(
            [c for c in pivot.columns if c != "cod_pdv"],
            key=periodo_sort_key,
        )
        pivot = pivot[["cod_pdv"] + periodos_cols]
    else:
        metricas = pd.DataFrame(columns=[
            "cod_pdv", "n_meses_impacto",
            "meses_con_venta", "venta_mes_actual", "eans_vendidos",
        ])
        pivot = pd.DataFrame(columns=["cod_pdv"])
        periodos_cols = []

    # Joins
    resultado = universo.merge(
        metricas, left_on="ID ECOM", right_on="cod_pdv", how="left",
    ).drop(columns=["cod_pdv"], errors="ignore")
    resultado = resultado.merge(
        pivot, left_on="ID ECOM", right_on="cod_pdv", how="left",
    ).drop(columns=["cod_pdv"], errors="ignore")

    # Llenar nulos
    resultado["tiene_impacto"]    = resultado["n_meses_impacto"].notna().map({True: "Sí", False: "No"})
    resultado["n_meses_impacto"]  = resultado["n_meses_impacto"].fillna(0).astype(int)
    resultado["meses_con_venta"]  = resultado["meses_con_venta"].fillna("")
    resultado["venta_mes_actual"] = resultado["venta_mes_actual"].fillna("No")
    resultado["eans_vendidos"]    = resultado["eans_vendidos"].fillna("")
    for col in periodos_cols:
        resultado[col] = resultado[col].fillna(0).astype(int)

    cols_rutero   = [c for c in RUTERO_COLS_SALIDA if c in resultado.columns]
    cols_metricas = [
        "tiene_impacto", "n_meses_impacto",
        "meses_con_venta", "venta_mes_actual", "eans_vendidos",
    ]
    resultado = resultado[cols_rutero + cols_metricas + periodos_cols]

    resultado = resultado.sort_values(
        ["tiene_impacto", "n_meses_impacto"],
        ascending=[True, False],
    ).reset_index(drop=True)

    return resultado


# ─────────────────────────────────────────────────────────────────────────────
# 3a. FORMATO HOJAS DE DETALLE
# ─────────────────────────────────────────────────────────────────────────────

def aplicar_formato(ws, df: pd.DataFrame) -> None:
    fill_header     = PatternFill("solid", fgColor=COLOR_HEADER)
    fill_impacto_si = PatternFill("solid", fgColor=COLOR_IMPACTO_SI)
    fill_impacto_no = PatternFill("solid", fgColor=COLOR_IMPACTO_NO)
    fill_altrow     = PatternFill("solid", fgColor=COLOR_ALTROW)
    font_header     = Font(bold=True, color="FFFFFF", size=10)
    font_normal     = Font(size=10)
    alin_centro     = Alignment(horizontal="center", vertical="center")
    alin_izq        = Alignment(horizontal="left",   vertical="center")
    thin   = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill      = fill_header
        cell.font      = font_header
        cell.alignment = alin_centro
        cell.border    = border

    headers = [cell.value for cell in ws[1]]
    idx_impacto = headers.index("tiene_impacto") + 1 if "tiene_impacto" in headers else None

    for i, row in enumerate(ws.iter_rows(min_row=2), start=1):
        impacto_val = row[idx_impacto - 1].value if idx_impacto else None
        fill_fila = (
            fill_impacto_si if impacto_val == "Sí"
            else fill_impacto_no if impacto_val == "No"
            else (fill_altrow if i % 2 == 0 else None)
        )
        for cell in row:
            cell.font      = font_normal
            cell.border    = border
            cell.alignment = alin_izq if cell.column <= 2 else alin_centro
            if fill_fila:
                cell.fill = fill_fila

    for col_idx, col_cells in enumerate(ws.iter_cols(), start=1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            (len(str(c.value)) if c.value is not None else 0) for c in col_cells
        )
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 45)

    ws.freeze_panes    = "A2"
    ws.auto_filter.ref = ws.dimensions


# ─────────────────────────────────────────────────────────────────────────────
# 3b. MÉTRICAS PARA EL DASHBOARD
# ─────────────────────────────────────────────────────────────────────────────

def _es_periodo(col: str) -> bool:
    parts = str(col).split("-")
    return len(parts) == 2 and parts[0] in MESES_ORDEN


def _resumen_por_segmento(bds: dict, mes_actual: str) -> pd.DataFrame:
    rows = []
    for nombre, df in bds.items():
        total   = len(df)
        imp     = int((df["tiene_impacto"] == "Sí").sum())
        pct     = round(imp / total * 100, 1) if total > 0 else 0.0
        mes_act = int((df["venta_mes_actual"] == "Sí").sum())
        rows.append({
            "Segmento NPI":          nombre,
            "PDVs Target":           total,
            "Con Impacto":           imp,
            "Sin Impacto":           total - imp,
            "% Impacto":             pct,
            f"Impacto {mes_actual}": mes_act,
        })
    return pd.DataFrame(rows)


def _tendencia_mensual(bds: dict) -> pd.DataFrame:
    """PDVs con ventas > 0 por período por segmento."""
    all_periods: set = set()
    for df in bds.values():
        all_periods.update(c for c in df.columns if _es_periodo(c))
    periods = sorted(all_periods, key=periodo_sort_key)
    if not periods:
        return pd.DataFrame()

    rows = []
    for p in periods:
        fila: dict = {"Período": p}
        for nombre, df in bds.items():
            fila[nombre] = int((df[p] > 0).sum()) if p in df.columns else 0
        rows.append(fila)

    tend = pd.DataFrame(rows)
    tend["Total"] = tend[list(bds.keys())].sum(axis=1)
    return tend


def _resumen_supervisor_por_segmento(bds: dict, mes_actual: str) -> dict:
    """Retorna {segmento: DataFrame} con fila por supervisor + cols mensuales."""
    COL_SUP = "SUPERVISOR RESPONSABLE"
    result  = {}

    for nombre, df in bds.items():
        if COL_SUP not in df.columns:
            result[nombre] = pd.DataFrame()
            continue

        period_cols = sorted(
            [c for c in df.columns if _es_periodo(c)], key=periodo_sort_key
        )

        agg = (
            df.groupby(COL_SUP, dropna=False)
            .agg(
                PDVs_Target    =("tiene_impacto",    "count"),
                Con_Impacto    =("tiene_impacto",    lambda x: (x == "Sí").sum()),
                Impacto_MesAct =("venta_mes_actual", lambda x: (x == "Sí").sum()),
            )
            .reset_index()
        )
        agg["Con_Impacto"]    = agg["Con_Impacto"].astype(int)
        agg["Impacto_MesAct"] = agg["Impacto_MesAct"].astype(int)
        agg["Sin Impacto"]    = agg["PDVs_Target"] - agg["Con_Impacto"]
        agg["% Impacto"]      = (agg["Con_Impacto"] / agg["PDVs_Target"] * 100).round(1)

        for p in period_cols:
            monthly = (
                df.groupby(COL_SUP, dropna=False)[p]
                .apply(lambda x: int((x > 0).sum()))
                .reset_index(name=p)
            )
            agg = agg.merge(monthly, on=COL_SUP, how="left")
            agg[p] = agg[p].fillna(0).astype(int)

        agg = agg.rename(columns={
            COL_SUP:          "Supervisor",
            "PDVs_Target":    "PDVs Target",
            "Con_Impacto":    "Con Impacto",
            "Impacto_MesAct": f"Impacto {mes_actual}",
        })
        cols_ord = (
            ["Supervisor", "PDVs Target", "Con Impacto", "Sin Impacto",
             "% Impacto", f"Impacto {mes_actual}"]
            + period_cols
        )
        agg = agg[[c for c in cols_ord if c in agg.columns]]
        result[nombre] = agg.sort_values("% Impacto", ascending=False).reset_index(drop=True)

    return result


# ─────────────────────────────────────────────────────────────────────────────
# 3c. ESCRITURA DEL DASHBOARD
# ─────────────────────────────────────────────────────────────────────────────

def _brd():
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _titulo_seccion(ws, row: int, texto: str, n_cols: int = 12) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
    c = ws.cell(row=row, column=1, value=texto)
    c.fill      = PatternFill("solid", fgColor=COLOR_SECCION)
    c.font      = Font(bold=True, color="FFFFFF", size=11)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 22
    return row + 1


def _escribir_tabla(ws, df: pd.DataFrame, start_row: int, start_col: int = 1) -> int:
    border = _brd()
    for j, col in enumerate(df.columns, start=start_col):
        c = ws.cell(row=start_row, column=j, value=col)
        c.fill      = PatternFill("solid", fgColor=COLOR_HEADER)
        c.font      = Font(bold=True, color="FFFFFF", size=10)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = border
    ws.row_dimensions[start_row].height = 22

    for i, (_, row_data) in enumerate(df.iterrows()):
        r  = start_row + 1 + i
        bg = COLOR_ALTROW if i % 2 == 1 else None
        for j, val in enumerate(row_data, start=start_col):
            c = ws.cell(row=r, column=j, value=val)
            c.font      = Font(size=10)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border    = border
            if bg:
                c.fill = PatternFill("solid", fgColor=bg)
        ws.row_dimensions[r].height = 16

    return start_row + 1 + len(df)


def _escribir_tabla_dyn(
    ws, df: pd.DataFrame, start_row: int, table_name: str, start_col: int = 1
) -> int:
    next_row = _escribir_tabla(ws, df, start_row, start_col)
    end_row  = next_row - 1
    end_col  = start_col + len(df.columns) - 1
    ref = (
        f"{get_column_letter(start_col)}{start_row}:"
        f"{get_column_letter(end_col)}{end_row}"
    )
    safe_name = "T" + "".join(c if c.isalnum() else "_" for c in table_name)
    tab = Table(displayName=safe_name, ref=ref)
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False, showLastColumn=False,
        showRowStripes=True,   showColumnStripes=False,
    )
    ws.add_table(tab)
    return next_row


def escribir_dashboard(
    wb,
    bds: dict,
    bds_sku_splits: dict,
    mes_actual: str,
) -> None:
    ws = wb.create_sheet("Resumen", 0)

    # Título principal
    ws.merge_cells("A1:L1")
    c = ws.cell(
        row=1, column=1,
        value=f"Dashboard - Seguimiento de Impacto por Segmento - {mes_actual}",
    )
    c.fill      = PatternFill("solid", fgColor=COLOR_HEADER)
    c.font      = Font(bold=True, color="FFFFFF", size=14)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    ts = ws.cell(row=2, column=1,
                 value=f"Generado: {datetime.today().strftime('%Y-%m-%d %H:%M')}")
    ts.font = Font(italic=True, size=9, color="888888")

    # KPIs
    total_pdvs = sum(len(df) for df in bds.values())
    total_imp  = sum(
        int((df["tiene_impacto"] == "Sí").sum())
        for df in bds.values() if "tiene_impacto" in df.columns
    )
    pct_global = round(total_imp / total_pdvs * 100, 1) if total_pdvs > 0 else 0.0

    kpis = [
        ("Segmentos NPI",    len(bds)),
        ("Total PDVs",       f"{total_pdvs:,}"),
        ("Con Impacto",      f"{total_imp:,}"),
        ("% Impacto Global", f"{pct_global}%"),
        ("Periodo Actual",   mes_actual),
    ]
    for k, (label, value) in enumerate(kpis):
        c1, c2 = k * 2 + 1, k * 2 + 2
        ws.merge_cells(start_row=4, start_column=c1, end_row=4, end_column=c2)
        lbl = ws.cell(row=4, column=c1, value=label)
        lbl.fill      = PatternFill("solid", fgColor=COLOR_SECCION)
        lbl.font      = Font(bold=True, color="FFFFFF", size=9)
        lbl.alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells(start_row=5, start_column=c1, end_row=5, end_column=c2)
        val = ws.cell(row=5, column=c1, value=value)
        val.fill      = PatternFill("solid", fgColor=COLOR_KPI_BG)
        val.font      = Font(bold=True, size=13, color=COLOR_HEADER)
        val.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[4].height = 18
    ws.row_dimensions[5].height = 30

    cur_row = 7

    # 1. Resumen por segmento
    cur_row = _titulo_seccion(ws, cur_row, "1. Resumen por Segmento", 12)
    df_seg  = _resumen_por_segmento(bds, mes_actual)
    cur_row = _escribir_tabla_dyn(ws, df_seg, cur_row, "ResumenSegmento") + 1

    # Subtablas por SKU
    for seg_nombre, skus in bds_sku_splits.items():
        ws.merge_cells(start_row=cur_row, start_column=1,
                        end_row=cur_row,   end_column=8)
        sub = ws.cell(row=cur_row, column=1,
                      value=f"{seg_nombre} - Desglose por SKU")
        sub.fill      = PatternFill("solid", fgColor=COLOR_CYAN_ACC)
        sub.font      = Font(bold=True, color="FFFFFF", size=10)
        sub.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[cur_row].height = 18
        cur_row += 1

        for sku_nombre, sku_df in skus.items():
            sku_total   = len(sku_df)
            sku_imp     = int((sku_df["tiene_impacto"] == "Sí").sum())
            sku_pct     = round(sku_imp / sku_total * 100, 1) if sku_total > 0 else 0.0
            sku_mes_act = int((sku_df["venta_mes_actual"] == "Sí").sum())
            df_sku = pd.DataFrame([{
                "SKU":                   sku_nombre,
                "PDVs Target":           sku_total,
                "Con Impacto":           sku_imp,
                "Sin Impacto":           sku_total - sku_imp,
                "% Impacto":             sku_pct,
                f"Impacto {mes_actual}": sku_mes_act,
            }])
            safe = "SKU_" + "".join(c if c.isalnum() else "_" for c in sku_nombre)
            cur_row = _escribir_tabla_dyn(ws, df_sku, cur_row, safe) + 1

    # 2. Tendencia mensual
    cur_row = _titulo_seccion(ws, cur_row, "2. Tendencia de Impacto Mensual", 12)
    df_tend = _tendencia_mensual(bds)
    if not df_tend.empty:
        cur_row = _escribir_tabla_dyn(ws, df_tend, cur_row, "TendenciaMensual") + 1
    else:
        ws.cell(row=cur_row, column=1, value="(sin datos de periodos disponibles)")
        cur_row += 2

    # 3. Resumen por zona (tabla por segmento, filtrable)
    cur_row = _titulo_seccion(ws, cur_row, "3. Resumen por Zona", 12)
    sups_por_seg = _resumen_supervisor_por_segmento(bds, mes_actual)
    for nombre, df_s in sups_por_seg.items():
        n_cols_sub = max(len(df_s.columns), 6) if not df_s.empty else 6
        ws.merge_cells(start_row=cur_row, start_column=1,
                        end_row=cur_row,   end_column=n_cols_sub)
        sub = ws.cell(row=cur_row, column=1, value=nombre)
        sub.fill      = PatternFill("solid", fgColor=COLOR_CYAN_ACC)
        sub.font      = Font(bold=True, color="FFFFFF", size=10)
        sub.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[cur_row].height = 18
        cur_row += 1

        if df_s.empty:
            ws.cell(row=cur_row, column=1,
                    value="('SUPERVISOR RESPONSABLE' no encontrada en el rutero)")
            cur_row += 2
        else:
            safe_name = "Sup_" + "".join(c if c.isalnum() else "_" for c in nombre)
            cur_row = _escribir_tabla_dyn(ws, df_s, cur_row, safe_name) + 1

    # Anchos de columna
    col_widths = {1: 35, 2: 14, 3: 14, 4: 14, 5: 12, 6: 18}
    for col_idx, w in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    for col_idx in range(7, 20):
        ws.column_dimensions[get_column_letter(col_idx)].width = 13

    ws.freeze_panes = "A6"


# ─────────────────────────────────────────────────────────────────────────────
# 4. ESCRITURA
# ─────────────────────────────────────────────────────────────────────────────

def escribir_excel(
    bds: dict,
    bds_detalle: dict,
    bds_sku_splits: dict,
    salida: str,
    mes_actual: str,
) -> None:
    """
    Construye el Excel en memoria (BytesIO) y lo sube a SharePoint en
    `salida` (ruta SharePoint tipo 'Equipo Información/.../impacto_segmentos.xlsx').
    Ya no escribe a disco local.
    """
    buffer = io.BytesIO()

    # 1) Escribir hojas de detalle
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        for nombre, df in bds_detalle.items():
            df.to_excel(writer, sheet_name=nombre, index=False)
    buffer.seek(0)

    # 2) Formato + dashboard (openpyxl puede recargar desde el mismo buffer)
    wb = load_workbook(buffer)
    for nombre, df in bds_detalle.items():
        aplicar_formato(wb[nombre], df)
    escribir_dashboard(wb, bds, bds_sku_splits, mes_actual)

    buffer_final = io.BytesIO()
    wb.save(buffer_final)
    contenido = buffer_final.getvalue()

    _subir_bytes_cloud(
        contenido, salida,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    print(f"\n✅ Archivo generado y subido a SharePoint: {salida}")
    print(f"\n   Resumen por pestaña (período {mes_actual}):")
    print(f"   {'Segmento':<32} {'PDVs':>8} {'Con impacto':>13} {'Sin impacto':>13}")
    print(f"   {'-' * 70}")
    for nombre, df in bds_detalle.items():
        total = len(df)
        con   = int((df["tiene_impacto"] == "Sí").sum()) if "tiene_impacto" in df.columns else 0
        sin   = total - con
        print(f"   {nombre:<32} {total:>8,} {con:>13,} {sin:>13,}")


# ─────────────────────────────────────────────────────────────────────────────
# 5. PUNTO DE ENTRADA
# ─────────────────────────────────────────────────────────────────────────────

def run() -> None:
    print("=" * 55)
    print("  ETL — Impacto por Segmento (V2, Cloud)")
    print("=" * 55)

    ruta_salida_cloud = f"{paths.RUTA_CARPETA_SALIDAS_DYP}/impacto_segmentos.xlsx"

    print("\n[1/5] Descargando rutero desde SharePoint...")
    try:
        rutero_bytes = _descargar_rutero_cloud()
    except Exception as e:
        print(f"❌ {e}")
        return
    rutero = cargar_rutero(io.BytesIO(rutero_bytes))
    print(f"      {len(rutero):,} PDVs en el rutero")

    print("\n[2/5] Descargando listas de referencia desde SharePoint...")
    try:
        listas_bytes = _descargar_listas_cloud()
    except Exception as e:
        print(f"❌ Falta el archivo de listas en SharePoint: {e}\n"
              f"   Hojas requeridas: MSL + {', '.join(SEGMENTOS_PRODNUEVOS)}")
        return
    listas = cargar_listas(io.BytesIO(listas_bytes))
    for nombre, datos in listas.items():
        pdv_info = (
            f"{len(datos['pdvs']):,} PDVs target"
            if datos["pdvs"] is not None else "todos los PDVs del rutero"
        )
        print(f"      {nombre:<20}: {len(datos['eans']):>3} EANs | {pdv_info}")

    print("\n[3/5] Descargando ventas consolidadas desde SharePoint...")
    try:
        ventas_bytes_list = _descargar_ventas_historico_cloud()
    except Exception as e:
        print(f"❌ {e}")
        return
    ventas, desc_map = cargar_ventas_desde_csv(ventas_bytes_list)
    print(f"      {len(ventas):,} registros (PDV × EAN × periodo)")
    print(f"      {len(desc_map):,} descripciones de producto capturadas")

    if ventas.empty:
        print("❌ El consolidado de ventas está vacío — nada que cruzar.")
        return

    # Período más reciente
    ventas["_sort"] = ventas.apply(
        lambda r: periodo_sort_key(f"{r['mes']}-{r['año']}"), axis=1
    )
    fila_max  = ventas.loc[ventas["_sort"].idxmax()]
    mes_actual = f"{fila_max['mes']}-{fila_max['año']}"
    ventas.drop(columns=["_sort"], inplace=True)
    print(f"      Período más reciente: {mes_actual}")

    print("\n[4/5] Construyendo BDs de impacto...")
    bds            = {}   # Para el dashboard (siempre por segmento completo)
    bds_detalle    = {}   # Para las pestañas de detalle (split por EAN si aplica)
    bds_sku_splits = {}   # {segmento: {sku_tab: df}} solo para segmentos con split

    for nombre, datos in listas.items():
        print(f"      → {nombre}...")
        df_seg = construir_bd(
            rutero        = rutero,
            ventas        = ventas,
            eans_segmento = datos["eans"],
            pdvs_target   = datos["pdvs"],
            mes_actual    = mes_actual,
        )
        bds[nombre] = df_seg

        eans = datos["eans"]
        if 1 < len(eans) <= SPLIT_SKU_MAX:
            bds_sku_splits[nombre] = {}
            for ean in sorted(eans):
                auto = _make_tab_name(nombre, desc_map.get(ean, ""), ean)
                tab  = CUSTOM_TAB_NAMES.get(auto, auto)
                print(f"         → SKU: {tab}")
                sku_df = construir_bd(
                    rutero        = rutero,
                    ventas        = ventas,
                    eans_segmento = {ean},
                    pdvs_target   = datos["pdvs"],
                    mes_actual    = mes_actual,
                )
                bds_detalle[tab]            = sku_df
                bds_sku_splits[nombre][tab] = sku_df
        else:
            bds_detalle[nombre] = df_seg

    print("\n[5/5] Escribiendo Excel + dashboard en SharePoint...")
    escribir_excel(bds, bds_detalle, bds_sku_splits, ruta_salida_cloud, mes_actual)


if __name__ == "__main__":
    run()