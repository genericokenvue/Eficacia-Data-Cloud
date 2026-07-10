"""
detector_anomalias.py
─────────────────────
Detecta anomalías en las salidas del ETL y genera el Excel de revisión que
alimenta la app web de supervisores (ALERTAS/REVISION/app_revision.html).

Las 14 reglas (CONTEXTO_PROYECTO §8) están agrupadas en 5 módulos. La sexta
posible (EXHIBICIONES_GRATIS) queda intencionalmente fuera porque la única
señal posible es "falta 2da captura" y es informativa, no error que requiera
decisión humana.

Outputs
───────
ALERTAS/REVISION/Casos_Revision_<MM>_<YYYY>.xlsx
    Hoja "Casos"     → tabla TblCasos con los casos para revisión humana.
    Hoja "Historial" → tabla TblHistorial con correcciones de meses previos
                       (vacío hasta que aplicar_correcciones.py la pueble).
ALERTAS/REVISION/auto_resueltos_<MM>_<YYYY>.csv
    Anomalías que se auto-resuelven por reglas de negocio (no requieren
    intervención humana). Sirve para auditoría.

Convenciones
────────────
• MERCADERISTA_O_MARCA  → nombre del gestor (CIF, PRECIOS, SOS, EXH pagadas)
                          o la marca afectada (NO_PRESENCIA por marca).
• SUPERVISOR            → SUPERVISOR_LIDER canónico (UPPER, strip).
                          Si en un PATRON cubre PDVs de varios supervisores,
                          se usa "MÚLTIPLE" como en histórico de Marzo 2026.
• PDVS_AFECTADOS        → string "PDV1 | PDV2 | ... | + N más" si > 5 PDVs.
• ID_CASO               → "<MODULO_PREFIX>-<TIPO_PREFIX>-<corr>" — único por
                          archivo. Estable mes a mes solo si el upsert
                          encuentra match (id_caso + mes + año).
• ID_ENCUESTA           → para SOS, se mete en METADATA (JSON) para que
                          aplicar_correcciones.py pueda replicar el universo
                          a todas las filas del mismo ID_ENCUESTA.

Volumen esperado
────────────────
~30-300 casos/mes (depende del mes). En Marzo 2026 fueron 226.
"""

from __future__ import annotations

import os
import sys
import json
import argparse
from datetime import datetime
from pathlib import Path

# Forzar UTF-8 en stdout para que los caracteres box-drawing (═, ─, ▶)
# no exploten en Windows con cp1252.
try:
    sys.stdout.reconfigure(encoding="utf-8")  # Python 3.7+
except AttributeError:
    pass

import pandas as pd
import numpy as np

# alertas_logger ya añade SCRIPTS al sys.path al importarse desde ALERTAS,
# pero como este script vive en SCRIPTS, lo aseguramos manualmente.
_SCRIPTS_DIR = Path(__file__).resolve().parent
if str(_SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(_SCRIPTS_DIR))

import paths


# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURACIÓN
# ─────────────────────────────────────────────────────────────────────────────
DIR_OUT = paths.ALERTAS_DIR / "REVISION"
DIR_OUT.mkdir(parents=True, exist_ok=True)

# Cota auto-resolución CIF: tiempos por encima de esto se consideran
# acumulación de turnos del día anterior, no error a revisar.
CIF_TIEMPO_AUTORESUELTO_H = 18.0
CIF_TIEMPO_UMBRAL_H        = 9.0

# Materialidad SOS — outlier solo si la marca supera este umbral en cms
SOS_MATERIALIDAD_CMS = 500.0
SOS_TUKEY_K          = 3.0
SOS_MIN_OBS_CATEG    = 15

# R8 — cms=0 con target>0: solo se reporta como PATRON cuando un gestor
# tiene >=N PDVs con esa marca en 0. Casos puntuales son ruido normal
# (la marca no se midió o no estaba), no anomalía.
SOS_R8_MIN_PDVS_PATRON = 10

# Precios — factor multiplicador sobre p99
PRECIOS_FACTOR_P99 = 2.0
PRECIOS_MIN_OBS_SUBCANAL = 5

# Exhibiciones pagadas — factor sobre cantidad negociada
EXH_FACTOR_SOBREEJECUCION = 1.5

# Columnas de salida del Excel (17 — incluye METADATA al final)
COLS_CASO = [
    "ID_CASO", "SUPERVISOR", "MERCADERISTA_O_MARCA",
    "MODULO", "TIPO", "NIVEL",
    "CAUSA", "DESCRIPCION", "N_AFECTADOS",
    "VALOR_ORIGINAL", "PDVS_AFECTADOS",
    "MES", "ANIO",
    "ESTADO", "DECISION", "VALOR_CORRECTO", "OBSERVACION",
    "METADATA",  # JSON con info auxiliar (ID_ENCUESTA, etc.) — no se ve en la app
]

COLS_HISTORIAL = [
    "ID_CORRECCION", "ID_PDV", "NOMBRE_PDV", "MES", "ANIO",
    "MODULO", "CAUSA", "DECISION",
    "VALOR_ORIGINAL", "VALOR_CORRECTO",
    "APROBADO_POR", "FECHA", "ID_CASO_ORIGEN",
]

# Prefijos de ID_CASO por módulo + tipo
PREFIJO_MODULO = {
    "CIF":                  "CIF",
    "NO_PRESENCIA":         "NP",
    "SOS":                  "SOS",
    "PRECIOS":              "PR",
    "EXHIBICIONES_PAGADAS": "EX-P",
    "EXHIBICIONES_GRATIS":  "EX-G",
}


# ─────────────────────────────────────────────────────────────────────────────
# UTILIDADES
# ─────────────────────────────────────────────────────────────────────────────

def _norm_str(s) -> str:
    """Strip + upper para nombres / IDs. NaN → ''."""
    if s is None:
        return ""
    if isinstance(s, float) and pd.isna(s):
        return ""
    return str(s).strip().upper()


def _norm_id(s) -> str:
    """ID a string sin trailing .0 (typical de Excel)."""
    if s is None:
        return ""
    try:
        if pd.isna(s):
            return ""
    except (TypeError, ValueError):
        pass
    out = str(s).strip()
    if out.endswith(".0"):
        out = out[:-2]
    return out


def _pdvs_lista(pdvs: list, max_visibles: int = 5) -> str:
    """'PDV1 | PDV2 | + N más' — formato canónico del histórico."""
    pdvs = [p for p in pdvs if p]
    if not pdvs:
        return ""
    if len(pdvs) <= max_visibles:
        return " | ".join(pdvs)
    extra = len(pdvs) - max_visibles
    return " | ".join(pdvs[:max_visibles]) + f" | + {extra} más"


def _gen_id(prefijo: str, tipo: str, contador: int) -> str:
    """Genera 'CIF-S-0023' o 'NP-0198' (PATRON con S/P opcional)."""
    n = f"{contador:04d}"
    if tipo == "PATRON" and prefijo == "CIF":
        return f"{prefijo}-S-{n}"
    if tipo == "PATRON" and prefijo == "NP":
        return f"{prefijo}-{n}"
    return f"{prefijo}-{n}"


def _periodo_default() -> tuple[int, int]:
    """Detecta mes/año del archivo CIF más reciente; fallback al actual."""
    try:
        df = pd.read_excel(paths.CIF_OUT_FINAL, usecols=["MES", "AÑO"], nrows=10000)
        m = int(pd.to_numeric(df["MES"], errors="coerce").dropna().mode().iloc[0])
        a = int(pd.to_numeric(df["AÑO"], errors="coerce").dropna().mode().iloc[0])
        return m, a
    except Exception:
        now = datetime.now()
        return now.month, now.year


def _resolver_supervisor(pdvs_sup: list[str]) -> str:
    """
    Si todos los PDVs del caso pertenecen al mismo supervisor → ese nombre.
    Si hay 2+ supervisores distintos → 'MÚLTIPLE'.
    """
    unicos = {s for s in pdvs_sup if s}
    if len(unicos) == 1:
        return next(iter(unicos))
    if len(unicos) == 0:
        return "SIN_SUPERVISOR"
    return "MÚLTIPLE"


# Lookups globales (lazy) ─────────────────────────────────────────────────
_PT_CACHE: dict | None = None


def _cargar_pt_lookups() -> dict:
    """
    Carga (una vez) los lookups derivados de Plan de trabajo:
      sup_por_nombre   → NOMBRE_GESTOR (upper) → SUPERVISOR_LIDER (upper)
      sup_por_pdv      → ID_PDV (str) → SUPERVISOR_LIDER (upper)
      acron_por_pdv    → ID_PDV → ACRONIMO (para derivar SUBCANAL)
      subcanal_por_pdv → ID_PDV → SUBCANAL (prefijo de letras del ACRONIMO)
      nombrepdv_por_pdv → ID_PDV → NOMBRE_PDV
    """
    global _PT_CACHE
    if _PT_CACHE is not None:
        return _PT_CACHE

    pt = pd.read_excel(paths.CIF_OUT_FINAL,
                       usecols=["ID_PDV_INVOLVES", "NOMBRE", "NOMBRE_PDV",
                                "ROL", "SUPERVISOR_LIDER", "ACRONIMO"])
    pt["ID_PDV_INVOLVES"]  = pt["ID_PDV_INVOLVES"].map(_norm_id)
    pt["NOMBRE"]           = pt["NOMBRE"].map(_norm_str)
    pt["NOMBRE_PDV"]       = pt["NOMBRE_PDV"].map(_norm_str)
    pt["SUPERVISOR_LIDER"] = pt["SUPERVISOR_LIDER"].map(_norm_str)
    pt["ACRONIMO"]         = pt["ACRONIMO"].astype(str).str.strip().str.upper()
    pt["SUBCANAL"]         = pt["ACRONIMO"].str.extract(r"^([A-Za-z]+)")[0]

    gestores = pt[pt["ROL"].astype(str).str.upper() == "GESTOR"]
    sup_por_nombre = (
        gestores[gestores["SUPERVISOR_LIDER"] != ""]
        .drop_duplicates(subset=["NOMBRE"])
        .set_index("NOMBRE")["SUPERVISOR_LIDER"].to_dict()
    )
    pt_dedup = pt.drop_duplicates(subset=["ID_PDV_INVOLVES"])
    sup_por_pdv       = pt_dedup.set_index("ID_PDV_INVOLVES")["SUPERVISOR_LIDER"].to_dict()
    acron_por_pdv     = pt_dedup.set_index("ID_PDV_INVOLVES")["ACRONIMO"].to_dict()
    subcanal_por_pdv  = pt_dedup.set_index("ID_PDV_INVOLVES")["SUBCANAL"].to_dict()
    nombrepdv_por_pdv = pt_dedup.set_index("ID_PDV_INVOLVES")["NOMBRE_PDV"].to_dict()

    _PT_CACHE = {
        "sup_por_nombre": sup_por_nombre,
        "sup_por_pdv":    sup_por_pdv,
        "acron_por_pdv":  acron_por_pdv,
        "subcanal_por_pdv": subcanal_por_pdv,
        "nombrepdv_por_pdv": nombrepdv_por_pdv,
    }
    return _PT_CACHE


def _resolver_sup_con_fallback(sup_directo: str,
                                empleado: str,
                                id_pdv: str = "") -> str:
    """
    Cascada de fallbacks para conseguir el supervisor de un caso:
      1) sup_directo si no está vacío
      2) Lookup por nombre del empleado en PT
      3) Lookup por ID_PDV en PT
      4) 'SIN_SUPERVISOR'
    """
    if sup_directo:
        return sup_directo
    lk = _cargar_pt_lookups()
    if empleado:
        s = lk["sup_por_nombre"].get(empleado, "")
        if s:
            return s
    if id_pdv:
        s = lk["sup_por_pdv"].get(id_pdv, "")
        if s:
            return s
    return "SIN_SUPERVISOR"


# ─────────────────────────────────────────────────────────────────────────────
# DETECTOR — CIF
# ─────────────────────────────────────────────────────────────────────────────

def detectar_cif(mes: int, anio: int) -> tuple[list[dict], list[dict]]:
    """
    Reglas CIF:
      R1) Tiempo de servicio > 9h en una visita individual (REVISIÓN, PATRON).
          Tiempos > 18h se auto-resuelven (acumulación de turnos).
      R2) Sobrecumplimiento de visitas: VISITAS_REAL > CANTIDAD_VISITAS (REVISIÓN, PATRON).

    Retorna (casos, auto_resueltos).
    """
    casos: list[dict] = []
    auto: list[dict] = []

    # ── R1: tiempo > 9h (con auto-resolución > 18h) ──────────────────────
    csv_visitas = paths.CIF_BASES / "INVOLVES" / "informe_visitas_procesado.csv"
    if csv_visitas.exists():
        df_v = pd.read_csv(csv_visitas, sep=";", encoding="utf-8-sig", low_memory=False)
        df_v.columns = [c.lstrip("﻿") for c in df_v.columns]
        if "TIEMPO_SERVICIO_MIN" in df_v.columns:
            df_v["tiempo_h"] = pd.to_numeric(df_v["TIEMPO_SERVICIO_MIN"], errors="coerce") / 60
            mask_critico = df_v["tiempo_h"] > CIF_TIEMPO_AUTORESUELTO_H
            mask_revisar = (df_v["tiempo_h"] > CIF_TIEMPO_UMBRAL_H) & (~mask_critico)

            # Auto-resoluciones
            for _, r in df_v[mask_critico].iterrows():
                auto.append({
                    "REGLA": "CIF tiempo > 18h",
                    "MODULO": "CIF",
                    "ID_PDV": _norm_id(r.get("ID_PDV_INVOLVES")),
                    "PDV":    _norm_str(r.get("PUNTO_DE_VENTA")),
                    "EMPLEADO": _norm_str(r.get("EMPLEADO")),
                    "FECHA":   str(r.get("FECHA_VISITA", "")),
                    "VALOR_ORIGINAL": f"{r['tiempo_h']:.1f}h",
                    "AUTO_RESOLUCION": "Acumulación de turnos (tiempo > 18h)",
                })

            # Casos para revisión: agrupar por EMPLEADO (PATRON)
            df_r1 = df_v[mask_revisar].copy()
            if not df_r1.empty:
                # Necesitamos cruzar con PT para conseguir el supervisor del gestor
                pt = pd.read_excel(paths.CIF_OUT_FINAL,
                                   usecols=["NOMBRE", "SUPERVISOR_LIDER",
                                            "ID_PDV_INVOLVES", "ROL"])
                pt["NOMBRE"]            = pt["NOMBRE"].map(_norm_str)
                pt["SUPERVISOR_LIDER"]  = pt["SUPERVISOR_LIDER"].map(_norm_str)
                pt["ID_PDV_INVOLVES"]   = pt["ID_PDV_INVOLVES"].map(_norm_id)
                sup_por_nombre = (
                    pt[pt["ROL"].astype(str).str.upper()=="GESTOR"]
                    .drop_duplicates(subset=["NOMBRE"])
                    .set_index("NOMBRE")["SUPERVISOR_LIDER"].to_dict()
                )

                df_r1["EMPLEADO_N"] = df_r1["EMPLEADO"].map(_norm_str)
                df_r1["ID_PDV"]     = df_r1["ID_PDV_INVOLVES"].map(_norm_id)
                df_r1["PDV"]        = df_r1["PUNTO_DE_VENTA"].map(_norm_str)

                for empleado, grupo in df_r1.groupby("EMPLEADO_N"):
                    pdvs_unicos = grupo["PDV"].dropna().unique().tolist()
                    horas_max   = grupo["tiempo_h"].max()
                    horas_min   = grupo["tiempo_h"].min()
                    sup = sup_por_nombre.get(empleado, "SIN_SUPERVISOR")
                    casos.append({
                        "MODULO": "CIF",
                        "TIPO":   "PATRON",
                        "NIVEL":  "REVISION",
                        "SUPERVISOR": sup,
                        "MERCADERISTA_O_MARCA": empleado,
                        "CAUSA": f"Tiempo de visita entre {CIF_TIEMPO_UMBRAL_H:.0f}-{CIF_TIEMPO_AUTORESUELTO_H:.0f}h",
                        "DESCRIPCION": (
                            f"{empleado} registró {len(grupo)} visita(s) con tiempo de "
                            f"servicio entre {CIF_TIEMPO_UMBRAL_H:.0f}h y {CIF_TIEMPO_AUTORESUELTO_H:.0f}h "
                            f"(rango: {horas_min:.1f}h–{horas_max:.1f}h). "
                            f"Posible error de digitación o jornada extendida."
                        ),
                        "N_AFECTADOS": len(grupo),
                        "VALOR_ORIGINAL": f"{horas_min:.1f}h–{horas_max:.1f}h",
                        "PDVS_AFECTADOS": _pdvs_lista(pdvs_unicos),
                        "MES": mes, "ANIO": anio,
                        "METADATA": {"regla": "R1", "n_visitas": int(len(grupo))},
                    })

    # ── R2: sobrecumplimiento de visitas ─────────────────────────────────
    df_pt = pd.read_excel(paths.CIF_OUT_FINAL,
                          usecols=["NOMBRE", "SUPERVISOR_LIDER", "ROL",
                                   "ID_PDV_INVOLVES", "NOMBRE_PDV",
                                   "CANTIDAD_VISITAS", "VISITAS_REAL"])
    df_pt["NOMBRE"]           = df_pt["NOMBRE"].map(_norm_str)
    df_pt["SUPERVISOR_LIDER"] = df_pt["SUPERVISOR_LIDER"].map(_norm_str)
    df_pt["NOMBRE_PDV"]       = df_pt["NOMBRE_PDV"].map(_norm_str)
    df_pt["CANTIDAD_VISITAS"] = pd.to_numeric(df_pt["CANTIDAD_VISITAS"], errors="coerce")
    df_pt["VISITAS_REAL"]     = pd.to_numeric(df_pt["VISITAS_REAL"], errors="coerce")

    es_gestor = df_pt["ROL"].astype(str).str.upper() == "GESTOR"
    sob = df_pt[
        es_gestor
        & (df_pt["CANTIDAD_VISITAS"] > 0)
        & (df_pt["VISITAS_REAL"] > df_pt["CANTIDAD_VISITAS"])
    ].copy()

    for nombre, grupo in sob.groupby("NOMBRE"):
        if not nombre:
            continue
        pdvs = grupo["NOMBRE_PDV"].dropna().unique().tolist()
        sup_unicos = grupo["SUPERVISOR_LIDER"].dropna().unique().tolist()
        total_plan = int(grupo["CANTIDAD_VISITAS"].sum())
        total_real = int(grupo["VISITAS_REAL"].sum())
        # Si el "patrón" cubre 1 solo PDV → PUNTUAL (no es realmente patrón)
        tipo_caso = "PATRON" if len(grupo) > 1 else "PUNTUAL"
        sup_resuelto = _resolver_supervisor(sup_unicos)
        if sup_resuelto == "SIN_SUPERVISOR":
            sup_resuelto = _resolver_sup_con_fallback("", nombre)
        casos.append({
            "MODULO": "CIF",
            "TIPO":   tipo_caso,
            "NIVEL":  "REVISION",
            "SUPERVISOR": sup_resuelto,
            "MERCADERISTA_O_MARCA": nombre,
            "CAUSA": "Sobrecumplimiento de visitas planeadas",
            "DESCRIPCION": (
                f"{nombre} ejecutó {total_real} visitas vs {total_plan} planeadas en "
                f"{len(grupo)} PDV(s). Posible ruta ampliada o error de planeación."
            ),
            "N_AFECTADOS": len(grupo),
            "VALOR_ORIGINAL": f"{total_real} reales / {total_plan} planeadas",
            "PDVS_AFECTADOS": _pdvs_lista(pdvs),
            "MES": mes, "ANIO": anio,
            "METADATA": {"regla": "R2", "visitas_plan": total_plan, "visitas_real": total_real},
        })

    return casos, auto


# ─────────────────────────────────────────────────────────────────────────────
# DETECTOR — NO PRESENCIA
# ─────────────────────────────────────────────────────────────────────────────

def detectar_np(mes: int, anio: int) -> tuple[list[dict], list[dict]]:
    """
    R3) 100% agotado en una marca×PDV (REVISIÓN, agrupado por marca → PATRON).
    R4) agotados > productos_medidos (CRÍTICO, PUNTUAL).
    """
    casos: list[dict] = []
    auto: list[dict] = []

    f_agotados = paths.NP_SALIDA / f"ANALISIS_AGOTADOS_{_nombre_mes(mes)}_{anio}.xlsx"
    if not f_agotados.exists():
        # fallback: tomar el más reciente
        candidatos = sorted(paths.NP_SALIDA.glob("ANALISIS_AGOTADOS_*.xlsx"))
        if candidatos:
            f_agotados = candidatos[-1]
        else:
            return casos, auto

    df = pd.read_excel(f_agotados)
    df["ID del PDV"] = df["ID del PDV"].map(_norm_id)
    df["PDV"]        = df["PDV"].map(_norm_str)
    df["Marca"]      = df["Marca"].map(_norm_str)
    df["PRODUCTOS_MEDIDOS"] = pd.to_numeric(df["PRODUCTOS_MEDIDOS"], errors="coerce").fillna(0)
    df["CANT_AGOTADOS"]    = pd.to_numeric(df["CANT_AGOTADOS"], errors="coerce").fillna(0)
    df["%_AGOTADOS"]       = pd.to_numeric(df["%_AGOTADOS"], errors="coerce")

    # Cruce para conseguir supervisor por PDV (desde Plan de trabajo)
    pt = pd.read_excel(paths.CIF_OUT_FINAL,
                       usecols=["ID_PDV_INVOLVES", "SUPERVISOR_LIDER"])
    pt["ID_PDV_INVOLVES"]  = pt["ID_PDV_INVOLVES"].map(_norm_id)
    pt["SUPERVISOR_LIDER"] = pt["SUPERVISOR_LIDER"].map(_norm_str)
    sup_por_pdv = (
        pt.drop_duplicates(subset=["ID_PDV_INVOLVES"])
        .set_index("ID_PDV_INVOLVES")["SUPERVISOR_LIDER"].to_dict()
    )

    # ── R3: 100% agotado, agrupar por MARCA ──────────────────────────────
    cien = df[df["%_AGOTADOS"] >= 1.0].copy()
    cien["SUP"] = cien["ID del PDV"].map(lambda i: sup_por_pdv.get(i, ""))

    for marca, grupo in cien.groupby("Marca"):
        if not marca:
            continue
        pdvs = grupo["PDV"].dropna().unique().tolist()
        sups = grupo["SUP"].dropna().unique().tolist()
        productos_max = int(grupo["PRODUCTOS_MEDIDOS"].max())
        casos.append({
            "MODULO": "NO_PRESENCIA",
            "TIPO":   "PATRON",
            "NIVEL":  "REVISION",
            "SUPERVISOR": _resolver_supervisor(sups),
            "MERCADERISTA_O_MARCA": marca,
            "CAUSA": "Marca con 100% de SKUs agotados",
            "DESCRIPCION": (
                f"La marca {marca} aparece con 100% de productos agotados en "
                f"{len(grupo)} PDV(s). Verificar si es desabastecimiento real "
                f"o error sistemático de captura."
            ),
            "N_AFECTADOS": len(grupo),
            "VALOR_ORIGINAL": f"100% agotado (hasta {productos_max} SKUs)",
            "PDVS_AFECTADOS": _pdvs_lista(pdvs),
            "MES": mes, "ANIO": anio,
            "METADATA": {"regla": "R3", "max_productos_medidos": productos_max},
        })

    # ── R4: agotados > medidos (CRÍTICO, PUNTUAL) ────────────────────────
    inco = df[df["CANT_AGOTADOS"] > df["PRODUCTOS_MEDIDOS"]].copy()
    for _, r in inco.iterrows():
        sup = sup_por_pdv.get(r["ID del PDV"], "")
        casos.append({
            "MODULO": "NO_PRESENCIA",
            "TIPO":   "PUNTUAL",
            "NIVEL":  "CRITICO",
            "SUPERVISOR": sup or "SIN_SUPERVISOR",
            "MERCADERISTA_O_MARCA": r["Marca"],
            "CAUSA": "Inconsistencia: agotados > medidos",
            "DESCRIPCION": (
                f"En {r['PDV']} se reportan {int(r['CANT_AGOTADOS'])} SKUs agotados "
                f"de {r['Marca']} cuando solo se midieron {int(r['PRODUCTOS_MEDIDOS'])} "
                f"productos. Error de captura."
            ),
            "N_AFECTADOS": 1,
            "VALOR_ORIGINAL": f"{int(r['CANT_AGOTADOS'])} agotados / {int(r['PRODUCTOS_MEDIDOS'])} medidos",
            "PDVS_AFECTADOS": r["PDV"],
            "MES": mes, "ANIO": anio,
            "METADATA": {"regla": "R4", "id_pdv": r["ID del PDV"]},
        })

    return casos, auto


# ─────────────────────────────────────────────────────────────────────────────
# DETECTOR — SOS  (5 reglas)
# ─────────────────────────────────────────────────────────────────────────────

def detectar_sos(mes: int, anio: int) -> tuple[list[dict], list[dict]]:
    """
    R5)  cms_marca > cms_universo                (CRÍTICO, PUNTUAL)
    R6)  IQR Tukey k=3 con cms_marca > 500       (REVISIÓN, PUNTUAL)
    R7)  universo=0 con marca>0                  (CRÍTICO, PUNTUAL)
    R8)  cms=0 con target>0                      (REVISIÓN, PUNTUAL)
    R9)  participacion > 100%                    (CRÍTICO, PUNTUAL)
    """
    casos: list[dict] = []
    auto: list[dict] = []

    # Reporte calculado (con SUBCANAL, TARGET, PARTICIPACION)
    f_calc = paths.SOS_SALIDA / "Reporte_SOS_Final_Calculado.xlsx"
    if not f_calc.exists():
        return casos, auto
    df = pd.read_excel(f_calc)

    # Encuesta cruda (para ID de la encuesta + Superior del empleado)
    f_cruda = paths.SOS_BASES / "Encuesta Sos Consolidada.xlsx"
    if f_cruda.exists():
        cruda = pd.read_excel(f_cruda, usecols=[
            "ID de la encuesta", "ID del PDV", "Marca",
            "Empleado", "Superior del empleado",
            "Categoría de producto",
        ])
        cruda["ID del PDV"] = cruda["ID del PDV"].map(_norm_id)
        cruda["Marca"]      = cruda["Marca"].map(_norm_str)
        cruda["CATEGORIA"]  = cruda["Categoría de producto"].map(_norm_str)
        cruda["EMPLEADO"]   = cruda["Empleado"].map(_norm_str)
        cruda["SUPERIOR"]   = cruda["Superior del empleado"].map(_norm_str)
        # Llave compuesta: PDV + Marca + Categoría
        cruda["_K"] = cruda["ID del PDV"] + "|" + cruda["Marca"] + "|" + cruda["CATEGORIA"]
        meta_lookup = cruda.drop_duplicates(subset=["_K"]).set_index("_K")[
            ["ID de la encuesta", "EMPLEADO", "SUPERIOR"]
        ].to_dict("index")
    else:
        meta_lookup = {}

    # Normalización del calculado
    df["ID INVOLVES"]   = df["ID INVOLVES"].map(_norm_id)
    df["PDV"]           = df["PUNTO DE VENTA"].map(_norm_str)
    df["MARCA"]         = df["MARCA"].map(_norm_str)
    df["CATEG"]         = df["CATEGORÍA DE PRODUCTO"].map(_norm_str)
    df["UNIV"]          = pd.to_numeric(df["¿Cual es el universo en cms de la categoria?"], errors="coerce")
    df["MARCA_CMS"]     = pd.to_numeric(df["¿Cuántos cms tiene la marca?"], errors="coerce")
    df["TARGET"]        = pd.to_numeric(df["TARGET"], errors="coerce")
    df["PARTICIP"]      = pd.to_numeric(df["PARTICIPACION_SOS"], errors="coerce")
    df["_K"]            = df["ID INVOLVES"] + "|" + df["MARCA"] + "|" + df["CATEG"]

    def _meta(row):
        m = meta_lookup.get(row["_K"], {})
        return (
            _norm_str(m.get("EMPLEADO", "")),
            _norm_str(m.get("SUPERIOR", "")),
            _norm_id(m.get("ID de la encuesta", "")),
        )

    def _add(row, tipo, nivel, causa, descripcion, valor_orig, extra_meta=None):
        empl, sup, id_enc = _meta(row)
        sup_final = _resolver_sup_con_fallback(sup, empl, row["ID INVOLVES"])
        meta = {"regla": causa, "id_pdv": row["ID INVOLVES"],
                "id_encuesta": id_enc, "categoria": row["CATEG"]}
        if extra_meta:
            meta.update(extra_meta)
        casos.append({
            "MODULO": "SOS",
            "TIPO":   tipo,
            "NIVEL":  nivel,
            "SUPERVISOR": sup_final,
            "MERCADERISTA_O_MARCA": empl or row["MARCA"],
            "CAUSA": causa,
            "DESCRIPCION": descripcion,
            "N_AFECTADOS": 1,
            "VALOR_ORIGINAL": valor_orig,
            "PDVS_AFECTADOS": f"{row['PDV']} — {row['MARCA']} [ID_ENCUESTA: {id_enc}]",
            "MES": mes, "ANIO": anio,
            "METADATA": meta,
        })

    # ── R5: marca > universo (CRÍTICO) ───────────────────────────────────
    m_r5 = (df["MARCA_CMS"] > df["UNIV"]) & df["UNIV"].notna() & (df["UNIV"] > 0)
    for _, r in df[m_r5].iterrows():
        _add(r, "PUNTUAL", "CRITICO",
             "cms reportado supera al universo del PDV",
             f"En {r['PDV']} se reportaron {r['MARCA_CMS']:.0f} cm de "
             f"{r['MARCA']} ({r['CATEG']}) pero el universo total es "
             f"{r['UNIV']:.0f} cm — físicamente imposible.",
             f"{r['MARCA_CMS']:.0f} cm (universo: {r['UNIV']:.0f} cm)")

    # ── R7: universo=0 con marca>0 (CRÍTICO) ─────────────────────────────
    m_r7 = (df["UNIV"].fillna(0) == 0) & (df["MARCA_CMS"].fillna(0) > 0)
    for _, r in df[m_r7].iterrows():
        _add(r, "PUNTUAL", "CRITICO",
             "Universo=0 con marca>0",
             f"En {r['PDV']} se reportó universo=0 cm para {r['CATEG']} pero "
             f"la marca {r['MARCA']} tiene {r['MARCA_CMS']:.0f} cm registrados.",
             f"{r['MARCA_CMS']:.0f} cm (universo: 0 cm)")

    # ── R9: participación > 100% (CRÍTICO) ──────────────────────────────
    # (cubre subconjunto de R5 — quitamos overlap para no duplicar)
    m_r9 = (df["PARTICIP"] > 1.0) & ~m_r5
    for _, r in df[m_r9].iterrows():
        _add(r, "PUNTUAL", "CRITICO",
             "Participación SOS > 100%",
             f"En {r['PDV']} la participación SOS de {r['MARCA']} dio "
             f"{r['PARTICIP']*100:.1f}% — superior al 100% sin que marca > universo.",
             f"{r['PARTICIP']*100:.1f}%")

    # ── R8: cms=0 con target>0 (REVISIÓN, PATRON por empleado×marca) ─────
    # Casos puntuales son ruido normal (la mayoría de marcas no están en la
    # mayoría de PDVs). Solo es patrón sospechoso cuando un mismo gestor
    # reporta cms=0 en muchos PDVs target de la misma marca.
    m_r8 = (df["MARCA_CMS"].fillna(0) == 0) & (df["TARGET"].fillna(0) > 0) & (df["UNIV"] > 0)
    df_r8 = df[m_r8].copy()
    if not df_r8.empty:
        # Necesitamos empleado por PDV — lo sacamos del lookup de la cruda
        df_r8["_K"] = df_r8["ID INVOLVES"] + "|" + df_r8["MARCA"] + "|" + df_r8["CATEG"]
        df_r8["EMPL"] = df_r8["_K"].map(lambda k: _norm_str(meta_lookup.get(k, {}).get("EMPLEADO", "")))
        df_r8["SUP"]  = df_r8["_K"].map(lambda k: _norm_str(meta_lookup.get(k, {}).get("SUPERIOR", "")))

        for (empl, marca), grupo in df_r8.groupby(["EMPL", "MARCA"]):
            if not empl or len(grupo) < SOS_R8_MIN_PDVS_PATRON:
                continue
            pdvs = grupo["PDV"].dropna().unique().tolist()
            categs = grupo["CATEG"].dropna().unique().tolist()
            sups = grupo["SUP"].dropna().unique().tolist()
            sup_resuelto = _resolver_supervisor(sups)
            if sup_resuelto == "SIN_SUPERVISOR":
                sup_resuelto = _resolver_sup_con_fallback("", empl)
            target_max = grupo["TARGET"].max() * 100
            casos.append({
                "MODULO": "SOS",
                "TIPO":   "PATRON",
                "NIVEL":  "REVISION",
                "SUPERVISOR": sup_resuelto,
                "MERCADERISTA_O_MARCA": empl,
                "CAUSA": f"Marca sin cms con target asignado ({marca})",
                "DESCRIPCION": (
                    f"{empl} reportó 0 cm de {marca} en {len(grupo)} PDVs target "
                    f"(target hasta {target_max:.0f}% en {', '.join(categs[:3])}). "
                    f"Probable desabastecimiento sistemático o falta de captura."
                ),
                "N_AFECTADOS": len(grupo),
                "VALOR_ORIGINAL": f"0 cm en {len(grupo)} PDVs (target hasta {target_max:.0f}%)",
                "PDVS_AFECTADOS": _pdvs_lista(pdvs),
                "MES": mes, "ANIO": anio,
                "METADATA": {"regla": "R8", "marca": marca,
                             "n_pdvs": int(len(grupo)),
                             "categorias": categs[:5]},
            })

    # ── R6: IQR Tukey k=3, marca > materialidad ──────────────────────────
    def _tukey(x: pd.Series) -> pd.Series:
        if len(x) < SOS_MIN_OBS_CATEG:
            return pd.Series(False, index=x.index)
        q1, q3 = x.quantile(0.25), x.quantile(0.75)
        return x > (q3 + SOS_TUKEY_K * (q3 - q1))

    excluidos = m_r5 | m_r7 | m_r9
    df_cand = df[~excluidos].copy()
    df_cand["_outlier"] = df_cand.groupby("CATEG")["MARCA_CMS"].transform(_tukey)
    m_r6 = df_cand["_outlier"] & (df_cand["MARCA_CMS"] > SOS_MATERIALIDAD_CMS)
    for _, r in df_cand[m_r6].iterrows():
        _add(r, "PUNTUAL", "REVISION",
             "Outlier estadístico (IQR Tukey k=3) en cms de marca",
             f"En {r['PDV']} se reportaron {r['MARCA_CMS']:.0f} cm de "
             f"{r['MARCA']} — valor atípico para la categoría {r['CATEG']} "
             f"(supera Q3 + 3·IQR).",
             f"{r['MARCA_CMS']:.0f} cm")

    return casos, auto


# ─────────────────────────────────────────────────────────────────────────────
# DETECTOR — PRECIOS  (3 reglas)
# ─────────────────────────────────────────────────────────────────────────────

def detectar_precios(mes: int, anio: int) -> tuple[list[dict], list[dict]]:
    """
    R10) precio_promo >= precio_regular           (REVISIÓN, PUNTUAL)
    R11) precio > 2× p99 por SKU+subcanal         (CRÍTICO, PUNTUAL)
    R12) presencia=1 con precio=0                 (REVISIÓN, PUNTUAL)
    """
    casos: list[dict] = []
    auto: list[dict] = []

    f_an = paths.PR_SALIDA / f"ANALISIS_PRECIOS_{_nombre_mes(mes)}_{anio}.xlsx"
    if not f_an.exists():
        candidatos = sorted(paths.PR_SALIDA.glob("ANALISIS_PRECIOS_*.xlsx"))
        if candidatos:
            f_an = candidatos[-1]
        else:
            return casos, auto

    df = pd.read_excel(f_an)
    df["ID del PDV"]      = df["ID del PDV"].map(_norm_id)
    df["PDV"]             = df["PDV"].map(_norm_str)
    df["EMPLEADO"]        = df["Empleado"].map(_norm_str)
    df["NOMBRE_PRODUCTO"] = df["NOMBRE_PRODUCTO"].map(_norm_str)
    df["MARCA"]           = df["Marca"].map(_norm_str)
    df["PRESENCIA"]       = pd.to_numeric(df["PRESENCIA"], errors="coerce").fillna(0)
    df["PRECIO_REGULAR"]  = pd.to_numeric(df["PRECIO_REGULAR"], errors="coerce").fillna(0)
    df["HAY_PROMO"]       = pd.to_numeric(df["HAY_PROMO"], errors="coerce").fillna(0)
    df["PRECIO_PROMO"]    = pd.to_numeric(df["PRECIO_PROMO"], errors="coerce").fillna(0)

    # Derivar subcanal: prefijo de letras del ACRONIMO en CIF
    pt = pd.read_excel(paths.CIF_OUT_FINAL,
                       usecols=["ID_PDV_INVOLVES", "ACRONIMO",
                                "SUPERVISOR_LIDER"])
    pt["ID_PDV_INVOLVES"]  = pt["ID_PDV_INVOLVES"].map(_norm_id)
    pt["SUPERVISOR_LIDER"] = pt["SUPERVISOR_LIDER"].map(_norm_str)
    pt["SUBCANAL"]         = pt["ACRONIMO"].astype(str).str.extract(r"^([A-Za-z]+)")
    pt_dedup = pt.drop_duplicates(subset=["ID_PDV_INVOLVES"])
    df = df.merge(
        pt_dedup[["ID_PDV_INVOLVES", "SUBCANAL", "SUPERVISOR_LIDER"]],
        left_on="ID del PDV", right_on="ID_PDV_INVOLVES", how="left",
    )

    # ── R10: promo >= regular ────────────────────────────────────────────
    m_r10 = (df["HAY_PROMO"] == 1) & (df["PRECIO_PROMO"] > 0) & (df["PRECIO_REGULAR"] > 0) \
            & (df["PRECIO_PROMO"] >= df["PRECIO_REGULAR"])
    for _, r in df[m_r10].iterrows():
        casos.append({
            "MODULO": "PRECIOS", "TIPO": "PUNTUAL", "NIVEL": "REVISION",
            "SUPERVISOR": _resolver_sup_con_fallback(r["SUPERVISOR_LIDER"], r.get("EMPLEADO", ""), r.get("ID del PDV", r.get("ID_PDV_INVOLVES", ""))),
            "MERCADERISTA_O_MARCA": r["EMPLEADO"] or r["MARCA"],
            "CAUSA": "Precio promo ≥ precio regular",
            "DESCRIPCION": (
                f"En {r['PDV']} el precio promo de {r['NOMBRE_PRODUCTO']} "
                f"(${r['PRECIO_PROMO']:,.0f}) iguala o supera al regular "
                f"(${r['PRECIO_REGULAR']:,.0f}). La promo deja de ser promo."
            ),
            "N_AFECTADOS": 1,
            "VALOR_ORIGINAL": f"Promo ${r['PRECIO_PROMO']:,.0f} / Regular ${r['PRECIO_REGULAR']:,.0f}",
            "PDVS_AFECTADOS": f"{r['PDV']} — {r['NOMBRE_PRODUCTO']}",
            "MES": mes, "ANIO": anio,
            "METADATA": {"regla": "R10", "id_pdv": r["ID del PDV"],
                         "sku": _norm_str(r.get("CODIGO_SKU", ""))},
        })

    # ── R11: precio > 2× p99 por SKU+SUBCANAL ────────────────────────────
    activos = df[(df["PRESENCIA"] == 1) & (df["PRECIO_REGULAR"] > 0)].copy()
    if not activos.empty:
        # p99 por SKU+SUBCANAL si n>=min, sino p99 por SKU global
        def _p99(g):
            if len(g) >= PRECIOS_MIN_OBS_SUBCANAL:
                return g.quantile(0.99)
            return np.nan

        p99_sk_sub = activos.groupby(["NOMBRE_PRODUCTO", "SUBCANAL"])["PRECIO_REGULAR"].transform(_p99)
        p99_sk_glob = activos.groupby("NOMBRE_PRODUCTO")["PRECIO_REGULAR"].transform("quantile", q=0.99)
        activos["P99"] = p99_sk_sub.fillna(p99_sk_glob)
        m_r11 = activos["PRECIO_REGULAR"] > (activos["P99"] * PRECIOS_FACTOR_P99)
        for _, r in activos[m_r11].iterrows():
            casos.append({
                "MODULO": "PRECIOS", "TIPO": "PUNTUAL", "NIVEL": "CRITICO",
                "SUPERVISOR": _resolver_sup_con_fallback(r["SUPERVISOR_LIDER"], r.get("EMPLEADO", ""), r.get("ID del PDV", r.get("ID_PDV_INVOLVES", ""))),
                "MERCADERISTA_O_MARCA": r["EMPLEADO"] or r["MARCA"],
                "CAUSA": "Precio anómalo (> 2× p99 del SKU)",
                "DESCRIPCION": (
                    f"En {r['PDV']} el precio regular de {r['NOMBRE_PRODUCTO']} "
                    f"(${r['PRECIO_REGULAR']:,.0f}) supera 2× el p99 del SKU "
                    f"en {r['SUBCANAL'] or 'mercado'} (p99=${r['P99']:,.0f}). "
                    f"Probable error de digitación."
                ),
                "N_AFECTADOS": 1,
                "VALOR_ORIGINAL": f"${r['PRECIO_REGULAR']:,.0f} (p99: ${r['P99']:,.0f})",
                "PDVS_AFECTADOS": f"{r['PDV']} — {r['NOMBRE_PRODUCTO']}",
                "MES": mes, "ANIO": anio,
                "METADATA": {"regla": "R11", "id_pdv": r["ID del PDV"],
                             "sku": _norm_str(r.get("CODIGO_SKU", "")),
                             "p99": float(r["P99"]) if not pd.isna(r["P99"]) else None},
            })

    # ── R12: presencia=1 con precio=0 ────────────────────────────────────
    m_r12 = (df["PRESENCIA"] == 1) & (df["PRECIO_REGULAR"] == 0)
    for _, r in df[m_r12].iterrows():
        casos.append({
            "MODULO": "PRECIOS", "TIPO": "PUNTUAL", "NIVEL": "REVISION",
            "SUPERVISOR": _resolver_sup_con_fallback(r["SUPERVISOR_LIDER"], r.get("EMPLEADO", ""), r.get("ID del PDV", r.get("ID_PDV_INVOLVES", ""))),
            "MERCADERISTA_O_MARCA": r["EMPLEADO"] or r["MARCA"],
            "CAUSA": "Presencia=1 con precio=0",
            "DESCRIPCION": (
                f"En {r['PDV']} se marca presencia de {r['NOMBRE_PRODUCTO']} "
                f"pero el precio quedó en $0. Falta captura del valor."
            ),
            "N_AFECTADOS": 1,
            "VALOR_ORIGINAL": f"Presencia=1, Precio=$0",
            "PDVS_AFECTADOS": f"{r['PDV']} — {r['NOMBRE_PRODUCTO']}",
            "MES": mes, "ANIO": anio,
            "METADATA": {"regla": "R12", "id_pdv": r["ID del PDV"],
                         "sku": _norm_str(r.get("CODIGO_SKU", ""))},
        })

    return casos, auto


# ─────────────────────────────────────────────────────────────────────────────
# DETECTOR — EXHIBICIONES PAGADAS  (2 reglas)
# ─────────────────────────────────────────────────────────────────────────────

def detectar_exh_pagadas(mes: int, anio: int) -> tuple[list[dict], list[dict]]:
    """
    R13) cantidad_ejecutada > 1.5 × cantidad_negociada    (REVISIÓN, PUNTUAL)
    R14) implementada='No' con cantidad_ejecutada > 0     (CRÍTICO, PUNTUAL)
    """
    casos: list[dict] = []
    auto: list[dict] = []

    candidatos = sorted(paths.EXHIB_SALIDA.glob("*pagadas*agrupado*.xlsx"))
    if not candidatos:
        return casos, auto
    df = pd.read_excel(candidatos[-1])

    # Normalizar
    df["ID_PDV_INVOLVES"] = df["ID_PDV_INVOLVES"].map(_norm_id)
    df["TIPO_FINAL"]      = df["TIPO_FINAL"].map(_norm_str)
    df["MARCA_FINAL"]     = df["MARCA_FINAL"].map(_norm_str)
    df["IMPLEMENTADA"]    = df["La Exhibicion esta implementada de acuerdo con el planning?"].map(_norm_str)
    df["CANTIDAD_PLANEADA"]  = pd.to_numeric(df["CANTIDAD_PLANEADA"], errors="coerce").fillna(0)
    df["CANTIDAD_EJECUTADA"] = pd.to_numeric(df["CANTIDAD_EJECUTADA"], errors="coerce").fillna(0)

    # Cruce supervisor + PDV
    pt = pd.read_excel(paths.CIF_OUT_FINAL,
                       usecols=["ID_PDV_INVOLVES", "NOMBRE_PDV",
                                "SUPERVISOR_LIDER"])
    pt["ID_PDV_INVOLVES"]  = pt["ID_PDV_INVOLVES"].map(_norm_id)
    pt["NOMBRE_PDV"]       = pt["NOMBRE_PDV"].map(_norm_str)
    pt["SUPERVISOR_LIDER"] = pt["SUPERVISOR_LIDER"].map(_norm_str)
    pt_dedup = pt.drop_duplicates(subset=["ID_PDV_INVOLVES"])
    df = df.merge(pt_dedup, on="ID_PDV_INVOLVES", how="left")

    # ── R13: sobre-ejecución ─────────────────────────────────────────────
    m_r13 = (df["CANTIDAD_PLANEADA"] > 0) & \
            (df["CANTIDAD_EJECUTADA"] > df["CANTIDAD_PLANEADA"] * EXH_FACTOR_SOBREEJECUCION)
    for _, r in df[m_r13].iterrows():
        ratio = r["CANTIDAD_EJECUTADA"] / r["CANTIDAD_PLANEADA"]
        casos.append({
            "MODULO": "EXHIBICIONES_PAGADAS", "TIPO": "PUNTUAL", "NIVEL": "REVISION",
            "SUPERVISOR": _resolver_sup_con_fallback(r["SUPERVISOR_LIDER"], r.get("EMPLEADO", ""), r.get("ID del PDV", r.get("ID_PDV_INVOLVES", ""))),
            "MERCADERISTA_O_MARCA": r["MARCA_FINAL"] or "SIN_MARCA",
            "CAUSA": "Cantidad ejecutada > 1.5× la negociada",
            "DESCRIPCION": (
                f"En {r['NOMBRE_PDV']} se ejecutaron {int(r['CANTIDAD_EJECUTADA'])} "
                f"unidades de exhibición {r['TIPO_FINAL']} cuando la negociación fue "
                f"de {int(r['CANTIDAD_PLANEADA'])} (factor {ratio:.1f}×). "
                f"Verificar si fue un acuerdo no documentado."
            ),
            "N_AFECTADOS": 1,
            "VALOR_ORIGINAL": f"{int(r['CANTIDAD_EJECUTADA'])} ejecutadas / "
                              f"{int(r['CANTIDAD_PLANEADA'])} negociadas",
            "PDVS_AFECTADOS": f"{r['NOMBRE_PDV']} — {r['TIPO_FINAL']} {r['MARCA_FINAL']}",
            "MES": mes, "ANIO": anio,
            "METADATA": {"regla": "R13", "id_pdv": r["ID_PDV_INVOLVES"],
                         "ratio": float(ratio)},
        })

    # ── R14: implementada=No con cantidad > 0 ────────────────────────────
    m_r14 = (df["IMPLEMENTADA"] == "NO") & (df["CANTIDAD_EJECUTADA"] > 0)
    for _, r in df[m_r14].iterrows():
        casos.append({
            "MODULO": "EXHIBICIONES_PAGADAS", "TIPO": "PUNTUAL", "NIVEL": "CRITICO",
            "SUPERVISOR": _resolver_sup_con_fallback(r["SUPERVISOR_LIDER"], r.get("EMPLEADO", ""), r.get("ID del PDV", r.get("ID_PDV_INVOLVES", ""))),
            "MERCADERISTA_O_MARCA": r["MARCA_FINAL"] or "SIN_MARCA",
            "CAUSA": "Implementada=No con cantidad capturada > 0",
            "DESCRIPCION": (
                f"En {r['NOMBRE_PDV']} se reporta exhibición {r['TIPO_FINAL']} "
                f"no implementada pero con {int(r['CANTIDAD_EJECUTADA'])} unidades "
                f"capturadas. Inconsistencia en el registro."
            ),
            "N_AFECTADOS": 1,
            "VALOR_ORIGINAL": f"Implementada=No, cantidad={int(r['CANTIDAD_EJECUTADA'])}",
            "PDVS_AFECTADOS": f"{r['NOMBRE_PDV']} — {r['TIPO_FINAL']} {r['MARCA_FINAL']}",
            "MES": mes, "ANIO": anio,
            "METADATA": {"regla": "R14", "id_pdv": r["ID_PDV_INVOLVES"]},
        })

    return casos, auto


# ─────────────────────────────────────────────────────────────────────────────
# HELPERS ADICIONALES
# ─────────────────────────────────────────────────────────────────────────────

_MESES_ES = {
    1:"ENERO", 2:"FEBRERO", 3:"MARZO", 4:"ABRIL", 5:"MAYO", 6:"JUNIO",
    7:"JULIO", 8:"AGOSTO", 9:"SEPTIEMBRE", 10:"OCTUBRE", 11:"NOVIEMBRE", 12:"DICIEMBRE",
}


def _nombre_mes(m: int) -> str:
    return _MESES_ES.get(int(m), "ABRIL")


# ─────────────────────────────────────────────────────────────────────────────
# ESCRITURA — Excel con tablas nombradas (TblCasos + TblHistorial)
# ─────────────────────────────────────────────────────────────────────────────

def _asignar_ids(casos: list[dict]) -> list[dict]:
    """
    Numera los casos respetando agrupación por módulo. Genera ID_CASO único.
    PATRON usa sufijo -S- en CIF y nada en NP (matchea histórico).
    """
    contadores: dict[str, int] = {}
    for c in casos:
        modulo = c["MODULO"]
        prefijo = PREFIJO_MODULO.get(modulo, "OTHER")
        tipo    = c["TIPO"]
        key     = f"{prefijo}-{tipo}"
        contadores[key] = contadores.get(key, 0) + 1
        c["ID_CASO"] = _gen_id(prefijo, tipo, contadores[key])
        # Defaults columnas que llena la app
        c.setdefault("ESTADO",         "PENDIENTE")
        c.setdefault("DECISION",       "")
        c.setdefault("VALOR_CORRECTO", "")
        c.setdefault("OBSERVACION",    "")
        # METADATA → JSON string
        meta = c.pop("METADATA", {})
        c["METADATA"] = json.dumps(meta, ensure_ascii=False) if meta else ""
    return casos


def _cargar_historial_acumulado() -> pd.DataFrame:
    """
    Lee Historial_Correcciones.xlsx si existe (acumulado entre meses).
    Sin esto, devuelve frame vacío con las columnas correctas.
    """
    f = DIR_OUT / "Historial_Correcciones.xlsx"
    if f.exists():
        try:
            return pd.read_excel(f)
        except Exception:
            pass
    return pd.DataFrame(columns=COLS_HISTORIAL)


def escribir_excel(casos: list[dict], mes: int, anio: int) -> Path:
    """
    Escribe Casos_Revision_<MM>_<YYYY>.xlsx con 2 hojas:
      • Casos     → tabla nombrada TblCasos
      • Historial → tabla nombrada TblHistorial (acumulado)

    Las "tablas nombradas" son requisito para que la app/PowerApps las
    consuma; openpyxl las soporta vía openpyxl.worksheet.table.Table.
    """
    from openpyxl import Workbook
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.utils import get_column_letter

    df_casos = pd.DataFrame(casos, columns=COLS_CASO)
    df_hist  = _cargar_historial_acumulado()

    ruta = DIR_OUT / f"Casos_Revision_{mes:02d}_{anio}.xlsx"
    wb = Workbook()

    # Hoja Casos
    ws_c = wb.active
    ws_c.title = "Casos"
    ws_c.append(list(df_casos.columns))
    for _, row in df_casos.iterrows():
        ws_c.append([row[c] if not pd.isna(row[c]) else "" for c in df_casos.columns])
    # Tabla nombrada
    if len(df_casos) > 0:
        rango = f"A1:{get_column_letter(len(df_casos.columns))}{len(df_casos)+1}"
        t_casos = Table(displayName="TblCasos", ref=rango)
        t_casos.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False,
            showRowStripes=True, showColumnStripes=False)
        ws_c.add_table(t_casos)
    # Anchos
    anchos = {"A":14, "B":30, "C":35, "D":22, "E":10, "F":12, "G":40,
              "H":60, "I":12, "J":28, "K":50, "L":6, "M":7, "N":12,
              "O":18, "P":22, "Q":40, "R":40}
    for col, w in anchos.items():
        ws_c.column_dimensions[col].width = w

    # Hoja Historial
    ws_h = wb.create_sheet("Historial")
    ws_h.append(COLS_HISTORIAL)
    for _, row in df_hist.iterrows():
        ws_h.append([row.get(c, "") for c in COLS_HISTORIAL])
    if len(df_hist) > 0:
        rango_h = f"A1:{get_column_letter(len(COLS_HISTORIAL))}{len(df_hist)+1}"
    else:
        # Tabla vacía: solo header + 1 fila vacía para que sea tabla válida
        ws_h.append([""] * len(COLS_HISTORIAL))
        rango_h = f"A1:{get_column_letter(len(COLS_HISTORIAL))}2"
    t_hist = Table(displayName="TblHistorial", ref=rango_h)
    t_hist.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False)
    ws_h.add_table(t_hist)

    wb.save(ruta)
    return ruta


def escribir_auto_resueltos(auto: list[dict], mes: int, anio: int) -> Path:
    ruta = DIR_OUT / f"auto_resueltos_{mes:02d}_{anio}.csv"
    if not auto:
        pd.DataFrame(columns=["REGLA", "MODULO"]).to_csv(ruta, index=False)
    else:
        pd.DataFrame(auto).to_csv(ruta, index=False)
    return ruta


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Detector de anomalías — Eficacia")
    parser.add_argument("--mes",  type=int, default=None, help="Mes activo (1-12). Default: detección automática.")
    parser.add_argument("--anio", type=int, default=None, help="Año. Default: detección automática.")
    parser.add_argument("--solo", nargs="+",
                        choices=["cif", "np", "sos", "precios", "exh"],
                        help="Correr solo algunos módulos (default: todos)")
    args = parser.parse_args()

    mes, anio = _periodo_default()
    if args.mes:  mes  = args.mes
    if args.anio: anio = args.anio
    modulos = args.solo or ["cif", "np", "sos", "precios", "exh"]

    print("═" * 70)
    print(f"  DETECTOR DE ANOMALÍAS — período {mes:02d}/{anio}")
    print("═" * 70)
    print(f"  Output: {DIR_OUT}")

    detectores = {
        "cif":     ("CIF",                  detectar_cif),
        "np":      ("NO PRESENCIA",         detectar_np),
        "sos":     ("SOS",                  detectar_sos),
        "precios": ("PRECIOS",              detectar_precios),
        "exh":     ("EXHIBICIONES PAGADAS", detectar_exh_pagadas),
    }

    casos_total: list[dict] = []
    auto_total:  list[dict] = []

    for key in modulos:
        nombre, fn = detectores[key]
        print(f"\n  ▶ Detectando {nombre}...")
        try:
            casos, auto = fn(mes, anio)
            print(f"     {len(casos)} caso(s) | {len(auto)} auto-resuelto(s)")
            casos_total.extend(casos)
            auto_total.extend(auto)
        except Exception as e:
            print(f"     ⚠️  Error: {e}")

    # Asignar IDs estables
    casos_total = _asignar_ids(casos_total)

    # Resumen
    print("\n" + "─" * 70)
    print("  RESUMEN")
    print("─" * 70)
    if casos_total:
        df_resumen = pd.DataFrame([
            {"MODULO": c["MODULO"], "TIPO": c["TIPO"], "NIVEL": c["NIVEL"]}
            for c in casos_total
        ])
        print(df_resumen.groupby(["MODULO", "TIPO", "NIVEL"]).size().to_string())

    print(f"\n  Total casos:           {len(casos_total)}")
    print(f"  Total auto-resueltos:  {len(auto_total)}")

    # Escritura
    ruta_xlsx = escribir_excel(casos_total, mes, anio)
    ruta_csv  = escribir_auto_resueltos(auto_total, mes, anio)
    print(f"\n  ✅ Casos:          {ruta_xlsx}")
    print(f"  ✅ Auto-resueltos: {ruta_csv}")
    print("═" * 70)

    return {"casos": casos_total, "auto": auto_total,
            "mes": mes, "anio": anio,
            "xlsx": ruta_xlsx, "csv": ruta_csv}


if __name__ == "__main__":
    main()
